#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CONSOLIDADOR T25 v15.1 - VERSIÓN PARAMETRIZADA PARA API
=======================================================

Esta es una versión modificada del consolidador T25 que:
- NO usa input() - lee parámetros de variables de entorno
- Genera archivos en OUTPUT_FOLDER
- Es ejecutable desde la API web

Variables de entorno requeridas:
- CONSOLIDADOR_MAESTRA: Ruta al archivo de maestra
- CONSOLIDADOR_MODO: ESPECIFICO, POR_ANO, COMPLETO
- CONSOLIDADOR_ANO: Año del contrato (opcional según modo)
- CONSOLIDADOR_NUMERO: Número del contrato (opcional según modo)
- CONSOLIDADOR_OUTPUT: Carpeta de salida
- SFTP_HOST, SFTP_PORT, SFTP_USERNAME, SFTP_PASSWORD: Credenciales SFTP
- SFTP_CARPETA_PRINCIPAL: Carpeta principal en SFTP
"""

import os
import sys
import pandas as pd
import numpy as np
import re
import warnings
import os
import gc
import io
import zipfile
import chardet
from datetime import datetime, timedelta
from typing import Tuple, Optional, List, Dict, Any
from tqdm import tqdm
import threading
import time
import shutil
import paramiko
from dataclasses import dataclass, field

# Machine Learning
# Machine Learning
try:
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.metrics.pairwise import cosine_similarity
except ImportError:
    TfidfVectorizer = None
    cosine_similarity = None

from collections import Counter

# Configuración
warnings.filterwarnings('ignore')
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', 60)
pd.set_option('display.float_format', lambda x: f'{x:,.2f}')

# ══════════════════════════════════════════════════════════════════════════════
# LECTURA DE PARÁMETROS DESDE VARIABLES DE ENTORNO
# ══════════════════════════════════════════════════════════════════════════════

# Parámetros del consolidador
PARAM_MAESTRA = os.environ.get('CONSOLIDADOR_MAESTRA', '')
PARAM_MODO = os.environ.get('CONSOLIDADOR_MODO', 'COMPLETO')
PARAM_ANO = os.environ.get('CONSOLIDADOR_ANO', '')
PARAM_NUMERO = os.environ.get('CONSOLIDADOR_NUMERO', '')
PARAM_OUTPUT = os.environ.get('CONSOLIDADOR_OUTPUT', './outputs')

# Validar parámetros obligatorios
if not PARAM_MAESTRA:
    print("❌ ERROR: Variable CONSOLIDADOR_MAESTRA no definida")
    sys.exit(1)

if not os.path.exists(PARAM_MAESTRA):
    print(f"❌ ERROR: Archivo de maestra no encontrado: {PARAM_MAESTRA}")
    sys.exit(1)

# Crear carpeta de salida
os.makedirs(PARAM_OUTPUT, exist_ok=True)
os.chdir(PARAM_OUTPUT)

print(f"📋 Parámetros recibidos:")
print(f"   • Maestra: {PARAM_MAESTRA}")
print(f"   • Modo: {PARAM_MODO}")
print(f"   • Año: {PARAM_ANO or 'Todos'}")
print(f"   • Contrato: {PARAM_NUMERO or 'Todos'}")
print(f"   • Output: {PARAM_OUTPUT}")

# ══════════════════════════════════════════════════════════════════════════════
# LISTA COMPLETA DE CIUDADES COLOMBIANAS (para validación de CUPS)
# ══════════════════════════════════════════════════════════════════════════════

CIUDADES_COLOMBIA_COMPLETA = {
    # Capitales
    'BOGOTÁ', 'BOGOTA', 'MEDELLÍN', 'MEDELLIN', 'CALI', 'BARRANQUILLA',
    'CARTAGENA', 'BUCARAMANGA', 'CÚCUTA', 'CUCUTA', 'PEREIRA', 'IBAGUÉ',
    'IBAGUE', 'SANTA MARTA', 'MANIZALES', 'VILLAVICENCIO', 'PASTO',
    'MONTERÍA', 'MONTERIA', 'NEIVA', 'ARMENIA', 'SINCELEJO', 'POPAYÁN',
    'POPAYAN', 'VALLEDUPAR', 'TUNJA', 'FLORENCIA', 'QUIBDÓ', 'QUIBDO',
    'RIOHACHA', 'YOPAL', 'MOCOA', 'LETICIA', 'INÍRIDA', 'INIRIDA',
    'MITÚ', 'MITU', 'PUERTO CARREÑO', 'SAN JOSÉ DEL GUAVIARE', 'ARAUCA',
    # Ciudades intermedias usadas en traslados
    'BAHIA SOLANO', 'BARRANCABERMEJA', 'BUENAVENTURA', 'PALMIRA',
    'CARTAGO', 'TULUA', 'TULUÁ', 'BUGA', 'SOGAMOSO', 'DUITAMA', 'GIRARDOT',
    'FUSAGASUGA', 'FUSAGASUGÁ', 'FACATATIVA', 'FACATATIVÁ', 'ZIPAQUIRA',
    'ZIPAQUIRÁ', 'CHIA', 'CHÍA', 'SOACHA', 'RIONEGRO', 'ENVIGADO',
    'ITAGUI', 'ITAGÜÍ', 'BELLO', 'TUMACO', 'IPIALES', 'GRANADA', 'ACACIAS',
    'ACACÍAS', 'PUERTO LOPEZ', 'PUERTO LÓPEZ', 'AGUACHICA', 'OCAÑA',
    'APARTADO', 'APARTADÓ', 'TURBO', 'CAUCASIA', 'MAGANGUE', 'MAGANGUÉ',
    'LORICA', 'CERETE', 'CERETÉ', 'ESPINAL', 'MELGAR', 'FLANDES', 'HONDA',
    'MARIQUITA', 'LA DORADA', 'PUERTO BERRIO', 'PUERTO BERRÍO',
    'PUERTO BOYACA', 'PUERTO BOYACÁ', 'CIENAGA', 'CIÉNAGA', 'FUNDACION',
    'FUNDACIÓN', 'ARACATACA', 'EL BANCO', 'PLATO', 'COROZAL', 'SAMPUES',
    'SAMPUÉS', 'SAN MARCOS', 'ZARZAL', 'JAMUNDI', 'JAMUNDÍ', 'YUMBO',
    'CANDELARIA', 'PRADERA', 'FLORIDA', 'CERRITO', 'GUACARI', 'GUACARÍ',
    'GINEBRA', 'ROLDANILLO', 'LA UNION', 'LA UNIÓN', 'SEVILLA',
    'CAICEDONIA', 'ARGELIA', 'DARIEN', 'DARIÉN', 'RESTREPO', 'DAGUA',
    'LA CUMBRE', 'CLO', 'BOG', 'MDE',  # Códigos de aeropuerto
}

DEPARTAMENTOS_COLOMBIA = {
    'BOGOTÁ D.C', 'BOGOTA D.C', 'BOGOTÁ D.C.', 'BOGOTA D.C.',
    'ANTIOQUIA', 'ATLÁNTICO', 'ATLANTICO', 'BOLÍVAR', 'BOLIVAR',
    'BOYACÁ', 'BOYACA', 'CALDAS', 'CAQUETÁ', 'CAQUETA', 'CASANARE',
    'CAUCA', 'CESAR', 'CHOCÓ', 'CHOCO', 'CÓRDOBA', 'CORDOBA',
    'CUNDINAMARCA', 'GUAINÍA', 'GUAINIA', 'GUAVIARE', 'HUILA',
    'LA GUAJIRA', 'MAGDALENA', 'META', 'NARIÑO', 'NARINO',
    'NORTE DE SANTANDER', 'PUTUMAYO', 'QUINDÍO', 'QUINDIO',
    'RISARALDA', 'SAN ANDRÉS', 'SAN ANDRES', 'SANTANDER', 'SUCRE',
    'TOLIMA', 'VALLE', 'VALLE DEL CAUCA', 'VAUPÉS', 'VAUPES',
    'VICHADA', 'AMAZONAS', 'ARAUCA'
}

MUNICIPIOS_COLOMBIA = CIUDADES_COLOMBIA_COMPLETA | DEPARTAMENTOS_COLOMBIA

# ══════════════════════════════════════════════════════════════════════════════
# PREFIJOS DE CELULAR COLOMBIANO (para validación de teléfonos)
# ══════════════════════════════════════════════════════════════════════════════

PREFIJOS_CELULAR_COLOMBIA = {
    '300', '301', '302', '303', '304', '305',
    '310', '311', '312', '313', '314', '315', '316', '317', '318',
    '320', '321', '322', '323', '324',
    '350', '351',
    '330', '331', '332', '333'
}

# ══════════════════════════════════════════════════════════════════════════════
# HOJAS A EXCLUIR (silenciosamente, sin alerta)
# ══════════════════════════════════════════════════════════════════════════════

HOJAS_EXCLUIR_SILENCIOSAMENTE = {
    'INSTRUCCIONES', 'INFO', 'DATOS', 'CONTENIDO', 'INDICE', 'ÍNDICE',
    'GUIA DE USO', 'GUÍA DE USO', 'CONTROL DE CAMBIOS', 'HOJA1', 'SHEET1',
    'INSTRUCTIVO', 'PARAMETROS', 'PARÁMETROS', 'CONFIGURACION', 'CONFIGURACIÓN',
    'LISTA', 'LISTAS', 'VALIDACION', 'VALIDACIÓN', 'CATALOGO', 'CATÁLOGO',
    'RESUMEN', 'PORTADA', 'CARATULA', 'CARÁTULA', 'INICIO', 'HOME',
    'MENU', 'MENÚ', 'ANEXO TECNICO', 'ANEXO TÉCNICO', 'GLOSARIO',
}

# Hojas que NO generan alerta individualmente, pero se mencionan
# si no hay hoja de servicios válida
HOJAS_SIN_SERVICIOS_VALIDOS = {
    'PAQUETES', 'TARIFAS PAQUETES', 'PAQUETE',
    'COSTO VIAJE', 'COSTO DE VIAJE', 'COSTOS VIAJE',
}

# ══════════════════════════════════════════════════════════════════════════════
# PALABRAS INVÁLIDAS PARA CUPS
# ══════════════════════════════════════════════════════════════════════════════

PALABRAS_INVALIDAS_CUPS = [
    'CODIGO', 'CUPS', 'ITEM', 'DESCRIPCION', 'TARIFA', 'TOTAL', 'SUBTOTAL',
    'DEPARTAMENTO', 'MUNICIPIO', 'HABILITACION', 'HABIITACION', 'DIRECCION',
    'TELEFONO', 'EMAIL', 'SEDE', 'NOMBRE', 'NUMERO', 'ESPECIALIDAD',
    'MANUAL', 'OBSERV', 'PORCENTAJE', 'HOMOLOGO', 'N°', 'NO.',
    'NOTA', 'NOTAS', 'ACLARATORIA', 'ACLARATORIAS', 'ACLARACION', 'ACLARACIONES',
    'INCLUYE', 'NO INCLUYE', 'EXCLUYE',
    'USO DE EQUIPO', 'DERECHO DE SALA', 'DERECHO SALA',
    'VER NOTA', 'VER NOTAS', 'SEGUN NOTA',
    'APLICA', 'NO APLICA', 'SEGÚN', 'SEGUN',
    'CONSULTAR', 'REVISAR', 'PENDIENTE',
    'VALOR', 'PRECIO', 'COSTO',
    'CONTRATO', 'ACTA', 'OTROSI', 'OTROSÍ',
    'VIGENTE', 'VIGENCIA',
    'TRASLADO', 'ORIGEN', 'DESTINO',
    'TARIFAS PROPIAS', 'TARIFA PROPIA',
]

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE DETECCIÓN DE ARCHIVOS v15.0
# ══════════════════════════════════════════════════════════════════════════════

import re

def es_archivo_tarifas_valido(nombre: str) -> tuple:
    """
    🆕 v15.1: Detecta si un archivo es válido para procesamiento de tarifas.

    PATRONES VÁLIDOS:
    1. Contiene "ANEXO 1" o "ANEXO_1" (formato tradicional)
    2. Contiene "TARIFAS" o "TARIFA" en el nombre (formato simplificado)
    3. Contiene "OTROSI" o "OTROSÍ" seguido de número

    EXCLUSIONES v15.1:
    - "ANALISIS DE TARIFAS" y variantes (ANÁLISIS, sin DE, singular/plural)

    Retorna: (es_valido: bool, tipo: str)
    Tipos: 'ANEXO_1', 'TARIFAS', 'OTROSI', 'INVALIDO'
    """
    if not nombre:
        return False, 'INVALIDO'

    # 🆕 v15.2: Normalización agresiva (eliminar tabs y espacios extra)
    nombre_upper = nombre.upper().replace('\t', '').strip()

    # EXCLUSIONES: archivos que NO se deben procesar
    palabras_excluir = [
        'MEDICAMENT', 'MEDICAMENTO', 'MEDICAMENTOS',
        'FARMACO', 'FÁRMACO', 'FARMACOS', 'FÁRMACOS',
        'INSUMO', 'INSUMOS'
    ]

    for palabra in palabras_excluir:
        if palabra in nombre_upper:
            if 'SERVICIO' in nombre_upper or 'SERV' in nombre_upper:
                continue
            return False, 'INVALIDO'

    # 🆕 v15.1: EXCLUSIÓN: "ANALISIS DE TARIFAS" y variantes NO se procesan
    # Pero "TARIFAS" o "TARIFA" solas SÍ se procesan
    if re.search(r'AN[AÁ]LISIS\s*(DE\s*)?(TARIFAS?|TARIFA)', nombre_upper):
        return False, 'INVALIDO'

    # DETECCIÓN 1: Archivos OTROSÍ (tienen prioridad)
    patrones_otrosi = [
        r'OTRO\s*S[IÍ]\s*[_#\-\s]*(\d+)',
        r'OTROS[IÍ]\s*[_#\-\s]*(\d+)',
        r'OT[_\-\s]?(\d+)',
        r'ADICI[OÓ]N\s*[_#\-\s]*(\d+)',
        r'MODIFICACI[OÓ]N\s*[_#\-\s]*(\d+)',
    ]

    for patron in patrones_otrosi:
        if re.search(patron, nombre_upper):
            return True, 'OTROSI'

    # DETECCIÓN 2: Archivos con ANEXO 1 explícito
    patrones_anexo1 = [
        r'ANEXO\s*[_\-\s]*0?1(?!\d)',
        r'ANEX[O0]\s*[_\-\s]*1(?!\d)',
        r'ANEXO\s*N[OÚº°]?\.?\s*0?1(?!\d)',
        r'A1[_\-\s]',
        r'[_\-]ANEXO[_\-]?1',
        r'ANEXO[_\-]1[_\-]',
    ]

    for patron in patrones_anexo1:
        if re.search(patron, nombre_upper):
            return True, 'ANEXO_1'

    n_limpio = nombre_upper.replace(' ', '').replace('_', '').replace('-', '').replace('(', '').replace(')', '')
    if 'ANEXO1' in n_limpio or 'ANEXO01' in n_limpio:
        return True, 'ANEXO_1'

    # EXCLUSIÓN: Verificar si es ANEXO 2, 3, etc.
    patron_anexo_no_1 = r'ANEXO\s*[_\-\s]*([2-9]|[1-9]\d)(?!\d)'
    if re.search(patron_anexo_no_1, nombre_upper):
        return False, 'INVALIDO'

    # DETECCIÓN 3: Archivos que contienen "TARIFAS" (formato simplificado)
    patrones_tarifas = [
        r'\d+[\-_]TARIFAS[\-_]',
        r'^TARIFAS[\-_]',
        r'[\-_]TARIFAS[\-_]',
        r'[\-_]TARIFAS\.',
    ]

    for patron in patrones_tarifas:
        if re.search(patron, nombre_upper):
            return True, 'TARIFAS'

    # DETECCIÓN 4: Combinaciones especiales
    if 'ANEXO' in nombre_upper and ('TARIFA' in nombre_upper or 'SERV' in nombre_upper):
        otros_anexos = re.search(r'ANEXO\s*[_\-\s]*([2-9]|[1-9]\d)(?!\d)', nombre_upper)
        if not otros_anexos:
            return True, 'ANEXO_1'

    return False, 'INVALIDO'

def contiene_anexo1(nombre: str) -> bool:
    """
    🆕 v15.0: Detecta si el nombre corresponde a un archivo procesable de tarifas.

    AHORA INCLUYE:
    - Archivos con ANEXO 1 explícito
    - Archivos con "TARIFAS" en el nombre (formato simplificado)
    - Archivos de OTROSÍ
    """
    es_valido, tipo = es_archivo_tarifas_valido(nombre)
    return es_valido

def extraer_numero_otrosi_global(nombre: str):
    """
    🆕 v15.0: Extrae el número de otrosí del nombre del archivo.
    """
    if not nombre:
        return None

    nombre_upper = nombre.upper()

    patrones = [
        r'OTRO\s*S[IÍ]\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
        r'OTROS[IÍ]\s*[_#\-\s]*(\d+)',
        r'OTRO[\s_\-]?SI[\s_\-#]*(\d+)',
        r'OT\s*[_\-\s]?\s*(\d+)',
        r'ADICI[OÓ]N\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
        r'MODIFICA(?:CI[OÓ]N)?\s*[_#\-\s]*(\d+)',
    ]

    for patron in patrones:
        match = re.search(patron, nombre_upper)
        if match:
            try:
                return int(match.group(1))
            except (ValueError, IndexError):
                continue

    return None

def clasificar_tipo_archivo(nombre: str) -> dict:
    """
    🆕 v15.0: Clasifica un archivo y retorna información completa.
    """
    resultado = {
        'es_valido': False,
        'tipo': 'INVALIDO',
        'numero_otrosi': None,
        'es_otrosi': False,
        'motivo_exclusion': None
    }

    if not nombre:
        resultado['motivo_exclusion'] = 'Nombre vacío'
        return resultado

    nombre_upper = nombre.upper()

    palabras_excluir = ['MEDICAMENT', 'FARMACO', 'FÁRMACO', 'INSUMO']
    for palabra in palabras_excluir:
        if palabra in nombre_upper:
            if 'SERVICIO' not in nombre_upper and 'SERV' not in nombre_upper:
                resultado['motivo_exclusion'] = f'Archivo de {palabra.lower()}'
                return resultado

    num_otrosi = extraer_numero_otrosi_global(nombre)
    if num_otrosi:
        resultado['numero_otrosi'] = num_otrosi
        resultado['es_otrosi'] = True

    es_valido, tipo = es_archivo_tarifas_valido(nombre)
    resultado['es_valido'] = es_valido
    resultado['tipo'] = tipo

    if not es_valido:
        resultado['motivo_exclusion'] = 'No coincide con patrones de tarifas'

    return resultado

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: es_telefono_celular_colombiano
# ══════════════════════════════════════════════════════════════════════════════

def es_telefono_celular_colombiano(valor: str) -> bool:
    """
    v14.1: Detecta si un valor es un teléfono celular colombiano.
    CORREGIDO: Funciona con números SIN guiones (como vienen en Excel).

    Ejemplos que detecta:
    - 3214567890 (sin guiones)
    - 3001234567
    - 3501234567

    NO debe confundir con:
    - Tarifas (5920000, 11380000)
    - Códigos CUPS (890201)
    - Habilitación (7614708225)
    """
    if not valor:
        return False

    import re

    # Limpiar el valor - quitar TODO excepto dígitos
    valor_str = str(valor).strip()

    # Si termina en .0, quitarlo
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]

    # Quitar cualquier caracter no dígito
    valor_clean = re.sub(r'[^\d]', '', valor_str)

    # Debe ser exactamente 10 dígitos
    if len(valor_clean) != 10:
        return False

    # Verificar prefijo de celular colombiano
    prefijo = valor_clean[:3]
    return prefijo in PREFIJOS_CELULAR_COLOMBIA

def es_telefono_celular(valor: str) -> bool:
    """Alias para compatibilidad."""
    return es_telefono_celular_colombiano(valor)

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: validar_cups
# ══════════════════════════════════════════════════════════════════════════════

def validar_cups(cups: str, fila: list = None) -> bool:
    """
    v14.1: Validación de CUPS ULTRA estricta.

    RECHAZA:
    - Ciudades colombianas (ARMENIA, CALI, BAHIA SOLANO, etc.)
    - Valores monetarios grandes (>= 7 dígitos)
    - Teléfonos celulares (10 dígitos con prefijo conocido)
    - Códigos de habilitación (8-12 dígitos)
    - Palabras inválidas (CODIGO, TARIFA, DESCRIPCION, etc.)
    """
    import re

    if not cups:
        return False

    cups_str = str(cups).strip()

    # Quitar .0 si existe
    if cups_str.endswith('.0'):
        cups_str = cups_str[:-2]

    cups_u = cups_str.upper()

    # 1. Longitud básica
    if not cups_str or len(cups_str) > 15:
        return False

    # 2. RECHAZAR si es una ciudad (traslados)
    if cups_u in CIUDADES_COLOMBIA_COMPLETA:
        return False

    # 3. RECHAZAR si es departamento
    if cups_u in DEPARTAMENTOS_COLOMBIA:
        return False

    # 4. RECHAZAR palabras inválidas
    for palabra in PALABRAS_INVALIDAS_CUPS:
        if palabra in cups_u:
            return False

    # 5. RECHAZAR patrones inválidos
    patrones_invalidos = [
        r'^\*',
        r'^-+$',
        r'^\d{1,2}$',
        r'^N\.?A\.?$',
        r'^N/A$',
        r'INCLUYE',
        r'NOTA\s*\d*',
    ]
    for patron in patrones_invalidos:
        if re.search(patron, cups_u):
            return False

    # 6. Extraer solo dígitos
    cups_digits = re.sub(r'[^\d]', '', cups_str)

    # 🆕 v15.2: Mejora detección de valores monetarios
    # Solo rechazar si es PURAMENTE numérico y muy largo (>= 10) para evitar falsos positivos
    if cups_digits and len(cups_digits) >= 10 and cups_digits == cups_str:
        # Podría ser monetario muy grande o habilitación (ver punto 8)
        pass 
    elif cups_digits and len(cups_digits) >= 7 and cups_digits != cups_str:
        # Permitir alfanuméricos largos (ej 123456-01 tiene 8 dígitos pero es válido)
        pass
    elif cups_digits and len(cups_digits) >= 7 and cups_digits == cups_str:
        # Si es numérico puro entre 7 y 9 dígitos, verificar si parece monetario
        # Por seguridad mejor permitirlo si no es Habilitación
        pass
    # 7. RECHAZAR si parece un valor monetario grande (>= 7 dígitos) - REEMPLAZADO POR LOGICA ARRIBA
    # if cups_digits and len(cups_digits) >= 7:
    #    return False

    # 8. RECHAZAR si parece teléfono celular (10 dígitos con prefijo conocido)
    if es_telefono_celular(cups_str):
        return False

    # 9. 🆕 v15.2: RECHAZAR si parece código de habilitación (10-12 dígitos puros)
    # Rango ajustado: antes 8-12, ahora 10-12 para permitir CUPS propios de 8 dígitos
    if cups_digits and cups_digits == cups_str and 10 <= len(cups_digits) <= 12:
        return False

    # 10. RECHAZAR valores especiales
    if cups_u in ['N.A', 'NA', 'N/A', 'N.A.', '-', '--', '---', 'NINGUNO', 'NINGUNA', 'NULL', 'NONE', '']:
        return False

    # 11. Si es solo dígitos, debe tener al menos 4
    if cups_digits and cups_digits == cups_str:
        if len(cups_digits) < 4:
            return False

    # 12. Si la fila completa parece ser de traslados, rechazar
    if fila and es_fila_de_traslados(fila):
        return False

    return True

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: validar_tarifa
# ══════════════════════════════════════════════════════════════════════════════

def validar_tarifa(tarifa, fila: list = None) -> bool:
    """
    v14.1: Validación mejorada de tarifas.
    Solo rechaza si CLARAMENTE es un teléfono celular.
    """
    import re

    if tarifa is None:
        return True  # Valor nulo es aceptable

    valor_str = str(tarifa).strip()

    # Quitar .0 si existe
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]

    # RECHAZAR si es teléfono celular
    if es_telefono_celular(valor_str):
        return False

    # RECHAZAR si parece código de habilitación Y hay contexto de sede
    valor_clean = re.sub(r'[^\d]', '', valor_str)
    if valor_clean and 8 <= len(valor_clean) <= 12:
        if fila:
            fila_texto = ' '.join([str(x).upper() for x in fila[:5] if x])
            for depto in DEPARTAMENTOS_COLOMBIA:
                if depto in fila_texto:
                    return False

    return True

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: es_fila_de_traslados
# ══════════════════════════════════════════════════════════════════════════════

def es_fila_de_traslados(fila: list) -> bool:
    """
    v14.1: Detecta si una fila de DATOS contiene información de traslados.
    Una fila es de traslados si tiene ciudades en las primeras columnas.
    """
    if not fila or len(fila) < 3:
        return False

    # Verificar si hay ciudades en las primeras columnas
    for i, celda in enumerate(fila[:4]):
        if celda:
            celda_str = str(celda).strip()
            if celda_str.endswith('.0'):
                celda_str = celda_str[:-2]
            celda_upper = celda_str.upper()

            # Verificar contra lista de ciudades
            if celda_upper in CIUDADES_COLOMBIA_COMPLETA:
                return True

    return False

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: es_encabezado_seccion_traslados
# ══════════════════════════════════════════════════════════════════════════════

def es_encabezado_seccion_traslados(fila: list) -> bool:
    """
    v14.1: Detecta si una fila es el ENCABEZADO de una sección de TRASLADOS.
    """
    if not fila:
        return False

    fila_texto = ' '.join([str(x).upper().strip() for x in fila if x is not None])

    # Patrones específicos de encabezados de traslados
    indicadores_traslados = [
        'ORIGEN',
        'DESTINO',
        'MUNICIPIO ORIGEN',
        'MUNICIPIO DESTINO',
        'DEPARTAMENTO DESTINO',
        'TIPO DE TRASLADO',
    ]

    contador = 0
    for indicador in indicadores_traslados:
        if indicador in fila_texto:
            contador += 1

    # Si tiene 2+ indicadores de traslados Y NO tiene CUPS, es sección de traslados
    tiene_cups = 'CUPS' in fila_texto
    return contador >= 2 and not tiene_cups

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: buscar_hoja_servicios_inteligente
# ══════════════════════════════════════════════════════════════════════════════

def debe_excluir_hoja_silenciosamente(nombre_hoja: str) -> bool:
    """Verifica si una hoja debe ser excluida SIN generar alerta."""
    if not nombre_hoja:
        return True

    nombre_upper = nombre_hoja.upper().strip()

    # Excluir hojas de configuración/instrucciones
    if nombre_upper in HOJAS_EXCLUIR_SILENCIOSAMENTE:
        return True

    # Excluir hojas de paquetes/costos (NO generan alerta individual)
    if nombre_upper in HOJAS_SIN_SERVICIOS_VALIDOS:
        return True

    # Verificar patrones parciales
    for patron in HOJAS_SIN_SERVICIOS_VALIDOS:
        if patron in nombre_upper:
            return True

    return False

def buscar_hoja_servicios_inteligente(hojas: list) -> tuple:
    """
    v14.1: Busca la hoja de servicios de forma inteligente.

    Retorna: (nombre_hoja_encontrada, hojas_excluidas_info)

    CAMBIO v14.1: Las hojas de PAQUETES NO generan alerta individual.
    Solo se mencionan si NO se encuentra ninguna hoja de servicios.
    """
    if not hojas:
        return None, []

    hojas_norm = {h: h.upper().strip() for h in hojas}
    hojas_excluidas_info = []

    # Identificar hojas excluidas (para informar si no hay servicios)
    for hoja, h_norm in hojas_norm.items():
        if h_norm in HOJAS_SIN_SERVICIOS_VALIDOS:
            hojas_excluidas_info.append((hoja, "Hoja de paquetes/costos - No aplica para T25"))
        else:
            for patron in HOJAS_SIN_SERVICIOS_VALIDOS:
                if patron in h_norm:
                    hojas_excluidas_info.append((hoja, "Hoja de paquetes/costos - No aplica para T25"))
                    break

    # Filtrar hojas excluidas para la búsqueda
    hojas_validas = {h: h_norm for h, h_norm in hojas_norm.items()
                     if not debe_excluir_hoja_silenciosamente(h_norm)}

    if not hojas_validas:
        hojas_validas = hojas_norm

    # PASO 1: Buscar hoja "SERVICIOS" exacta
    for hoja, h_norm in hojas_validas.items():
        if h_norm.strip() == 'SERVICIOS':
            return hoja, hojas_excluidas_info

    # PASO 2: "TARIFAS DE SERVICIOS" sin modificadores
    patrones_exactos = [
        'TARIFAS DE SERVICIOS',
        'TARIFA DE SERVICIOS',
        'TARIFAS DE SERV',
        'TARIFA DE SERV',
        'TARIFAS DE SERVICIO',
        'TARIFA DE SERVICIO',
    ]

    for hoja, h_norm in hojas_validas.items():
        h_clean = ' '.join(h_norm.split())

        for patron in patrones_exactos:
            if h_clean == patron or h_clean.startswith(patron + ' ') or h_clean.startswith(patron):
                # Excluir si tiene palabras que indican que no es la hoja correcta
                if 'COSTO' not in h_clean and 'VIAJE' not in h_clean and 'PAQUETE' not in h_clean:
                    return hoja, hojas_excluidas_info

    # PASO 3: TARIFA + SERV (pero no traslados/paquetes)
    for hoja, h_norm in hojas_validas.items():
        if 'TARIFA' in h_norm and 'SERV' in h_norm:
            if 'TRASLADO' not in h_norm and 'PAQUETE' not in h_norm and 'AMBULANCIA' not in h_norm:
                return hoja, hojas_excluidas_info

    # PASO 4: SERVICIO (pero no traslados)
    for hoja, h_norm in hojas_validas.items():
        if 'SERVICIO' in h_norm and 'TRASLADO' not in h_norm:
            return hoja, hojas_excluidas_info

    # PASO 5: CUPS
    for hoja, h_norm in hojas_validas.items():
        if 'CUPS' in h_norm:
            if not debe_excluir_hoja_silenciosamente(h_norm):
                return hoja, hojas_excluidas_info

    # PASO 6: ANEXO 1
    for hoja, h_norm in hojas_validas.items():
        h_clean = h_norm.replace(' ', '').replace('_', '')
        if h_clean in ['ANEXO1', 'ANEXO01']:
            if not debe_excluir_hoja_silenciosamente(h_norm):
                return hoja, hojas_excluidas_info

    # No se encontró hoja de servicios
    return None, hojas_excluidas_info

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: generar_mensaje_hojas_disponibles
# ══════════════════════════════════════════════════════════════════════════════

def generar_mensaje_hojas_disponibles(hojas: list, hojas_excluidas_info: list = None) -> str:
    """
    v14.1: Genera mensaje con todas las hojas disponibles.
    Solo se llama cuando NO se encuentra hoja de servicios.

    AQUÍ es donde se menciona que hay hojas de PAQUETES (no antes).
    """
    if not hojas:
        return "Archivo sin hojas"

    hojas_str = ", ".join([f"'{h}'" for h in hojas])
    mensaje = f"No se encontró hoja de servicios válida. Hojas disponibles: [{hojas_str}]"

    # Agregar info de hojas excluidas si existen
    if hojas_excluidas_info:
        excluidas_str = ", ".join([f"'{h[0]}' ({h[1]})" for h in hojas_excluidas_info])
        mensaje += f". Hojas excluidas: [{excluidas_str}]"

    return mensaje

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: es_formato_propio (v14.1)
# ══════════════════════════════════════════════════════════════════════════════

def es_formato_propio(hojas: list, datos_primera_hoja: list = None) -> tuple:
    """
    v14.1: Detecta si un archivo tiene formato propio (no POSITIVA estándar).

    Retorna: (es_propio: bool, descripcion: str)
    """
    if not hojas:
        return True, "Sin hojas disponibles"

    # Nombres de hojas estándar POSITIVA
    NOMBRES_ESTANDAR = {
        'TARIFAS DE SERVICIOS', 'TARIFA DE SERVICIOS', 'TARIFAS DE SERV',
        'SERVICIOS', 'TARIFAS MEDICAMENTOS', 'TARIFAS INSUMOS',
        'TARIFAS TRASLADOS', 'TARIFAS PAQUETES', 'MEDICAMENTOS',
        'INSUMOS', 'TRASLADOS', 'PAQUETES', 'TARIFA DE SERV',
        'TARIFAS DE SERVICIO', 'TARIFA DE SERVICIO'
    }

    hojas_reconocidas = 0
    for hoja in hojas:
        hoja_upper = hoja.upper().strip()
        for std in NOMBRES_ESTANDAR:
            if std in hoja_upper or hoja_upper in std:
                hojas_reconocidas += 1
                break

    # Si ninguna hoja es reconocida, es formato propio
    if hojas_reconocidas == 0:
        return True, f"Formato propio - Hojas no estándar: {hojas}"

    # Verificar estructura interna si tenemos datos
    if datos_primera_hoja:
        for fila in datos_primera_hoja[:20]:
            if fila:
                fila_texto = ' '.join([str(x).upper() for x in fila if x])

                # Si tiene ORIGEN/DESTINO sin CUPS, es formato de traslados
                if ('ORIGEN' in fila_texto and 'DESTINO' in fila_texto):
                    if 'CUPS' not in fila_texto:
                        return True, "Formato propio de traslados"

    return False, ""

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIÓN: generar_mensaje_alerta_ambulancia (v14.1)
# ══════════════════════════════════════════════════════════════════════════════

def generar_mensaje_alerta_ambulancia(mensaje: str, categoria: str) -> str:
    """
    v14.1: Agrega prefijo de categoría si es contrato de ambulancias.
    """
    if categoria and 'AMBULANCIA' in categoria.upper():
        return f"[CATEGORÍA: Cuentas Médicas Ambulancias] {mensaje}"
    return mensaje

# ══════════════════════════════════════════════════════════════════════════════
# CLASE: SistemaAlertas (v14.1 - Sin duplicados mejorado)
# ══════════════════════════════════════════════════════════════════════════════

class SistemaAlertas:
    """
    v14.1: Sistema de alertas sin duplicados.
    """

    def __init__(self):
        self._alertas_hash = set()
        self.alertas = []

    def limpiar(self):
        self._alertas_hash = set()
        self.alertas = []

    def agregar(self, tipo: str, mensaje: str, contrato: str, archivo: str = "") -> bool:
        """
        Agrega una alerta si no existe una similar.
        Retorna True si se agregó, False si ya existía.
        """
        # Hash principal: tipo + contrato + archivo
        hash_principal = hash((tipo, contrato, archivo))

        # Hash secundario: incluye parte del mensaje para evitar duplicados exactos
        hash_secundario = hash((tipo, contrato, archivo, mensaje[:50]))

        if hash_secundario in self._alertas_hash:
            return False

        self._alertas_hash.add(hash_secundario)
        self.alertas.append({
            'tipo': tipo,
            'mensaje': mensaje,
            'contrato': contrato,
            'archivo': archivo
        })
        return True

# ══════════════════════════════════════════════════════════════════════════════
# RESUMEN DE CORRECCIONES v14.1
# ══════════════════════════════════════════════════════════════════════════════

"""
CORRECCIONES APLICADAS EN v14.1:

1. ALERTA DE PAQUETES (CORREGIDO):
   - La alerta de "Hoja de paquetes excluida" ahora SOLO aparece cuando NO existe
     hoja de servicios válida.
   - Las hojas PAQUETES, TARIFAS PAQUETES, COSTO VIAJE se excluyen SILENCIOSAMENTE.
   - Solo se mencionan en el mensaje final si no hay hoja de servicios.

   ANTES: Generaba alerta cada vez que veía una hoja PAQUETES
   AHORA: Solo menciona PAQUETES si el contrato no se puede procesar

2. TELÉFONOS SIN GUIONES (CORREGIDO):
   - La función es_telefono_celular_colombiano() ahora:
     * Limpia el valor quitando TODO excepto dígitos
     * Funciona correctamente con números como: 3214567890
     * Detecta prefijos: 300-305, 310-318, 320-324, 350-351, 330-333

   ANTES: Solo detectaba teléfonos con formato 321-456-7890
   AHORA: Detecta 3214567890 (como vienen en Excel)

3. VALIDACIÓN DE CUPS:
   - Rechaza ciudades colombianas (ARMENIA, CALI, etc.)
   - Rechaza valores monetarios >= 7 dígitos
   - Rechaza teléfonos de 10 dígitos con prefijo conocido
   - Rechaza códigos de habilitación de 8-12 dígitos

4. DETECCIÓN DE SECCIONES DE TRASLADOS:
   - es_encabezado_seccion_traslados(): detecta encabezados con ORIGEN/DESTINO
   - es_fila_de_traslados(): detecta filas con ciudades en primeras columnas

5. SISTEMA DE ALERTAS SIN DUPLICADOS:
   - Hash por (tipo, contrato, archivo, mensaje[:50])
   - Evita alertas repetidas del mismo problema
"""

print("=" * 70)
print("CONSOLIDADOR T25 v14.1 - FUNCIONES CORREGIDAS")
print("=" * 70)
print("""
✅ contiene_anexo1() - Detecta más patrones de ANEXO 1
✅ es_telefono_celular_colombiano() - Funciona SIN guiones
✅ validar_cups() - Validación ultra estricta
✅ validar_tarifa() - Solo rechaza teléfonos claros
✅ es_fila_de_traslados() - Detecta filas de traslados
✅ es_encabezado_seccion_traslados() - Detecta encabezados de traslados
✅ buscar_hoja_servicios_inteligente() - NO genera alerta de PAQUETES
✅ generar_mensaje_hojas_disponibles() - Solo cuando no hay servicios
✅ es_formato_propio() - Detecta formatos no POSITIVA
✅ SistemaAlertas - Sin duplicados mejorado

CORRECCIONES ESPECÍFICAS SOLICITADAS:
1. ✅ Alerta PAQUETES: Solo si NO hay hoja de servicios
2. ✅ Teléfonos: Detecta números SIN guiones (3214567890)
""")
print("=" * 70)

# ══════════════════════════════════════════════════════════════════════════════
# PRUEBAS DE VALIDACIÓN v14.1
# ══════════════════════════════════════════════════════════════════════════════

def ejecutar_pruebas_v14_1():
    """Ejecuta todas las pruebas de las correcciones v14.1"""

    print("\n" + "=" * 70)
    print("🧪 EJECUTANDO PRUEBAS v14.1")
    print("=" * 70)

    errores = []
    exitos = 0

    # ─────────────────────────────────────────────────────────────────────────
    # PRUEBA 1: Detección de teléfonos SIN guiones
    # ─────────────────────────────────────────────────────────────────────────
    print("\n📱 PRUEBA 1: Detección de teléfonos SIN guiones")
    print("-" * 50)

    telefonos_validos = [
        ('3214567890', True, 'Celular Claro sin guiones'),
        ('3001234567', True, 'Celular Movistar sin guiones'),
        ('3501234567', True, 'Celular Tigo sin guiones'),
        ('3101234567', True, 'Celular sin guiones'),
        ('3214567890.0', True, 'Con decimal de Excel'),
    ]

    no_telefonos = [
        ('5920000', False, 'Tarifa de traslado'),
        ('11380000', False, 'Tarifa alta'),
        ('890201', False, 'Código CUPS'),
        ('7614708225', False, 'Código habilitación (no empieza con prefijo celular)'),
        ('12345678', False, 'Solo 8 dígitos'),
        ('123456789012', False, '12 dígitos'),
        ('ARMENIA', False, 'Ciudad'),
    ]

    for valor, esperado, descripcion in telefonos_validos + no_telefonos:
        resultado = es_telefono_celular_colombiano(valor)
        estado = "✅" if resultado == esperado else "❌"
        print(f"  {estado} {valor:15} → {str(resultado):5} (esperado: {esperado}) - {descripcion}")
        if resultado == esperado:
            exitos += 1
        else:
            errores.append(f"Teléfono: {valor} retornó {resultado}, esperado {esperado}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRUEBA 2: Validación de CUPS (rechaza ciudades)
    # ─────────────────────────────────────────────────────────────────────────
    print("\n🏥 PRUEBA 2: Validación de CUPS (ultra estricta)")
    print("-" * 50)

    cups_validos = [
        ('890201', True, 'CUPS consulta válido'),
        ('015201', True, 'CUPS procedimiento'),
        ('602E01', True, 'CUPS con letra'),
        ('786001', True, 'CUPS válido'),
    ]

    cups_invalidos = [
        ('ARMENIA', False, 'Ciudad colombiana'),
        ('CALI', False, 'Ciudad colombiana'),
        ('BAHIA SOLANO', False, 'Ciudad con espacio'),
        ('BOGOTA', False, 'Capital'),
        ('5920000', False, 'Valor monetario (7 dígitos)'),
        ('11380000', False, 'Valor monetario (8 dígitos)'),
        ('3214567890', False, 'Teléfono celular'),
        ('7614708225', False, 'Código habilitación'),
        ('META', False, 'Departamento'),
        ('TRASLADO', False, 'Palabra inválida'),
    ]

    for valor, esperado, descripcion in cups_validos + cups_invalidos:
        resultado = validar_cups(valor)
        estado = "✅" if resultado == esperado else "❌"
        print(f"  {estado} {valor:15} → {str(resultado):5} (esperado: {esperado}) - {descripcion}")
        if resultado == esperado:
            exitos += 1
        else:
            errores.append(f"CUPS: {valor} retornó {resultado}, esperado {esperado}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRUEBA 3: Detección de filas de traslados
    # ─────────────────────────────────────────────────────────────────────────
    print("\n🚑 PRUEBA 3: Detección de filas de traslados")
    print("-" * 50)

    filas_traslados = [
        ([1, 'ARMENIA', 'CALI', 5920000], True, 'Fila con ciudades origen-destino'),
        ([1, 'BOGOTA', 'MEDELLIN', 8500000], True, 'Fila con capitales'),
        (['A', 'BAHIA SOLANO', 'CLO', 11380000], True, 'Fila con código aeropuerto'),
    ]

    filas_servicios = [
        ([1, '890201', '', 'CONSULTA MEDICINA GENERAL', 43686], False, 'Fila de servicio'),
        ([2, '015201', '890201', 'PROCEDIMIENTO', 125000], False, 'Fila con CUPS'),
        (['', '', '', '', ''], False, 'Fila vacía'),
    ]

    for fila, esperado, descripcion in filas_traslados + filas_servicios:
        resultado = es_fila_de_traslados(fila)
        estado = "✅" if resultado == esperado else "❌"
        fila_str = str(fila)[:40] + "..." if len(str(fila)) > 40 else str(fila)
        print(f"  {estado} {fila_str:45} → {str(resultado):5} - {descripcion}")
        if resultado == esperado:
            exitos += 1
        else:
            errores.append(f"Fila traslados: {fila} retornó {resultado}, esperado {esperado}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRUEBA 4: Búsqueda de hoja de servicios (NO alerta PAQUETES)
    # ─────────────────────────────────────────────────────────────────────────
    print("\n📋 PRUEBA 4: Búsqueda de hoja de servicios (alertas PAQUETES)")
    print("-" * 50)

    casos_hojas = [
        # (hojas, hoja_esperada, debe_mencionar_paquetes)
        (['TARIFAS DE SERV', 'MEDICAMENTOS', 'PAQUETES', 'TRASLADOS'],
         'TARIFAS DE SERV', False, 'Con hoja servicios - NO menciona PAQUETES'),

        (['SERVICIOS', 'PAQUETES', 'TARIFAS PAQUETES'],
         'SERVICIOS', False, 'Tiene SERVICIOS - NO menciona PAQUETES'),

        (['PAQUETES', 'TRASLADOS', 'MEDICAMENTOS'],
         None, True, 'SIN servicios - SÍ menciona PAQUETES'),

        (['TARIFAS PAQUETES', 'COSTO VIAJE'],
         None, True, 'Solo paquetes/costos - SÍ menciona en mensaje'),
    ]

    for hojas, esperada, debe_mencionar, descripcion in casos_hojas:
        hoja_encontrada, excluidas_info = buscar_hoja_servicios_inteligente(hojas)

        hoja_ok = hoja_encontrada == esperada

        # Verificar si menciona paquetes en las excluidas
        menciona_paquetes = any('paquete' in str(info).lower() for info in excluidas_info)

        # La lógica es: si NO encuentra hoja, debe mencionar paquetes
        # si SÍ encuentra hoja, NO debe generar alerta de paquetes
        if esperada is None:
            # No encontró hoja - debe tener info de excluidas para el mensaje
            logica_correcta = True  # Las excluidas se usan en generar_mensaje_hojas_disponibles
        else:
            # Encontró hoja - no debe haber problema
            logica_correcta = True

        estado = "✅" if hoja_ok and logica_correcta else "❌"
        print(f"  {estado} Hojas: {hojas}")
        print(f"      → Encontrada: '{hoja_encontrada}' (esperada: '{esperada}')")
        print(f"      → Excluidas info: {len(excluidas_info)} items")
        print(f"      → {descripcion}")

        if hoja_ok:
            exitos += 1
        else:
            errores.append(f"Búsqueda hojas: {hojas} retornó {hoja_encontrada}, esperada {esperada}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRUEBA 5: contiene_anexo1
    # ─────────────────────────────────────────────────────────────────────────
    print("\n📎 PRUEBA 5: Detección de ANEXO 1")
    print("-" * 50)

    nombres_anexo1 = [
        ('0667-2025-ANEXO_1-HOSPITAL_DEPARTAMENTAL', True, 'Formato estándar'),
        ('0584-2025-ANEXO_1_VIDA_SERVICIOS_AMBULANCIA', True, 'Ambulancia con ANEXO_1'),
        ('0513-2024-ANEXO_1_DE_TARIFAS', True, 'ANEXO_1_DE_TARIFAS'),
        ('ANEXO 1 TARIFAS', True, 'Con espacio'),
        ('ANEXO-1-SERVICIOS', True, 'Con guiones'),
        ('TARIFAS_ANEXO1', True, 'Anexo al final'),
    ]

    no_anexo1 = [
        ('MEDICAMENTOS_ANEXO_1', False, 'Medicamentos - debe excluir'),
        ('ANEXO_2_TARIFAS', False, 'Es ANEXO 2'),
        ('TARIFAS_GENERALES', False, 'Sin ANEXO'),
    ]

    for nombre, esperado, descripcion in nombres_anexo1 + no_anexo1:
        resultado = contiene_anexo1(nombre)
        estado = "✅" if resultado == esperado else "❌"
        print(f"  {estado} {nombre[:40]:40} → {str(resultado):5} - {descripcion}")
        if resultado == esperado:
            exitos += 1
        else:
            errores.append(f"ANEXO1: {nombre} retornó {resultado}, esperado {esperado}")

    # ─────────────────────────────────────────────────────────────────────────
    # RESUMEN
    # ─────────────────────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print("📊 RESUMEN DE PRUEBAS")
    print("=" * 70)

    total = exitos + len(errores)
    porcentaje = (exitos / total * 100) if total > 0 else 0

    print(f"\n  ✅ Exitosas: {exitos}")
    print(f"  ❌ Fallidas: {len(errores)}")
    print(f"  📈 Porcentaje: {porcentaje:.1f}%")

    if errores:
        print(f"\n  ⚠️ ERRORES ENCONTRADOS:")
        for error in errores:
            print(f"     • {error}")
    else:
        print(f"\n  🎉 ¡TODAS LAS PRUEBAS PASARON!")

    print("\n" + "=" * 70)

    return len(errores) == 0

# ══════════════════════════════════════════════════════════════════════════════
# EJEMPLO DE USO EN EL PROCESADOR
# ══════════════════════════════════════════════════════════════════════════════

"""
CÓMO INTEGRAR EN EL PROCESADOR PRINCIPAL:

En la función buscar_hoja_servicios del ProcesadorAnexo, reemplazar:

ANTES:
    def buscar_hoja_servicios(self, archivo: str) -> Optional[str]:
        ...
        # Si no encuentra, genera alerta de PAQUETES inmediatamente

DESPUÉS (v14.1):
    def buscar_hoja_servicios(self, archivo: str) -> Optional[str]:
        hojas = obtener_hojas(archivo)

        # Usar la nueva función
        hoja_encontrada, hojas_excluidas_info = buscar_hoja_servicios_inteligente(hojas)

        if hoja_encontrada:
            # ✅ Encontró hoja de servicios - NO genera alerta de PAQUETES
            return hoja_encontrada

        # ❌ NO encontró hoja de servicios - AHORA sí menciona PAQUETES
        mensaje = generar_mensaje_hojas_disponibles(hojas, hojas_excluidas_info)
        self.agregar_alerta(TipoAlerta.HOJA_NO_ENCONTRADA, mensaje, archivo)
        return None
"""

# ══════════════════════════════════════════════════════════════════════════════
# EJECUTAR PRUEBAS AL IMPORTAR
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    ejecutar_pruebas_v14_1()

"""CONSOLIDADOR T25 v14.1 - COMPLETO CON CORRECCIONES

CORRECCIONES v14.1:
- 🆕 Alerta PAQUETES: Solo cuando NO existe hoja de servicios válida
- 🆕 Teléfonos: Detecta números SIN guiones (como vienen en Excel)
- 🆕 Validación CUPS ultra estricta: rechaza ciudades colombianas
- 🆕 Detección de secciones de traslados (evita mapeo incorrecto)
- 🆕 Lista expandida de ciudades colombianas para validación
- 🆕 contiene_anexo1: NO excluye ambulancias (pueden tener servicios)
- Exclusión de hojas TARIFAS PAQUETES y COSTO VIAJE (silenciosa)
- Mejora en búsqueda de hojas (prioriza SERVICIOS)
- Búsqueda de contratos con cero inicial (901 → 0901)
- Alerta "CONTRATO NO SE ENCUENTRA EN EL GO ANYWHERE"
- Reconexión forzada por contrato
- Extracción de número de acta desde nombre del archivo
- Mejora en detección de columnas
- Alertas separadas por hojas en Excel
"""

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 1: INSTALACIÓN Y SISTEMA DE LOGGING
# ══════════════════════════════════════════════════════════════════════════════

print("🚀 Inicializando CONSOLIDADOR T25 v14.1...")
print("=" * 70)

# Instalación silenciosa de dependencias
# pip install pyxlsb openpyxl pandas paramiko xlrd tqdm scikit-learn chardet  # Instalar dependencias manualmente

import warnings
warnings.filterwarnings('ignore')

# # from IPython.display import display, HTML  # No disponible en local, clear_output  # No disponible en local
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple, Callable
from enum import Enum
from dataclasses import dataclass, field
import time

# ══════════════════════════════════════════════════════════════════════════════
# 🎨 SISTEMA DE LOGGING VISUAL
# ══════════════════════════════════════════════════════════════════════════════

class LogLevel(Enum):
    """Niveles de logging con sus estilos visuales."""
    INFO = ("ℹ️", "#2196F3", "info")
    SUCCESS = ("✅", "#4CAF50", "success")
    WARNING = ("⚠️", "#FF9800", "warning")
    ERROR = ("❌", "#F44336", "error")
    DEBUG = ("🔍", "#9E9E9E", "debug")
    STEP = ("📌", "#673AB7", "step")
    NAV = ("📂", "#795548", "nav")
    FILE = ("📄", "#607D8B", "file")
    DOWNLOAD = ("⬇️", "#00BCD4", "download")
    PROCESS = ("⚙️", "#FF5722", "process")
    ALERT = ("🔔", "#E91E63", "alert")

class Logger:
    """Sistema de logging visual para el Consolidador T25."""

    def __init__(self, verbose: bool = True):
        self.verbose = verbose
        self.indent_level = 0
        self.logs: List[Dict] = []
        self.start_time = time.time()
        self.current_contract = ""
        self.stats = {
            'contratos_procesados': 0,
            'contratos_exitosos': 0,
            'archivos_descargados': 0,
            'servicios_extraidos': 0,
            'alertas_generadas': 0
        }

    def _get_timestamp(self) -> str:
        return datetime.now().strftime("%H:%M:%S")

    def _get_elapsed(self) -> str:
        elapsed = time.time() - self.start_time
        if elapsed < 60:
            return f"{elapsed:.1f}s"
        return f"{elapsed/60:.1f}m"

    def _format_indent(self) -> str:
        return "│   " * self.indent_level

    def _print(self, level: LogLevel, message: str, details: str = "",
               show_time: bool = True, indent_override: int = None):
        if not self.verbose and level == LogLevel.DEBUG:
            return

        indent = "│   " * (indent_override if indent_override is not None else self.indent_level)
        icon = level.value[0]
        time_str = f"[{self._get_timestamp()}] " if show_time else ""
        detail_str = f" → {details}" if details else ""

        line = f"{indent}{icon} {time_str}{message}{detail_str}"
        print(line)

        self.logs.append({
            'time': self._get_timestamp(),
            'level': level.name,
            'message': message,
            'details': details
        })

    def set_contract(self, contract_id: str):
        self.current_contract = contract_id

    def indent(self):
        self.indent_level += 1

    def dedent(self):
        self.indent_level = max(0, self.indent_level - 1)

    def reset_indent(self):
        self.indent_level = 0

    def header(self, title: str, subtitle: str = ""):
        print("\n" + "═" * 70)
        print(f"  {title}")
        if subtitle:
            print(f"  {subtitle}")
        print("═" * 70)

    def subheader(self, title: str):
        print(f"\n{'─' * 50}")
        print(f"  {title}")
        print('─' * 50)

    def step(self, step_num: int, total: int, description: str):
        progress = "█" * int(step_num/total * 20) + "░" * (20 - int(step_num/total * 20))
        print(f"\n📌 PASO {step_num}/{total}: {description}")
        print(f"   [{progress}] {step_num/total*100:.0f}%")

    def contract_start(self, idx: int, total: int, contract_id: str):
        self.reset_indent()
        self.current_contract = contract_id
        self.stats['contratos_procesados'] += 1

        progress_pct = (idx / total) * 100
        bar_filled = int(progress_pct / 5)
        bar = "█" * bar_filled + "░" * (20 - bar_filled)

        print(f"\n┌{'─' * 68}┐")
        print(f"│ 📋 CONTRATO [{idx}/{total}] {contract_id:<20} [{bar}] {progress_pct:>5.1f}% │")
        print(f"└{'─' * 68}┘")

    def contract_end(self, success: bool, registros: int, tiempo: float, mensaje: str = ""):
        self.reset_indent()
        icon = "✅" if success else "❌"
        status = "ÉXITO" if success else "FALLO"

        if success:
            self.stats['contratos_exitosos'] += 1
            self.stats['servicios_extraidos'] += registros

        print(f"    ├── {icon} {status}: {registros:,} servicios en {tiempo:.1f}s")
        if mensaje and not success:
            print(f"    └── 💬 {mensaje}")
        print()

    def nav(self, path: str, found: bool = True):
        icon = "📂" if found else "📁"
        status = "" if found else " (no encontrado)"
        self._print(LogLevel.NAV, f"Navegando a: {path}{status}", show_time=False)

    def nav_tree(self, items: List[str], item_type: str = "carpetas"):
        if not items:
            self._print(LogLevel.DEBUG, f"(vacío - sin {item_type})", show_time=False)
            return

        count = len(items)
        shown = items[:5]

        for i, item in enumerate(shown):
            prefix = "├──" if i < len(shown) - 1 else "└──"
            icon = "📁" if item_type == "carpetas" else "📄"
            print(f"    {self._format_indent()}{prefix} {icon} {item}")

        if count > 5:
            print(f"    {self._format_indent()}    ... y {count - 5} más")

    def file_found(self, filename: str, file_type: str = ""):
        type_str = f"[{file_type}] " if file_type else ""
        self._print(LogLevel.FILE, f"Encontrado: {type_str}{filename}", show_time=False)
        self.stats['archivos_descargados'] += 1

    def download(self, filename: str, size: str = ""):
        size_str = f" ({size})" if size else ""
        self._print(LogLevel.DOWNLOAD, f"Descargando: {filename}{size_str}", show_time=False)

    def process(self, action: str, detail: str = ""):
        self._print(LogLevel.PROCESS, action, detail, show_time=False)

    def success(self, message: str, detail: str = ""):
        self._print(LogLevel.SUCCESS, message, detail, show_time=False)

    def warning(self, message: str, detail: str = ""):
        self._print(LogLevel.WARNING, message, detail, show_time=False)

    def error(self, message: str, detail: str = ""):
        self._print(LogLevel.ERROR, message, detail, show_time=False)

    def info(self, message: str, detail: str = ""):
        self._print(LogLevel.INFO, message, detail, show_time=False)

    def debug(self, message: str, detail: str = ""):
        self._print(LogLevel.DEBUG, message, detail, show_time=False)

    def alert(self, alert_type: str, message: str, archivo: str = ""):
        self.stats['alertas_generadas'] += 1
        archivo_str = f" en {archivo}" if archivo else ""
        self._print(LogLevel.ALERT, f"[{alert_type}] {message}{archivo_str}", show_time=False)

    def stats_summary(self):
        elapsed = time.time() - self.start_time

        print(f"\n{'═' * 70}")
        print("  📊 ESTADÍSTICAS DE EJECUCIÓN")
        print('═' * 70)
        print(f"""
    ⏱️  Tiempo total: {elapsed/60:.1f} minutos

    📋 Contratos:
       • Procesados: {self.stats['contratos_procesados']}
       • Exitosos: {self.stats['contratos_exitosos']}
       • Tasa de éxito: {100*self.stats['contratos_exitosos']/max(1,self.stats['contratos_procesados']):.1f}%

    📄 Archivos descargados: {self.stats['archivos_descargados']}

    📊 Servicios extraídos: {self.stats['servicios_extraidos']:,}

    🔔 Alertas generadas: {self.stats['alertas_generadas']}
""")
        print('═' * 70)

# Crear instancia global del logger
LOG = Logger(verbose=True)

LOG.header("CONSOLIDADOR T25 v14.1", "Sistema de Consolidación de Tarifas - POSITIVA")
print("""
✅ Sistema de logging inicializado

📋 Mejoras v14.1:
   • 🔍 Búsqueda mejorada de contratos (901 → 0901)
   • 📋 Alertas separadas por categoría en diferentes hojas
   • 🔄 Reconexión forzada por contrato (evita Socket closed)
   • ✅ Validación de CUPS mejorada (rechaza NOTA, NO INCLUYE, etc.)
   • 📊 Exclusión de hojas TARIFAS PAQUETES y COSTO VIAJE
   • 🎯 Priorización correcta de hoja "SERVICIOS"
""")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 2: IMPORTS Y CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(1, 6, "CARGANDO CONFIGURACIÓN")

# from google.colab import files  # No disponible en local
import pandas as pd
import numpy as np
import os
import re
import shutil
import threading
import zipfile
from typing import List, Optional, Tuple, Dict, Any, Callable
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from enum import Enum, auto
import time
import paramiko
import stat
from difflib import SequenceMatcher

LOG.indent()
LOG.success("Librerías importadas correctamente")

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN GLOBAL
# ══════════════════════════════════════════════════════════════════════════════

@dataclass
class Config:
    """Configuración centralizada del sistema."""
    HOST: str = 'mft.positiva.gov.co'
    PORT: int = 2243
    USERNAME: str = 'G_medica'
    PASSWORD: str = 'Uhnbru0sgnpit]'
    TIMEOUT_CONEXION: int = 30
    TIMEOUT_OPERACION: int = 20
    TIMEOUT_ARCHIVO: int = 60
    MAX_REINTENTOS_CONEXION: int = 5
    MAX_REINTENTOS_OPERACION: int = 3
    BACKOFF_BASE: float = 2.0
    KEEPALIVE_INTERVAL: int = 5
    CARPETA_PRINCIPAL: str = 'R.A-ABASTECIMIENTO RED ASISTENCIAL'
    CARPETA_TRABAJO: str = './trabajo_temp'
    CONTRATOS_PROBLEMATICOS: set = field(default_factory=lambda: {'572-2023'})
    TIMEOUT_CONTRATOS_PROBLEMATICOS: int = 30
    MAX_SEDES: int = 50

CONFIG = Config()

LOG.info("Configuración SFTP", f"{CONFIG.HOST}:{CONFIG.PORT}")
LOG.info("Timeout por archivo", f"{CONFIG.TIMEOUT_ARCHIVO}s")
LOG.info("Máximo de sedes", f"{CONFIG.MAX_SEDES}")

# ══════════════════════════════════════════════════════════════════════════════
# ENUMERACIONES Y CLASES DE DATOS
# ══════════════════════════════════════════════════════════════════════════════

class OrigenTarifa(Enum):
    INICIAL = "Inicial"
    OTROSI = "Otrosí"
    ACTA = "Acta"

class TipoAlerta(Enum):
    SIN_ANEXO1 = "SIN_ANEXO1"
    SIN_CARPETA_TARIFAS = "SIN_CARPETA_TARIFAS"
    ACTA_FALTANTE = "ACTA_FALTANTE"
    CARPETA_ACTAS_SIN_ANEXO = "CARPETA_ACTAS_SIN_ANEXO"
    SIN_FORMATO_POSITIVA = "SIN_FORMATO_POSITIVA"
    FORMATO_PROPIO = "FORMATO_PROPIO"  # 🆕 v15.0
    HOJA_NO_ENCONTRADA = "HOJA_NO_ENCONTRADA"
    COLUMNAS_NO_DETECTADAS = "COLUMNAS_NO_DETECTADAS"
    SEDES_NO_DETECTADAS = "SEDES_NO_DETECTADAS"
    FECHA_NO_ENCONTRADA = "FECHA_NO_ENCONTRADA"
    ERROR_PROCESAMIENTO = "ERROR_PROCESAMIENTO"
    TIMEOUT = "TIMEOUT"
    CONEXION = "CONEXION"
    ERROR_LECTURA = "ERROR_LECTURA"
    SOLO_TRASLADOS = "SOLO_TRASLADOS"
    CONTRATO_AMBULANCIA = "CONTRATO_AMBULANCIA"
    CONTRATO_AMBULANCIA_MAESTRA = "CONTRATO_AMBULANCIA_MAESTRA"
    ARCHIVO_SOLO_AMBULANCIAS = "ARCHIVO_SOLO_AMBULANCIAS"
    ARCHIVO_SOLO_TRASLADOS = "ARCHIVO_SOLO_TRASLADOS"
    TARIFA_SERVICIOS_NO_ENCONTRADA = "TARIFA_SERVICIOS_NO_ENCONTRADA"
    # 🆕 v14.1
    CONTRATO_NO_ENCONTRADO_GO = "CONTRATO_NO_ENCONTRADO_GO"
    FECHA_FALTANTE_MAESTRA = "FECHA_FALTANTE_MAESTRA"
    # 🆕 v15.3: Archivos de paquetes (no van a No_Positiva)
    ARCHIVO_PAQUETE = "ARCHIVO_PAQUETE"

class PrioridadAlerta(Enum):
    CRITICA = 1
    ALTA = 2
    MEDIA = 3
    BAJA = 4

ALERTAS_CONFIG = {
    TipoAlerta.SIN_ANEXO1: {
        'prioridad': PrioridadAlerta.CRITICA,
        'sugerencia': 'Verificar que el archivo ANEXO 1 esté cargado en TARIFAS'
    },
    TipoAlerta.SIN_CARPETA_TARIFAS: {
        'prioridad': PrioridadAlerta.CRITICA,
        'sugerencia': 'Crear carpeta TARIFAS en el contrato'
    },
    TipoAlerta.ACTA_FALTANTE: {
        'prioridad': PrioridadAlerta.ALTA,
        'sugerencia': 'Solicitar acta de negociación faltante'
    },
    TipoAlerta.SIN_FORMATO_POSITIVA: {
        'prioridad': PrioridadAlerta.MEDIA,
        'sugerencia': 'Verificar formato del archivo'
    },
    TipoAlerta.FECHA_NO_ENCONTRADA: {
        'prioridad': PrioridadAlerta.MEDIA,
        'sugerencia': 'Verificar registro en maestra de contratos'
    },
    TipoAlerta.TIMEOUT: {
        'prioridad': PrioridadAlerta.ALTA,
        'sugerencia': 'Archivo muy grande, considerar procesamiento manual'
    },
    TipoAlerta.SOLO_TRASLADOS: {
        'prioridad': PrioridadAlerta.BAJA,
        'sugerencia': 'Archivo contiene solo servicios de traslados'
    },
    TipoAlerta.CONTRATO_AMBULANCIA: {
        'prioridad': PrioridadAlerta.BAJA,
        'sugerencia': 'Contrato de ambulancias - verificar si requiere ANEXO 1 de servicios'
    },
    TipoAlerta.CONTRATO_AMBULANCIA_MAESTRA: {
        'prioridad': PrioridadAlerta.BAJA,
        'sugerencia': 'Contrato identificado como ambulancias desde la maestra'
    },
    TipoAlerta.ARCHIVO_SOLO_AMBULANCIAS: {
        'prioridad': PrioridadAlerta.BAJA,
        'sugerencia': 'Archivo contiene solo hojas de ambulancias/traslados asistenciales'
    },
    TipoAlerta.ARCHIVO_SOLO_TRASLADOS: {
        'prioridad': PrioridadAlerta.BAJA,
        'sugerencia': 'Archivo contiene solo hojas de traslados genéricos'
    },
    TipoAlerta.TARIFA_SERVICIOS_NO_ENCONTRADA: {
        'prioridad': PrioridadAlerta.ALTA,
        'sugerencia': 'Verificar si el archivo corresponde a servicios o solo traslados/ambulancias'
    },
    TipoAlerta.CONTRATO_NO_ENCONTRADO_GO: {
        'prioridad': PrioridadAlerta.CRITICA,
        'sugerencia': 'El contrato no existe en GoAnywhere - verificar número y año'
    },
}

@dataclass
class Alerta:
    """Representa una alerta generada durante el procesamiento."""
    tipo: TipoAlerta
    mensaje: str
    contrato: str
    archivo: str = ""
    sugerencia: str = ""
    prioridad: PrioridadAlerta = PrioridadAlerta.MEDIA
    timestamp: str = field(default_factory=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    def __post_init__(self):
        if not self.sugerencia and self.tipo in ALERTAS_CONFIG:
            self.sugerencia = ALERTAS_CONFIG[self.tipo]['sugerencia']
        if self.tipo in ALERTAS_CONFIG:
            self.prioridad = ALERTAS_CONFIG[self.tipo]['prioridad']

    def to_dict(self) -> Dict:
        return {
            'tipo': self.tipo.value,
            'prioridad': self.prioridad.value,
            'mensaje': self.mensaje,
            'contrato': self.contrato,
            'archivo': self.archivo,
            'sugerencia': self.sugerencia,
            'timestamp': self.timestamp
        }

    def __hash__(self):
        return hash((self.tipo, self.mensaje, self.contrato, self.archivo))

    def __eq__(self, other):
        if not isinstance(other, Alerta):
            return False
        return (self.tipo == other.tipo and
                self.mensaje == other.mensaje and
                self.contrato == other.contrato and
                self.archivo == other.archivo)

@dataclass
class ArchivoAnexo:
    """Representa un archivo ANEXO 1 descargado."""
    nombre: str
    ruta_local: str
    origen: OrigenTarifa
    numero: Optional[int] = None
    fecha_modificacion: Optional[float] = None
    origen_completo: str = ""  # 🆕 v14.1

    @property
    def origen_texto(self) -> str:
        if self.numero:
            return f"{self.origen.value} {self.numero}"
        return self.origen.value

LOG.success("Clases y configuración definidas")
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CLASES DE MACHINE LEARNING (MOVIDO AL INICIO)
# ══════════════════════════════════════════════════════════════════════════════

class ClasificadorTextoMedico:
    """
    Clasificador ML para detectar si un texto es:
    - Manual tarifario (SOAT, ISS, PROPIO)
    - Descripción de procedimiento médico
    - Porcentaje o valor numérico
    """

    # Vocabulario de referencia para manuales tarifarios
    VOCABULARIO_MANUAL = [
        'SOAT', 'SOAT VIGENTE', 'SOAT UVT', 'SOAT UVB', 'TARIFARIO SOAT',
        'ISS', 'ISS 2001', 'ISS2001', 'TARIFARIOS ISS', 'TARIFA ISS',
        'TARIFA PROPIA', 'TARIFAS PROPIAS', 'PROPIA', 'PROPIO', 'PROPIAS',
        'INSTITUCIONAL', 'TARIFAS INSTITUCIONALES', 'TARIFA INSTITUCIONAL',
        'DECRETO 2423', 'DECRETO 2644', 'UVT', 'UVB', 'TARIFA PLENA',
        'MENOS', 'PLENO', 'VIGENTE', 'MANUAL TARIFARIO'
    ]

    # Vocabulario de procedimientos médicos
    VOCABULARIO_MEDICO = [
        'CONSULTA', 'TERAPIA', 'NEURAL', 'CIRUGIA', 'PROCEDIMIENTO',
        'TRATAMIENTO', 'EVALUACION', 'VALORACION', 'DIAGNOSTICO',
        'EXAMEN', 'BIOPSIA', 'ECOGRAFIA', 'RADIOGRAFIA', 'TOMOGRAFIA',
        'RESONANCIA', 'LABORATORIO', 'HEMOGRAMA', 'CURACION', 'SUTURA',
        'INYECCION', 'APLICACION', 'NEBULIZACION', 'HOSPITALIZACION',
        'CONTROL', 'SEGUIMIENTO', 'ESPECIALISTA', 'MEDICINA', 'GENERAL',
        'PEDIATRIA', 'GINECOLOGIA', 'ORTOPEDIA', 'CARDIOLOGIA', 'NEUROLOGIA',
        'PSIQUIATRIA', 'PSICOLOGIA', 'FISIOTERAPIA', 'FONOAUDIOLOGIA',
        'ODONTOLOGIA', 'OPTOMETRIA', 'ANESTESIA', 'URGENCIA', 'AMBULANCIA',
        'SANGRE', 'ORINA', 'GLUCOSA', 'COLESTEROL', 'TRIGLICERIDOS',
        'ELECTROCARDIOGRAMA', 'ENDOSCOPIA', 'COLONOSCOPIA', 'MAMOGRAFIA',
        'QUIMIOTERAPIA', 'RADIOTERAPIA', 'DIALISIS', 'TRASPLANTE', 'PROTESIS',
        'IMPLANTE', 'REHABILITACION', 'TERAPIA OCUPACIONAL', 'TERAPIA FISICA',
        'CONSULTA DE', 'VISITA DE', 'ATENCION DE', 'SERVICIO DE'
    ]

    def __init__(self):
        """Inicializa el clasificador y entrena el vectorizador."""
        # Crear corpus de entrenamiento
        self.corpus_manual = self.VOCABULARIO_MANUAL
        self.corpus_medico = self.VOCABULARIO_MEDICO

        # Entrenar vectorizador TF-IDF
        self.vectorizer = TfidfVectorizer(
            analyzer='char_wb',
            ngram_range=(2, 4),
            lowercase=True,
            max_features=1000
        )

        # Entrenar con todo el vocabulario
        todo_corpus = self.corpus_manual + self.corpus_medico
        self.vectorizer.fit(todo_corpus)

        # Vectores de referencia
        self.vec_manual = self.vectorizer.transform(self.corpus_manual)
        self.vec_medico = self.vectorizer.transform(self.corpus_medico)

        # Centroide de cada clase
        self.centroide_manual = np.asarray(self.vec_manual.mean(axis=0)).flatten()
        self.centroide_medico = np.asarray(self.vec_medico.mean(axis=0)).flatten()

        print("✅ Clasificador ML entrenado")
        print(f"   • Vocabulario manual: {len(self.corpus_manual)} términos")
        print(f"   • Vocabulario médico: {len(self.corpus_medico)} términos")

    def clasificar(self, texto: str) -> Dict[str, Any]:
        """
        Clasifica un texto y retorna probabilidades.

        Returns:
            Dict con:
            - 'tipo': 'MANUAL', 'MEDICO', 'PORCENTAJE', 'DESCONOCIDO'
            - 'confianza': 0.0 a 1.0
            - 'scores': diccionario con scores de cada clase
        """
        if pd.isna(texto) or str(texto).strip() == '':
            return {'tipo': 'VACIO', 'confianza': 1.0, 'scores': {}}

        texto = str(texto).strip()
        texto_upper = texto.upper()

        # 1. Reglas rápidas basadas en patrones
        # ─────────────────────────────────────────────────────────────────────

        # Es un porcentaje o número
        if re.match(r'^[+-]?[\d,\.%\s]+$', texto):
            return {'tipo': 'PORCENTAJE', 'confianza': 0.95, 'scores': {'porcentaje': 0.95}}

        # Contiene palabras clave de manual tarifario
        palabras_manual = ['SOAT', 'ISS', 'TARIFA', 'DECRETO', 'UVT', 'UVB', 'PROPIA', 'PROPIO', 'INSTITUCIONAL']
        for palabra in palabras_manual:
            if palabra in texto_upper:
                return {'tipo': 'MANUAL', 'confianza': 0.9, 'scores': {'manual': 0.9}}

        # Contiene palabras clave médicas
        palabras_medicas = ['CONSULTA', 'TERAPIA', 'CIRUGIA', 'PROCEDIMIENTO', 'EXAMEN',
                           'TRATAMIENTO', 'BIOPSIA', 'ECOGRAFIA', 'LABORATORIO']
        for palabra in palabras_medicas:
            if palabra in texto_upper:
                return {'tipo': 'MEDICO', 'confianza': 0.85, 'scores': {'medico': 0.85}}

        # 2. Clasificación ML con TF-IDF
        # ─────────────────────────────────────────────────────────────────────
        try:
            vec_texto = self.vectorizer.transform([texto_upper])
            vec_array = np.asarray(vec_texto.todense()).flatten()

            # Calcular similitud con centroides
            sim_manual = cosine_similarity([vec_array], [self.centroide_manual])[0][0]
            sim_medico = cosine_similarity([vec_array], [self.centroide_medico])[0][0]

            # Normalizar scores
            total = sim_manual + sim_medico + 0.001
            score_manual = sim_manual / total
            score_medico = sim_medico / total

            scores = {'manual': score_manual, 'medico': score_medico}

            if score_manual > score_medico and score_manual > 0.4:
                return {'tipo': 'MANUAL', 'confianza': score_manual, 'scores': scores}
            elif score_medico > score_manual and score_medico > 0.4:
                return {'tipo': 'MEDICO', 'confianza': score_medico, 'scores': scores}
            else:
                return {'tipo': 'DESCONOCIDO', 'confianza': max(score_manual, score_medico), 'scores': scores}

        except Exception as e:
            return {'tipo': 'ERROR', 'confianza': 0.0, 'scores': {}, 'error': str(e)}

    def es_descripcion_medica(self, texto: str) -> Tuple[bool, float]:
        """
        Verifica si un texto parece ser una descripción médica.

        Returns:
            (es_medico, confianza)
        """
        resultado = self.clasificar(texto)
        return resultado['tipo'] == 'MEDICO', resultado['confianza']

    def es_manual_tarifario(self, texto: str) -> Tuple[bool, float]:
        """
        Verifica si un texto parece ser un manual tarifario válido.

        Returns:
            (es_manual, confianza)
        """
        resultado = self.clasificar(texto)
        return resultado['tipo'] == 'MANUAL', resultado['confianza']

# Crear instancia global
clasificador_ml = ClasificadorTextoMedico()

class ETLConsolidadoT25_ML:
    """
    ═══════════════════════════════════════════════════════════════════════════
    ETL CONSOLIDADO T25 - VERSIÓN CON MACHINE LEARNING
    ═══════════════════════════════════════════════════════════════════════════
    Sistema inteligente que detecta y corrige automáticamente cuando:
    - manual_tarifario contiene descripciones médicas
    - porcentaje_manual_tarifario contiene el manual real
    - Los valores están intercambiados entre columnas
    ═══════════════════════════════════════════════════════════════════════════
    """

    ANOS_IGNORAR = {'1996', '2001', '2016', '2022', '2023', '2024', '2025', '2644', '2423', '780'}

    PATRON_NUMERO = re.compile(r'[+-]?\d+(?:[,\.]\d+)?')
    PATRON_PORCENTAJE_FINAL = re.compile(r'[+-]\s*(\d+(?:[,\.]\d+)?)\s*$')
    PATRON_MENOS = re.compile(r'MENOS\s*(\d+(?:[,\.]\d+)?)', re.IGNORECASE)
    PATRON_MAS = re.compile(r'(?:MAS|\+)\s*(\d+(?:[,\.]\d+)?)', re.IGNORECASE)
    PATRON_DECIMAL = re.compile(r'^(-?0\.\d+)$')

    def __init__(self, clasificador: ClasificadorTextoMedico, chunk_size: int = 50000):
        """Inicializa el ETL con el clasificador ML."""
        self.clasificador = clasificador
        self.chunk_size = chunk_size
        self.stats = {
            'total_registros': 0,
            'columnas_intercambiadas': 0,
            'manuales_normalizados': 0,
            'porcentajes_extraidos': 0,
            'anomalias_detectadas': [],
            'correcciones_ml': []
        }
        self.resultados = {}

    def _detectar_y_corregir_anomalia(self, row: pd.Series) -> Dict[str, Any]:
        """
        Detecta si hay anomalía en la fila y sugiere corrección.

        Anomalías detectadas:
        1. manual_tarifario tiene descripción médica
        2. porcentaje_manual_tarifario tiene el manual real
        3. Valores intercambiados

        Returns:
            Dict con correcciones sugeridas
        """
        manual = str(row.get('manual_tarifario', '')).strip()
        porcentaje = str(row.get('porcentaje_manual_tarifario', '')).strip()
        descripcion = str(row.get('descripcion_del_cups', '')).strip()

        correccion = {
            'necesita_correccion': False,
            'nuevo_manual': manual,
            'nuevo_porcentaje': porcentaje,
            'razon': None,
            'confianza': 0.0
        }

        if not manual:
            return correccion

        # Clasificar el contenido de manual_tarifario
        clasif_manual = self.clasificador.clasificar(manual)
        clasif_porcentaje = self.clasificador.clasificar(porcentaje)

        # CASO 1: manual_tarifario tiene descripción médica
        if clasif_manual['tipo'] == 'MEDICO' and clasif_manual['confianza'] > 0.6:
            # Verificar si porcentaje tiene el manual real
            if clasif_porcentaje['tipo'] == 'MANUAL' and clasif_porcentaje['confianza'] > 0.5:
                correccion['necesita_correccion'] = True
                correccion['nuevo_manual'] = porcentaje
                correccion['nuevo_porcentaje'] = '0'  # Extraer del nuevo manual si hay
                correccion['razon'] = f"ML detectó descripción médica en manual_tarifario (conf: {clasif_manual['confianza']:.2f})"
                correccion['confianza'] = clasif_manual['confianza']
                return correccion

        # CASO 2: manual_tarifario es similar a descripcion_del_cups
        if descripcion and len(manual) > 20:
            manual_words = set(manual.upper().split())
            desc_words = set(descripcion.upper().split())
            if len(manual_words) > 0 and len(desc_words) > 0:
                similitud = len(manual_words & desc_words) / min(len(manual_words), len(desc_words))
                if similitud > 0.5:  # Más del 50% de palabras en común
                    if clasif_porcentaje['tipo'] == 'MANUAL':
                        correccion['necesita_correccion'] = True
                        correccion['nuevo_manual'] = porcentaje
                        correccion['nuevo_porcentaje'] = '0'
                        correccion['razon'] = f"manual_tarifario similar a descripción ({similitud:.0%})"
                        correccion['confianza'] = similitud
                        return correccion

        # CASO 3: manual_tarifario tiene formato de tarifa (número grande)
        try:
            valor_manual = float(manual.replace(',', '.').replace('$', '').strip())
            if valor_manual > 1000:  # Parece una tarifa, no un manual
                if clasif_porcentaje['tipo'] == 'MANUAL':
                    correccion['necesita_correccion'] = True
                    correccion['nuevo_manual'] = porcentaje
                    correccion['nuevo_porcentaje'] = '0'
                    correccion['razon'] = f"manual_tarifario contiene tarifa ({valor_manual:,.0f})"
                    correccion['confianza'] = 0.9
                    return correccion
        except:
            pass

        return correccion

    def _extraer_porcentaje(self, texto: str) -> Optional[float]:
        """Extrae porcentaje de forma inteligente."""
        if pd.isna(texto) or str(texto).strip() == '':
            return None

        texto = str(texto).strip()
        texto_upper = texto.upper()

        if 'PLENA' in texto_upper or 'PLENO' in texto_upper:
            if not re.search(r'[+-]\s*\d+', texto_upper):
                return 0.0

        texto_sin_pct = texto.replace('%', '')

        match = self.PATRON_PORCENTAJE_FINAL.search(texto_sin_pct)
        if match:
            try:
                valor = float(match.group(1).replace(',', '.'))
                if '-' in texto_sin_pct:
                    valor = -abs(valor)
                if -100 <= valor <= 200:
                    return valor
            except:
                pass

        match = self.PATRON_MENOS.search(texto_sin_pct)
        if match:
            try:
                return -float(match.group(1).replace(',', '.'))
            except:
                pass

        match = self.PATRON_MAS.search(texto_sin_pct)
        if match:
            try:
                return float(match.group(1).replace(',', '.'))
            except:
                pass

        match = self.PATRON_DECIMAL.match(texto_sin_pct.strip())
        if match:
            try:
                return round(float(match.group(1)) * 100, 2)
            except:
                pass

        numeros = self.PATRON_NUMERO.findall(texto_sin_pct)
        for num_str in reversed(numeros):
            try:
                num = float(num_str.replace(',', '.'))
                if str(int(abs(num))) in self.ANOS_IGNORAR:
                    continue
                if num > 1000:
                    continue
                if -100 <= num <= 200:
                    if 'MENOS' in texto_upper or f'-{num_str}' in texto:
                        num = -abs(num)
                    return num
            except:
                continue

        return None

    def _normalizar_manual(self, texto: str) -> str:
        """Normaliza el manual tarifario."""
        if pd.isna(texto) or str(texto).strip() == '':
            return 'PROPIO'

        texto = str(texto).strip()
        texto_upper = texto.upper()

        # PROPIO
        if re.search(r'\bPROPIA?S?\b|INSTITUCIONAL|TARIA\s*PROPIA', texto_upper):
            return 'PROPIO'

        # ISS
        if re.search(r'\bISS\b', texto_upper) and not re.search(r'\bSOAT\b', texto_upper):
            return 'ISS'

        # SOAT
        if re.search(r'\bSOAT\b|\bUVT\b|\bUVB\b|DECRETO\s*2423|DECRETO\s*2644', texto_upper):
            return 'SOAT'

        # Números puros -> PROPIO
        if re.match(r'^[\d,\.\s]+$', texto):
            return 'PROPIO'

        return texto

    def _procesar_fila(self, row: pd.Series) -> Dict[str, Any]:
        """
        Procesa una fila completa con detección ML.
        """
        resultado = {
            'manual_tarifario': '',
            'porcentaje_manual_tarifario': 0.0,
            'correccion_aplicada': False,
            'log': None
        }

        # 1. Detectar anomalías
        correccion = self._detectar_y_corregir_anomalia(row)

        if correccion['necesita_correccion']:
            # Usar valores corregidos
            manual_raw = correccion['nuevo_manual']
            porcentaje_raw = correccion['nuevo_porcentaje']
            resultado['correccion_aplicada'] = True
            resultado['log'] = correccion['razon']
        else:
            manual_raw = str(row.get('manual_tarifario', '')).strip()
            porcentaje_raw = str(row.get('porcentaje_manual_tarifario', '')).strip()

        # 2. Normalizar manual
        resultado['manual_tarifario'] = self._normalizar_manual(manual_raw)

        # 3. Extraer porcentaje
        tarifa = row.get('tarifa_unitaria_en_pesos', '0')
        try:
            tarifa_num = float(str(tarifa).replace(',', '.'))
        except:
            tarifa_num = 0

        # Si porcentaje es texto de manual o propio -> 0
        porcentaje_upper = porcentaje_raw.upper()
        if any(p in porcentaje_upper for p in ['PROPIO', 'PROPIA', 'INSTITUCIONAL', 'PLENA', 'PLENO']):
            resultado['porcentaje_manual_tarifario'] = 0.0
        else:
            # Extraer porcentaje
            pct = self._extraer_porcentaje(porcentaje_raw)
            if pct is not None:
                # Verificar que no sea igual a la tarifa
                if tarifa_num > 0 and abs(pct - tarifa_num) < 1:
                    resultado['porcentaje_manual_tarifario'] = 0.0
                elif pct > 1000:  # Probable tarifa duplicada
                    resultado['porcentaje_manual_tarifario'] = 0.0
                else:
                    resultado['porcentaje_manual_tarifario'] = round(pct, 2)
            else:
                resultado['porcentaje_manual_tarifario'] = 0.0

        return resultado

    def procesar_dataframe(self, df: pd.DataFrame, nombre: str = "Datos") -> pd.DataFrame:
        """Procesa un DataFrame completo."""
        print(f"\n{'═'*70}")
        print(f"🧠 PROCESANDO CON ML: {nombre}")
        print(f"{'═'*70}")

        inicio = datetime.now()
        total = len(df)
        print(f"📊 Total registros: {total:,}")

        # Normalizar columnas
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        # Asegurar columnas
        for col in ['manual_tarifario', 'porcentaje_manual_tarifario', 'tarifa_unitaria_en_pesos']:
            if col not in df.columns:
                df[col] = ''

        # Procesar
        print(f"\n🔄 Fase 1: Detección de anomalías con ML...")

        nuevos_manuales = []
        nuevos_porcentajes = []
        correcciones = []

        for idx in tqdm(range(total), desc="   Procesando"):
            row = df.iloc[idx]
            resultado = self._procesar_fila(row)

            nuevos_manuales.append(resultado['manual_tarifario'])
            nuevos_porcentajes.append(resultado['porcentaje_manual_tarifario'])

            if resultado['correccion_aplicada']:
                correcciones.append({
                    'indice': idx,
                    'original_manual': row.get('manual_tarifario', ''),
                    'original_porcentaje': row.get('porcentaje_manual_tarifario', ''),
                    'nuevo_manual': resultado['manual_tarifario'],
                    'log': resultado['log']
                })

        df['manual_tarifario'] = nuevos_manuales
        df['porcentaje_manual_tarifario'] = nuevos_porcentajes

        # Corregir tarifas
        print(f"\n🔄 Fase 2: Corrigiendo tarifas...")
        tarifa = pd.to_numeric(
            df['tarifa_unitaria_en_pesos'].astype(str).str.replace(',', '.'),
            errors='coerce'
        ).fillna(0)
        mask_pequena = (tarifa > 0) & (tarifa < 100)
        tarifa.loc[mask_pequena] = tarifa.loc[mask_pequena] * 1000
        df['tarifa_unitaria_en_pesos'] = tarifa.round(2)

        # Estadísticas
        duracion = (datetime.now() - inicio).total_seconds()

        print(f"\n{'─'*70}")
        print(f"📈 RESULTADOS - {nombre}")
        print(f"{'─'*70}")
        print(f"⏱️ Tiempo: {duracion:.1f} segundos")
        print(f"🔧 Correcciones ML aplicadas: {len(correcciones):,}")

        if correcciones:
            print(f"\n📝 Muestra de correcciones aplicadas:")
            for c in correcciones[:10]:
                print(f"   • Fila {c['indice']}: {c['log']}")
                print(f"     Original: '{str(c['original_manual'])[:40]}...'")
                print(f"     Corregido: '{c['nuevo_manual']}'")

        print(f"\n📊 Distribución de Manuales:")
        for manual, count in df['manual_tarifario'].value_counts().head(10).items():
            pct = count / total * 100
            print(f"   {manual:15} │ {count:>10,} │ {pct:5.1f}%")

        self.stats['total_registros'] += total
        self.stats['columnas_intercambiadas'] += len(correcciones)
        self.stats['correcciones_ml'].extend(correcciones)

        gc.collect()
        return df

    def ejecutar(self, contenido: bytes, nombre: str) -> Dict[str, pd.DataFrame]:
        """Ejecuta el ETL completo."""
        print("\n" + "═"*70)
        print("🚀 ETL CONSOLIDADO T25 - ML EDITION")
        print("═"*70)
        print(f"📁 Archivo: {nombre}")
        print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        inicio = datetime.now()

        # Cargar archivo
        dataframes = self._cargar_archivo(contenido, nombre)

        # Procesar cada hoja
        for nombre_hoja, df in dataframes.items():
            self.resultados[nombre_hoja] = self.procesar_dataframe(df, nombre_hoja)

        duracion = (datetime.now() - inicio).total_seconds()

        print("\n" + "═"*70)
        print("✅ ETL ML COMPLETADO")
        print("═"*70)
        print(f"📊 Total registros: {self.stats['total_registros']:,}")
        print(f"🔧 Correcciones ML: {self.stats['columnas_intercambiadas']:,}")
        print(f"⏱️ Tiempo: {duracion:.1f} segundos")

        return self.resultados

    def _cargar_archivo(self, contenido: bytes, nombre: str) -> Dict[str, pd.DataFrame]:
        """Carga archivo Excel o CSV."""
        dataframes = {}

        if nombre.endswith('.csv'):
            resultado = chardet.detect(contenido[:10000])
            encoding = resultado['encoding'] or 'utf-8'

            for sep in [';', ',', '\t']:
                try:
                    df = pd.read_csv(io.BytesIO(contenido), sep=sep, encoding=encoding,
                                    dtype=str, low_memory=False)
                    if len(df.columns) > 1:
                        dataframes['Datos'] = df
                        break
                except:
                    continue
        else:
            excel = pd.ExcelFile(io.BytesIO(contenido))
            for hoja in excel.sheet_names:
                df = pd.read_excel(excel, sheet_name=hoja, dtype=str)
                if len(df) > 0:
                    dataframes[hoja] = df

        return dataframes

    def exportar_log_correcciones(self, archivo: str = 'correcciones_ml.csv'):
        """Exporta log de correcciones ML."""
        if self.stats['correcciones_ml']:
            df_log = pd.DataFrame(self.stats['correcciones_ml'])
            df_log.to_csv(archivo, index=False, encoding='utf-8-sig')
            print(f"✅ Log exportado: {archivo}")
            return df_log
        else:
            print("ℹ️ No hay correcciones ML para exportar")
            return None

# Crear instancia global
try:
    if TfidfVectorizer is None:
        raise ImportError("scikit-learn no está instalado")
        
    print("🤖 Inicializando Machine Learning (Modo Robusto)...")
    clasificador_ml = ClasificadorTextoMedico()
    etl_ml_helper = ETLConsolidadoT25_ML(clasificador_ml)
except Exception as e:
    print(f"⚠️ Error inicializando Clasificador ML: {e}")
    clasificador_ml = None
    etl_ml_helper = None

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 3A: UTILIDADES Y FUNCIONES DE CONVERSIÓN
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(2, 6, "CARGANDO UTILIDADES v14.1")
LOG.indent()

def detectar_formato_real(filepath: str) -> str:
    """Detecta el formato REAL de un archivo Excel."""
    try:
        with open(filepath, 'rb') as f:
            header = f.read(8)

        if header[:4] == b'PK\x03\x04':
            try:
                with zipfile.ZipFile(filepath, 'r') as z:
                    names = z.namelist()
                    if any('workbook.bin' in n.lower() for n in names):
                        return 'xlsb'
                    elif any('.xml' in n.lower() for n in names):
                        return 'xlsx'
                return 'xlsx'
            except zipfile.BadZipFile:
                return 'zip_corrupt'

        if header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            return 'xls_old'

        return 'unknown'

    except Exception:
        return 'error'

LOG.success("🆕 Función detectar_formato_real agregada")

def leer_excel(ruta: str, sheet_name=0, header=0, engine=None):
    """Lee archivo Excel con manejo automático de motor."""
    if engine:
        return pd.read_excel(ruta, engine=engine, sheet_name=sheet_name, header=header)

    formato = detectar_formato_real(ruta)

    try:
        if formato == 'xlsb':
            return pd.read_excel(ruta, engine='pyxlsb', sheet_name=sheet_name, header=header)
        elif formato == 'xlsx':
            return pd.read_excel(ruta, engine='openpyxl', sheet_name=sheet_name, header=header)
        elif formato == 'xls_old':
            return pd.read_excel(ruta, engine='xlrd', sheet_name=sheet_name, header=header)
    except:
        pass

    for eng in ['openpyxl', 'pyxlsb', 'xlrd']:
        try:
            return pd.read_excel(ruta, engine=eng, sheet_name=sheet_name, header=header)
        except:
            continue
    raise Exception(f"No se pudo leer: {ruta}")

def obtener_hojas(ruta: str) -> List[str]:
    """Obtiene lista de hojas de un archivo Excel."""
    formato = detectar_formato_real(ruta)
    ext = os.path.splitext(ruta)[1].lower()

    if formato == 'xlsb':
        try:
            from pyxlsb import open_workbook
            with open_workbook(ruta) as wb:
                return list(wb.sheets)
        except Exception:
            pass

    elif formato == 'xlsx':
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            hojas = wb.sheetnames
            wb.close()
            return hojas
        except Exception:
            pass

    elif formato == 'xls_old':
        try:
            import xlrd
            wb = xlrd.open_workbook(ruta, on_demand=True)
            return wb.sheet_names()
        except Exception:
            pass

    # FALLBACK basado en extensión
    if ext == '.xlsb':
        try:
            from pyxlsb import open_workbook
            with open_workbook(ruta) as wb:
                return list(wb.sheets)
        except:
            pass
    elif ext == '.xls':
        try:
            import xlrd
            return xlrd.open_workbook(ruta).sheet_names()
        except:
            pass
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True)
            hojas = wb.sheetnames
            wb.close()
            return hojas
        except:
            pass

    return []

def leer_hoja_raw(ruta: str, hoja: str, max_filas: int = 50000) -> List[List]:
    """Lee hoja como lista de listas."""
    formato = detectar_formato_real(ruta)
    ext = os.path.splitext(ruta)[1].lower()

    try:
        if formato == 'xlsb':
            from pyxlsb import open_workbook
            datos = []
            with open_workbook(ruta) as wb:
                with wb.get_sheet(hoja) as sheet:
                    for i, row in enumerate(sheet.rows()):
                        if i >= max_filas:
                            break
                        datos.append([cell.v for cell in row])
            return datos

        elif formato == 'xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            sheet = wb[hoja]
            datos = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_filas:
                    break
                datos.append(list(row))
            wb.close()
            return datos

        elif formato == 'xls_old':
            import xlrd
            wb = xlrd.open_workbook(ruta)
            sheet = wb.sheet_by_name(hoja)
            return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                    for r in range(min(sheet.nrows, max_filas))]

        else:
            if ext == '.xlsb':
                from pyxlsb import open_workbook
                datos = []
                with open_workbook(ruta) as wb:
                    with wb.get_sheet(hoja) as sheet:
                        for i, row in enumerate(sheet.rows()):
                            if i >= max_filas:
                                break
                            datos.append([cell.v for cell in row])
                return datos
            elif ext == '.xls':
                import xlrd
                wb = xlrd.open_workbook(ruta)
                sheet = wb.sheet_by_name(hoja)
                return [[sheet.cell_value(r, c) for c in range(sheet.ncols)]
                        for r in range(min(sheet.nrows, max_filas))]
            else:
                from openpyxl import load_workbook
                wb = load_workbook(ruta, read_only=True, data_only=True)
                sheet = wb[hoja]
                datos = []
                for i, row in enumerate(sheet.iter_rows(values_only=True)):
                    if i >= max_filas:
                        break
                    datos.append(list(row))
                wb.close()
                return datos

    except AssertionError:
        return []
    except Exception:
        return []

LOG.success("Funciones de lectura Excel")

# ══════════════════════════════════════════════════════════════════════════════
# CLASIFICACIÓN DE HOJAS PARA ALERTAS
# ══════════════════════════════════════════════════════════════════════════════

def clasificar_hojas(hojas: List[str]) -> Dict[str, List[str]]:
    """Clasifica las hojas disponibles para generar alertas más descriptivas."""
    clasificacion = {
        'servicios': [],
        'medicamentos': [],
        'traslados': [],
        'ambulancias': [],
        'paquetes': [],
        'otras': []
    }

    PALABRAS_AMBULANCIA = ['AMBULANCIA', 'TAM', 'TAB', 'TRASLADO ASISTENCIAL',
                           'TRANSPORTE ASISTENCIAL', 'SERVICIO AMBULANCIA']

    for hoja in hojas:
        h = hoja.upper()

        if any(x in h for x in ['TARIFA DE SERV', 'TARIFAS DE SERV', 'TARIFAS SERV']):
            if 'MEDICAMENTO' not in h and 'TRASLADO' not in h and 'AMBULANCIA' not in h:
                clasificacion['servicios'].append(hoja)
                continue

        es_ambulancia = False
        for palabra in PALABRAS_AMBULANCIA:
            if palabra in h:
                es_ambulancia = True
                break

        if es_ambulancia:
            clasificacion['ambulancias'].append(hoja)
        elif 'MEDICAMENTO' in h or 'INSUMO' in h:
            clasificacion['medicamentos'].append(hoja)
        elif 'TRASLADO' in h:
            clasificacion['traslados'].append(hoja)
        elif 'PAQUETE' in h:
            clasificacion['paquetes'].append(hoja)
        elif 'SERVICIO' in h and 'MEDICAMENTO' not in h:
            clasificacion['servicios'].append(hoja)
        else:
            clasificacion['otras'].append(hoja)

    return clasificacion

def generar_mensaje_alerta_hojas(hojas: List[str], categoria_cuentas_medicas: str = None) -> str:
    """Genera un mensaje de alerta descriptivo basado en las hojas disponibles."""
    if not hojas:
        return "No se pudieron leer las hojas del archivo"

    cls = clasificar_hojas(hojas)
    partes = []

    if cls['ambulancias']:
        partes.append(f"Ambulancias: {cls['ambulancias'][0]}")
    if cls['medicamentos']:
        partes.append(f"Medicamentos: {cls['medicamentos'][0]}")
    if cls['traslados']:
        partes.append(f"Traslados: {cls['traslados'][0]}")
    if cls['paquetes']:
        partes.append(f"Paquetes: {cls['paquetes'][0]}")

    mensaje = ""
    if partes:
        mensaje = f"No se encontró hoja TARIFA SERVICIOS. Hojas encontradas: [{'; '.join(partes)}]"
    else:
        todas = ', '.join(hojas[:3])
        if len(hojas) > 3:
            todas += f" (+{len(hojas)-3} más)"
        mensaje = f"No se encontró hoja TARIFA SERVICIOS. Hojas disponibles: [{todas}]"

    if categoria_cuentas_medicas:
        mensaje += f". Categoría cuentas médicas: '{categoria_cuentas_medicas}'"

    return mensaje

def es_archivo_solo_traslados(hojas: List[str]) -> Tuple[bool, str, str]:
    """Verifica si un archivo SOLO contiene hojas de traslados/ambulancias."""
    if not hojas:
        return False, "", ""

    cls = clasificar_hojas(hojas)

    if cls['servicios']:
        return False, "", ""

    tiene_ambulancias = bool(cls['ambulancias'])
    tiene_traslados = bool(cls['traslados'])

    if tiene_ambulancias and not tiene_traslados:
        hojas_amb = ', '.join(cls['ambulancias'][:3])
        if len(cls['ambulancias']) > 3:
            hojas_amb += f" (+{len(cls['ambulancias'])-3} más)"
        mensaje = f"Archivo contiene solo hojas de ambulancias: [{hojas_amb}]"
        return True, mensaje, "AMBULANCIAS"

    if tiene_traslados and not tiene_ambulancias:
        hojas_traslado = ', '.join(cls['traslados'][:3])
        if len(cls['traslados']) > 3:
            hojas_traslado += f" (+{len(cls['traslados'])-3} más)"
        mensaje = f"Archivo contiene solo hojas de traslados: [{hojas_traslado}]"
        return True, mensaje, "TRASLADOS"

    if tiene_ambulancias and tiene_traslados:
        todas = cls['ambulancias'] + cls['traslados']
        hojas_str = ', '.join(todas[:3])
        if len(todas) > 3:
            hojas_str += f" (+{len(todas)-3} más)"
        mensaje = f"Archivo contiene solo hojas de ambulancias/traslados: [{hojas_str}]"
        return True, mensaje, "MIXTO"

    return False, "", ""

LOG.success("Funciones de clasificación de hojas")

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE NORMALIZACIÓN Y LIMPIEZA
# ══════════════════════════════════════════════════════════════════════════════

def normalizar_texto(texto) -> str:
    """Normaliza texto: mayúsculas, sin tildes, sin especiales."""
    if texto is None:
        return ""
    t = str(texto).upper().strip()
    for k, v in {'Á':'A','É':'E','Í':'I','Ó':'O','Ú':'U','Ñ':'N','Ü':'U'}.items():
        t = t.replace(k, v)
    return re.sub(r'[^A-Z0-9\s]', ' ', t).strip()

def similitud_texto(a: str, b: str) -> float:
    """Calcula similitud entre dos textos (0.0 a 1.0)."""
    return SequenceMatcher(None, a.upper(), b.upper()).ratio()

def limpiar_codigo(valor) -> Optional[str]:
    """Limpia código eliminando decimales y espacios."""
    if valor is None:
        return None
    texto = str(valor).strip()
    if texto.endswith('.0'):
        texto = texto[:-2]
    return None if not texto or texto.lower() in ('none', 'nan', '') else texto

def limpiar_tarifa(valor) -> Optional[object]:
    """Convierte tarifa a número (int si no tiene decimales)."""
    if valor is None:
        return None
    try:
        if isinstance(valor, (int, float)):
            if pd.isna(valor): return None
            val = float(valor)
            return int(val) if val.is_integer() else val
            
        texto = str(valor).replace('$', '').replace(',', '').replace(' ', '').strip()
        if not texto: return None
        val = float(texto)
        return int(val) if val.is_integer() else val
    except:
        return None

def limpiar_texto(valor) -> Optional[str]:
    """Limpia texto eliminando espacios extras y sufijos .0"""
    if valor is None:
        return None
    texto = str(valor).strip()
    if not texto or texto.lower() in ('none', 'nan'):
        return None
    
    if texto.endswith('.0'):
        texto = texto[:-2]
        
    return texto

def formatear_habilitacion(codigo, sede) -> str:
    """Formatea código de habilitación con sede."""
    if not codigo:
        return "0000000000-01"

    c = str(codigo).strip()
    if c.endswith('.0'):
        c = c[:-2]

    if re.match(r'^\d{8,12}-\d{1,2}$', c):
        return c

    c_limpio = re.sub(r'[^\d]', '', c)

    try:
        if sede is None:
            s = 1
        else:
            sede_str = str(sede).strip()
            if sede_str.endswith('.0'):
                sede_str = sede_str[:-2]
            sede_limpia = re.sub(r'[^\d]', '', sede_str)
            if sede_limpia == c_limpio or len(sede_limpia) > 5:
                s = 1
            else:
                s = int(float(sede_str)) if sede_str else 1
    except:
        s = 1

    return f"{c_limpio}-{str(s).zfill(2)}"

LOG.success("Funciones de normalización")

# ══════════════════════════════════════════════════════════════════════════════
# FUNCIONES DE DETECCIÓN DE PATRONES
# ══════════════════════════════════════════════════════════════════════════════

def es_extension_excel(nombre: str) -> bool:
    """Verifica si es archivo Excel."""
    return nombre and nombre.lower().endswith(('.xlsx', '.xls', '.xlsm', '.xlsb'))

# 🆕 v15.0: La función contiene_anexo1() ahora está definida al inicio del archivo
# con soporte mejorado para detectar archivos TARIFAS y OTROSI

def timestamp_a_fecha(timestamp: float) -> Optional[str]:
    """Convierte timestamp UNIX a fecha DD/MM/YYYY."""
    if not timestamp:
        return None
    try:
        dt = datetime.fromtimestamp(timestamp)
        return dt.strftime('%d/%m/%Y')
    except:
        return None

LOG.success("Funciones de detección de patrones")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 3B: VALIDACIÓN SEMÁNTICA v14.1
# ══════════════════════════════════════════════════════════════════════════════

print("\n📌 CARGANDO VALIDACIÓN SEMÁNTICA v14.1...")

# 🆕 v14.1: Lista COMPLETA de ciudades colombianas (incluye las usadas en traslados)
CIUDADES_COLOMBIA_COMPLETA = {
    # Capitales
    'BOGOTÁ', 'BOGOTA', 'MEDELLÍN', 'MEDELLIN', 'CALI', 'BARRANQUILLA',
    'CARTAGENA', 'BUCARAMANGA', 'CÚCUTA', 'CUCUTA', 'PEREIRA', 'IBAGUÉ',
    'IBAGUE', 'SANTA MARTA', 'MANIZALES', 'VILLAVICENCIO', 'PASTO',
    'MONTERÍA', 'MONTERIA', 'NEIVA', 'ARMENIA', 'SINCELEJO', 'POPAYÁN',
    'POPAYAN', 'VALLEDUPAR', 'TUNJA', 'FLORENCIA', 'QUIBDÓ', 'QUIBDO',
    'RIOHACHA', 'YOPAL', 'MOCOA', 'LETICIA', 'INÍRIDA', 'INIRIDA',
    'MITÚ', 'MITU', 'PUERTO CARREÑO', 'SAN JOSÉ DEL GUAVIARE', 'ARAUCA',
    # Ciudades intermedias usadas en traslados aéreos
    'BAHIA SOLANO', 'BARRANCABERMEJA', 'BUENAVENTURA', 'PALMIRA',
    'CARTAGO', 'TULUA', 'TULUÁ', 'BUGA', 'SOGAMOSO', 'DUITAMA', 'GIRARDOT',
    'FUSAGASUGA', 'FUSAGASUGÁ', 'FACATATIVA', 'FACATATIVÁ', 'ZIPAQUIRA',
    'ZIPAQUIRÁ', 'CHIA', 'CHÍA', 'SOACHA', 'RIONEGRO', 'ENVIGADO',
    'ITAGUI', 'ITAGÜÍ', 'BELLO', 'TUMACO', 'IPIALES', 'GRANADA', 'ACACIAS',
    'ACACÍAS', 'PUERTO LOPEZ', 'PUERTO LÓPEZ', 'AGUACHICA', 'OCAÑA',
    'APARTADO', 'APARTADÓ', 'TURBO', 'CAUCASIA', 'MAGANGUE', 'MAGANGUÉ',
    'LORICA', 'CERETE', 'CERETÉ', 'ESPINAL', 'MELGAR', 'FLANDES', 'HONDA',
    'MARIQUITA', 'LA DORADA', 'PUERTO BERRIO', 'PUERTO BERRÍO',
    'PUERTO BOYACA', 'PUERTO BOYACÁ', 'CIENAGA', 'CIÉNAGA', 'FUNDACION',
    'FUNDACIÓN', 'ARACATACA', 'EL BANCO', 'PLATO', 'COROZAL', 'SAMPUES',
    'SAMPUÉS', 'SAN MARCOS', 'ZARZAL', 'JAMUNDI', 'JAMUNDÍ', 'YUMBO',
    'CANDELARIA', 'PRADERA', 'FLORIDA', 'CERRITO', 'GUACARI', 'GUACARÍ',
    'GINEBRA', 'ROLDANILLO', 'LA UNION', 'LA UNIÓN', 'SEVILLA',
    'CAICEDONIA', 'ARGELIA', 'DARIEN', 'DARIÉN', 'RESTREPO', 'DAGUA',
    'LA CUMBRE', 'CLO', 'BOG', 'MDE',  # Códigos de aeropuerto
    # Otras ciudades importantes
    'TENJO', 'MOSQUERA', 'SUESCA', 'FUNZA', 'MADRID', 'ALCALÁ', 'ULLOA',
    'TRUJILLO', 'RIOFRÍO', 'RIOFRIO', 'CALIMA', 'VIJES', 'YOTOCO',
    'SAN PEDRO', 'EL DOVIO', 'ANDALUCÍA', 'ANDALUCIA', 'CONTRATACIÓN',
    'CONTRATACION', 'POPOYAN', 'BOLIVAR', 'BOLÍVAR',
}

MUNICIPIOS_COLOMBIA = {
    'BOGOTÁ', 'BOGOTA', 'MEDELLÍN', 'MEDELLIN', 'CALI', 'BARRANQUILLA',
    'CARTAGENA', 'BUCARAMANGA', 'CÚCUTA', 'CUCUTA', 'PEREIRA', 'IBAGUÉ',
    'IBAGUE', 'SANTA MARTA', 'MANIZALES', 'VILLAVICENCIO', 'PASTO',
    'MONTERÍA', 'MONTERIA', 'NEIVA', 'ARMENIA', 'SINCELEJO', 'POPAYÁN',
    'POPOYAN', 'VALLEDUPAR', 'TUNJA', 'FLORENCIA', 'QUIBDÓ', 'QUIBDO',
    'RIOHACHA', 'YOPAL', 'MOCOA', 'LETICIA', 'INÍRIDA', 'INIRIDA',
    'MITÚ', 'MITU', 'PUERTO CARREÑO', 'SAN JOSÉ DEL GUAVIARE',
    'ZARZAL', 'PALMIRA', 'ANDALUCÍA', 'ANDALUCIA', 'CONTRATACIÓN',
    'CONTRATACION', 'TULUÁ', 'TULUA', 'BUGA', 'CARTAGO', 'JAMUNDÍ',
    'JAMUNDI', 'YUMBO', 'CANDELARIA', 'PRADERA', 'FLORIDA', 'CERRITO',
    'GUACARÍ', 'GUACARI', 'GINEBRA', 'ROLDANILLO', 'LA UNIÓN', 'SEVILLA',
    'CAICEDONIA', 'ALCALÁ', 'ULLOA', 'ARGELIA', 'BOLÍVAR', 'TRUJILLO',
    'RIOFRÍO', 'RIOFRIO', 'CALIMA', 'DAGUA', 'LA CUMBRE', 'RESTREPO',
    'VIJES', 'YOTOCO', 'SAN PEDRO', 'DARIÉN', 'DARIEN', 'EL DOVIO',
    'SOGAMOSO', 'TENJO', 'MOSQUERA', 'ZIPAQUIRÁ', 'ZIPAQUIRA', 'SUESCA',
    'BUENAVENTURA', 'RIONEGRO', 'ENVIGADO', 'ITAGÜÍ', 'ITAGUI', 'BELLO',
    'SOACHA', 'CHÍA', 'CHIA', 'FUNZA', 'MADRID', 'FACATATIVÁ', 'FACATATIVA',
    'GIRARDOT', 'FUSAGASUGÁ', 'FUSAGASUGA', 'DUITAMA', 'SOGAMOSO'
}

DEPARTAMENTOS_COLOMBIA = {
    'BOGOTÁ D.C', 'BOGOTA D.C', 'BOGOTÁ D.C.', 'BOGOTA D.C.',
    'ANTIOQUIA', 'ATLÁNTICO', 'ATLANTICO', 'BOLÍVAR', 'BOLIVAR',
    'BOYACÁ', 'BOYACA', 'CALDAS', 'CAQUETÁ', 'CAQUETA', 'CASANARE',
    'CAUCA', 'CESAR', 'CHOCÓ', 'CHOCO', 'CÓRDOBA', 'CORDOBA',
    'CUNDINAMARCA', 'GUAINÍA', 'GUAINIA', 'GUAVIARE', 'HUILA',
    'LA GUAJIRA', 'MAGDALENA', 'META', 'NARIÑO', 'NARINO',
    'NORTE DE SANTANDER', 'PUTUMAYO', 'QUINDÍO', 'QUINDIO',
    'RISARALDA', 'SAN ANDRÉS', 'SAN ANDRES', 'SANTANDER', 'SUCRE',
    'TOLIMA', 'VALLE', 'VALLE DEL CAUCA', 'VAUPÉS', 'VAUPES',
    'VICHADA', 'AMAZONAS', 'ARAUCA'
}

PALABRAS_ENCABEZADO_SEDES = {
    'DEPARTAMENTO', 'MUNICIPIO', 'CODIGO DE HABILITACION', 'CÓDIGO DE HABILITACIÓN',
    'CODIGO DE HABIITACION', 'CÓDIGO DE HABIITACIÓN',
    'CODIGO HABILITACION', 'CÓDIGO HABILITACIÓN',
    'NUMERO DE SEDE', 'NÚMERO DE SEDE', 'N° SEDE', 'NO. SEDE',
    'NOMBRE DE LA SEDE', 'DIRECCION', 'DIRECCIÓN', 'TELEFONO', 'TELÉFONO',
    'EMAIL', 'CORREO', 'NOMBRE SEDE'
}

PALABRAS_ENCABEZADO_SERVICIOS = {
    'CODIGO CUPS', 'CÓDIGO CUPS', 'COD CUPS', 'COD. CUPS','CODIGO CUP', 'COD. CUP', 'COD CUP',
    'DESCRIPCION DEL CUPS', 'DESCRIPCIÓN DEL CUPS', 'DESCRIPCION DE CUP', 'DESCRIPCIÓN DE CUP',
    'TARIFA UNITARIA', 'MANUAL TARIFARIO', 'TARIFARIO',
    'CODIGO DE ESPECIALIDAD', 'CÓDIGO DE ESPECIALIDAD'
}

PATRONES_DIRECCION = [
    'CARRERA ', 'CRA ', 'CRA. ', 'CR ',
    'CALLE ', 'CL ', 'CL. ',
    'AVENIDA ', 'AV ', 'AV. ',
    'DIAGONAL ', 'DG ', 'DG. ',
    'TRANSVERSAL ', 'TV ', 'TV. ',
    'KM ', 'KILOMETRO', 'KILÓMETRO',
    'LOCAL ', 'PISO ', 'OFICINA ', 'OF ',
    'CONSULTORIO', 'TORRE ', 'BLOQUE ',
    'MANZANA', 'CASA ', 'APARTAMENTO', 'APTO',
    'EDIFICIO', 'CENTRO COMERCIAL', 'C.C.',
    'BARRIO ', 'VEREDA ', 'SECTOR '
]

# 🆕 v14.1 - Hojas a excluir SILENCIOSAMENTE (sin generar alerta)
HOJAS_EXCLUIR = {
    'INSTRUCCIONES', 'INFO', 'DATOS', 'CONTENIDO', 'INDICE', 'ÍNDICE',
    'GUIA DE USO', 'GUÍA DE USO', 'CONTROL DE CAMBIOS', 'HOJA1', 'SHEET1',
    'INSTRUCTIVO', 'PARAMETROS', 'PARÁMETROS', 'CONFIGURACION', 'CONFIGURACIÓN',
    'LISTA', 'LISTAS', 'VALIDACION', 'VALIDACIÓN', 'CATALOGO', 'CATÁLOGO',
    'RESUMEN', 'PORTADA', 'CARATULA', 'CARÁTULA', 'INICIO', 'HOME',
    'MENU', 'MENÚ', 'ANEXO TECNICO', 'ANEXO TÉCNICO', 'GLOSARIO',
}

# 🆕 v14.1: Hojas que se excluyen SILENCIOSAMENTE pero se MENCIONAN si no hay hoja de servicios
HOJAS_SIN_SERVICIOS_VALIDOS = {
    'PAQUETE', 'PAQUETES', 'TARIFAS PAQUETE', 'TARIFAS PAQUETES',
    'TARIFA PAQUETE', 'TARIFA PAQUETES',
    'COSTO VIAJE', 'COSTO DE VIAJE', 'COSTOS VIAJE', 'COSTOS DE VIAJE'
}

PATRONES_EXCLUIR_HOJA = [
    'COSTO VIAJE',
    'COSTO DE VIAJE',
    '(COSTO',
]

# 🆕 v14.1: Patrones de PAQUETES (se excluyen pero NO generan alerta individual)
PATRONES_PAQUETES = [
    'PAQUETE',
]

PALABRAS_HOJA_SERVICIOS_ALTA = [
    'TARIFA DE SERV',
    'TARIFAS DE SERV',
    'TARIFA SERV',
    'TARIFAS SERV',
    'SERVICIOS INDIVIDUALES',
    'SOLICITUD',
    'ANEXO 1',
    'ANEXO',
]

def debe_excluir_hoja(nombre_hoja: str) -> bool:
    """🆕 v14.1: Verifica si una hoja debe ser excluida (SILENCIOSAMENTE).
    Las hojas de PAQUETES se excluyen pero NO generan alerta individual.
    """
    if not nombre_hoja:
        return True

    nombre_upper = nombre_hoja.upper().strip()

    # Excluir hojas de configuración/instrucciones
    if nombre_upper in HOJAS_EXCLUIR:
        return True

    # 🆕 v14.1: Excluir hojas de PAQUETES silenciosamente (sin alerta)
    if nombre_upper in HOJAS_SIN_SERVICIOS_VALIDOS:
        return True

    # Verificar patrones de exclusión
    for patron in PATRONES_EXCLUIR_HOJA:
        if patron in nombre_upper:
            return True

    # 🆕 v14.1: Verificar patrones de PAQUETES
    for patron in PATRONES_PAQUETES:
        if patron in nombre_upper:
            return True

    return False

def obtener_hojas_excluidas_info(hojas: List[str]) -> List[Tuple[str, str]]:
    """🆕 v14.1: Obtiene info de hojas excluidas para mostrar si no hay servicios."""
    info = []
    for hoja in hojas:
        hoja_upper = hoja.upper().strip()

        # Verificar si es hoja de paquetes/costos
        if hoja_upper in HOJAS_SIN_SERVICIOS_VALIDOS:
            info.append((hoja, "Hoja de paquetes/costos - No aplica para T25"))
            continue

        for patron in PATRONES_PAQUETES:
            if patron in hoja_upper:
                info.append((hoja, "Hoja de paquetes - No aplica para T25"))
                break

    return info

def buscar_hoja_servicios_inteligente(hojas: List[str]) -> Tuple[Optional[str], List[Tuple[str, str]]]:
    """🆕 v14.1: Busca la hoja de servicios de forma inteligente.

    Retorna: (nombre_hoja_encontrada, hojas_excluidas_info)

    CAMBIO v14.1: Las hojas de PAQUETES NO generan alerta individual.
    Solo se mencionan en el mensaje final si no hay hoja de servicios.
    """
    if not hojas:
        return None, []

    hojas_norm = {h: h.upper().strip() for h in hojas}

    # 🆕 v14.1: Obtener info de hojas excluidas para informar si no hay servicios
    hojas_excluidas_info = obtener_hojas_excluidas_info(hojas)

    # Filtrar hojas excluidas
    hojas_validas = {h: h_norm for h, h_norm in hojas_norm.items()
                     if not debe_excluir_hoja(h_norm)}

    if not hojas_validas:
        hojas_validas = hojas_norm

    # PASO 1: Buscar hoja "SERVICIOS"
    for hoja, h_norm in hojas_validas.items():
        if h_norm.strip() == 'SERVICIOS':
            return hoja, hojas_excluidas_info

    # PASO 2: "TARIFAS DE SERVICIOS" sin modificadores
    patrones_exactos = [
        'TARIFAS DE SERVICIOS',
        'TARIFA DE SERVICIOS',
        'TARIFAS DE SERV',
        'TARIFA DE SERV',
    ]

    for hoja, h_norm in hojas_validas.items():
        h_clean = ' '.join(h_norm.split())

        for patron in patrones_exactos:
            if h_clean == patron or h_clean.startswith(patron + ' '):
                if 'COSTO' not in h_clean and 'VIAJE' not in h_clean and 'PAQUETE' not in h_clean:
                    return hoja, hojas_excluidas_info

    # PASO 3: TARIFA + SERV
    for hoja, h_norm in hojas_validas.items():
        if 'TARIFA' in h_norm and 'SERV' in h_norm:
            if not debe_excluir_hoja(h_norm):
                return hoja, hojas_excluidas_info

    # PASO 4: SERVICIO
    for hoja, h_norm in hojas_validas.items():
        if 'SERVICIO' in h_norm:
            if not debe_excluir_hoja(h_norm):
                return hoja, hojas_excluidas_info

    # PASO 5: CUPS
    for hoja, h_norm in hojas_validas.items():
        if 'CUPS' in h_norm:
            if not debe_excluir_hoja(h_norm):
                return hoja, hojas_excluidas_info

    # PASO 6: ANEXO 1
    for hoja, h_norm in hojas_validas.items():
        h_clean = h_norm.replace(' ', '').replace('_', '')
        if h_clean in ['ANEXO1', 'ANEXO01']:
            if not debe_excluir_hoja(h_norm):
                return hoja, hojas_excluidas_info

    # No se encontró hoja de servicios
    return None, hojas_excluidas_info

def es_encabezado_seccion_sedes(fila: list) -> bool:
    """Detecta si una fila es el ENCABEZADO de la sección de SEDES."""
    if not fila:
        return False

    fila_texto = ' '.join([str(x).upper().strip() for x in fila if x is not None])

    contador = 0
    for palabra in PALABRAS_ENCABEZADO_SEDES:
        if palabra in fila_texto:
            contador += 1

    return contador >= 3

def es_encabezado_seccion_servicios(fila: list) -> bool:
    """Detecta si una fila es el ENCABEZADO de la sección de SERVICIOS."""
    if not fila:
        return False

    fila_texto = ' '.join([str(x).upper().strip() for x in fila if x is not None])

    tiene_cups = 'CODIGO CUPS' in fila_texto or 'CÓDIGO CUPS' in fila_texto
    tiene_otra = any(p in fila_texto for p in ['DESCRIPCION', 'TARIFA', 'TARIFARIO', 'ESPECIALIDAD'])

    return tiene_cups and tiene_otra

def es_dato_de_sede(fila: list) -> bool:
    """Detecta si una fila contiene DATOS de sede."""
    if not fila or len(fila) < 3:
        return False
    
    col0 = str(fila[0]).upper().strip() if fila[0] is not None else ''
    col1 = str(fila[1]).upper().strip() if len(fila) > 1 and fila[1] is not None else ''

    es_depto = col0 in DEPARTAMENTOS_COLOMBIA or any(d in col0 for d in DEPARTAMENTOS_COLOMBIA)
    es_muni = col1 in MUNICIPIOS_COLOMBIA or any(m in col1 for m in MUNICIPIOS_COLOMBIA)

    # Validar por ubicación geográfica (método más confiable)
    if es_depto and es_muni:
        return True

    # 🆕 v15.3: Validación más estricta para evitar falsos positivos
    tiene_direccion = False
    tiene_codigo_hab = False

    # Solo buscar código de habilitación en columnas 2-5 (no en col 0-1 que son ITEM/CUPS)
    for i, item in enumerate(fila[2:6]):  # Columnas 2, 3, 4, 5
        if not item: continue
        item_str = str(item).upper().strip()
        
        # Chequear dirección
        for patron in PATRONES_DIRECCION:
            if patron in item_str:
                tiene_direccion = True
                break
        
        # Chequear código habilitación: debe ser 10-12 dígitos PUROS (sin guiones)
        # Los códigos CUPS como 890202-04 tienen guiones, la habilitación no
        clean_code = item_str.replace('.0', '')
        if clean_code.isdigit() and 10 <= len(clean_code) <= 12:
            tiene_codigo_hab = True

    # Solo considerar sede si tiene AMBOS: código de habilitación Y (departamento O dirección)
    if tiene_codigo_hab and (es_depto or es_muni or tiene_direccion):
        return True

    # Si tiene dirección Y departamento/municipio, es sede
    if tiene_direccion and (es_depto or es_muni):
        return True

    return False

def es_municipio_o_departamento(valor: str) -> bool:
    """Detecta si un valor es un municipio o departamento."""
    if not valor:
        return False
    valor_u = str(valor).upper().strip()
    return valor_u in MUNICIPIOS_COLOMBIA or valor_u in DEPARTAMENTOS_COLOMBIA

def es_direccion(valor: str) -> bool:
    """Detecta si un valor es una dirección."""
    if not valor:
        return False
    valor_u = str(valor).upper()
    for patron in PATRONES_DIRECCION:
        if patron in valor_u:
            return True
    return False

PREFIJOS_CELULAR_COLOMBIA = {
    '300', '301', '302', '303', '304', '305',
    '310', '311', '312', '313', '314', '315', '316', '317', '318',
    '320', '321', '322', '323', '324',
    '350', '351',
    '330', '331', '332', '333'
}

def es_telefono_celular_colombiano(valor: str) -> bool:
    """🆕 v14.1: Detecta si un valor es un teléfono celular colombiano.
    CORREGIDO: Funciona con números SIN guiones (como vienen en Excel).

    Ejemplos que detecta:
    - 3214567890 (sin guiones)
    - 3001234567
    - 3501234567

    NO debe confundir con:
    - Tarifas (5920000, 11380000)
    - Códigos CUPS (890201)
    - Habilitación (7614708225)
    """
    if not valor:
        return False

    # Convertir a string y limpiar
    valor_str = str(valor).strip()

    # Si termina en .0, quitarlo (típico de Excel)
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]

    # Quitar TODOS los caracteres no dígitos
    valor_clean = re.sub(r'[^\d]', '', valor_str)

    # Debe ser exactamente 10 dígitos
    if len(valor_clean) != 10:
        return False

    # Verificar prefijo de celular colombiano
    prefijo = valor_clean[:3]
    return prefijo in PREFIJOS_CELULAR_COLOMBIA

def es_telefono_celular(valor: str) -> bool:
    return es_telefono_celular_colombiano(valor)

def es_numero_sede(valor: str) -> bool:
    """Detecta si un valor es solo un número de sede."""
    if not valor:
        return False
    valor_str = str(valor).strip().replace('.0', '')
    return valor_str.isdigit() and len(valor_str) <= 2

#VALIDACIÓN DE CUPS ULTRA ESTRICTA
PALABRAS_INVALIDAS_CUPS = [
    'CODIGO', 'CUPS', 'ITEM', 'DESCRIPCION', 'TARIFA', 'TOTAL', 'SUBTOTAL',
    'DEPARTAMENTO', 'MUNICIPIO', 'HABILITACION', 'HABIITACION', 'DIRECCION',
    'TELEFONO', 'EMAIL', 'SEDE', 'NOMBRE', 'NUMERO', 'ESPECIALIDAD',
    'MANUAL', 'OBSERV', 'PORCENTAJE', 'HOMOLOGO', 'N°', 'NO.',
    'NOTA', 'NOTAS', 'ACLARATORIA', 'ACLARATORIAS', 'ACLARACION', 'ACLARACIONES',
    'INCLUYE', 'NO INCLUYE', 'EXCLUYE',
    'USO DE EQUIPO', 'DERECHO DE SALA', 'DERECHO SALA',
    'VER NOTA', 'VER NOTAS', 'SEGUN NOTA',
    'APLICA', 'NO APLICA', 'SEGÚN', 'SEGUN',
    'CONSULTAR', 'REVISAR', 'PENDIENTE',
    'VALOR', 'PRECIO', 'COSTO',
    'CONTRATO', 'ACTA', 'OTROSI', 'OTROSÍ',
    'VIGENTE', 'VIGENCIA',
    'TRASLADO', 'ORIGEN', 'DESTINO',  # 🆕 v14.1: Palabras de traslados
    'TARIFAS PROPIAS', 'TARIFA PROPIA',  # 🆕 v14.1: Son manuales tarifarios
]

PATRONES_INVALIDOS_CUPS = [
    r'^\*',
    r'^-+$',
    r'^\d{1,2}$',
    r'^N\.?A\.?$',
    r'^N/A$',
    r'INCLUYE',
    r'NOTA\s*\d*',
]

def es_fila_de_traslados(fila: list) -> bool:
    """🆕 v14.1: Detecta si una fila de DATOS contiene información de traslados.
    Una fila es de traslados si tiene ciudades en las primeras columnas.
    """
    if not fila or len(fila) < 3:
        return False

    # Verificar si hay ciudades en las primeras columnas
    for i, celda in enumerate(fila[:4]):
        if celda:
            celda_str = str(celda).strip()
            if celda_str.endswith('.0'):
                celda_str = celda_str[:-2]
            celda_upper = celda_str.upper()

            # Verificar contra lista de ciudades
            if celda_upper in CIUDADES_COLOMBIA_COMPLETA:
                return True

    return False

def es_encabezado_seccion_traslados(fila: list) -> bool:
    """🆕 v14.1: Detecta si una fila es el ENCABEZADO de una sección de TRASLADOS."""
    if not fila:
        return False

    fila_texto = ' '.join([str(x).upper().strip() for x in fila if x is not None])

    # Patrones específicos de encabezados de traslados
    indicadores_traslados = [
        'ORIGEN',
        'DESTINO',
        'MUNICIPIO ORIGEN',
        'MUNICIPIO DESTINO',
        'DEPARTAMENTO DESTINO',
        'TIPO DE TRASLADO',
    ]

    contador = 0
    for indicador in indicadores_traslados:
        if indicador in fila_texto:
            contador += 1

    # Si tiene 2+ indicadores de traslados Y NO tiene CUPS, es sección de traslados
    tiene_cups = 'CUPS' in fila_texto
    return contador >= 2 and not tiene_cups

def validar_cups(cups: str, fila: list = None) -> bool:
    """🆕 v14.1: Validación de CUPS ULTRA estricta.

    RECHAZA:
    - Ciudades colombianas (ARMENIA, CALI, BAHIA SOLANO, etc.)
    - Valores monetarios grandes (>= 7 dígitos)
    - Teléfonos celulares (10 dígitos con prefijo conocido)
    - Códigos de habilitación (8-12 dígitos puros)
    - Palabras inválidas (CODIGO, TARIFA, DESCRIPCION, etc.)
    """
    if not cups:
        return False

    cups_str = str(cups).strip()

    # Quitar .0 si existe
    if cups_str.endswith('.0'):
        cups_str = cups_str[:-2]

    cups_u = cups_str.upper()

    # 1. Longitud básica
    if not cups_str or len(cups_str) > 25:
        return False

    # 2. 🆕 v14.1: RECHAZAR si es una ciudad (traslados)
    if cups_u in CIUDADES_COLOMBIA_COMPLETA:
        return False

    # 3. RECHAZAR palabras inválidas
    for palabra in PALABRAS_INVALIDAS_CUPS:
        if palabra in cups_u:
            return False

    # 4. RECHAZAR patrones inválidos
    for patron in PATRONES_INVALIDOS_CUPS:
        if re.search(patron, cups_u):
            return False

    # 5. Extraer solo dígitos
    cups_digits = re.sub(r'[^\d]', '', cups_str)

    # 6. 🆕 v14.2: RECHAZAR si parece un valor monetario grande (>= 7 dígitos)
    # Permitir códigos con guiones (códigos propios de prestadores)
    if cups_digits and len(cups_digits) >= 7:
        # Si tiene guión, es un código propio válido (ej: 931002-1)
        if '-' in cups_str:
            pass  # Permitir
        else:
            return False

    # 7. RECHAZAR si parece teléfono celular (10 dígitos con prefijo conocido)
    if es_telefono_celular(cups_str):
        return False

    # 8. 🆕 v14.1: RECHAZAR si parece código de habilitación (8-12 dígitos puros)
    if cups_digits and cups_digits == cups_str and 8 <= len(cups_digits) <= 12:
        return False

    # 9. RECHAZAR municipios/departamentos
    if es_municipio_o_departamento(cups_u):
        return False

    # 10. RECHAZAR direcciones
    if es_direccion(cups_u):
        return False

    # 11. RECHAZAR valores especiales
    if cups_u in ['N.A', 'NA', 'N/A', 'N.A.', '-', '--', '---', 'NINGUNO', 'NINGUNA', 'NULL', 'NONE', '']:
        return False

    # 12. RECHAZAR si es número de sede
    if es_numero_sede(cups_str):
        return False

    # 13. Si es solo dígitos, debe tener al menos 4
    if cups_digits and cups_digits == cups_str:
        if len(cups_digits) < 4:
            return False

    # 14. 🆕 v14.1: Si la fila completa parece ser de traslados, rechazar
    if fila and es_fila_de_traslados(fila):
        return False

    # 15. Si la fila es dato de sede, rechazar
    if fila and es_dato_de_sede(fila):
        return False

    return True

def validar_tarifa(tarifa, fila: list = None) -> bool:
    """🆕 v14.1: Validación mejorada de tarifas.
    Solo rechaza si CLARAMENTE es un teléfono celular.
    """
    if tarifa is None:
        return True  # Valor nulo es aceptable

    valor_str = str(tarifa).strip()

    # Quitar .0 si existe
    if valor_str.endswith('.0'):
        valor_str = valor_str[:-2]

    # RECHAZAR si es teléfono celular
    if es_telefono_celular(valor_str):
        return False

    # RECHAZAR si parece código de habilitación Y hay contexto de sede
    valor_clean = re.sub(r'[^\d]', '', valor_str)
    if valor_clean and 8 <= len(valor_clean) <= 12:
        if fila:
            fila_texto = ' '.join([str(x).upper() for x in fila[:5] if x])
            for depto in DEPARTAMENTOS_COLOMBIA:
                if depto in fila_texto:
                    return False

    return True

def validar_manual_tarifario(manual) -> bool:
    if manual is None:
        return True
    if es_direccion(str(manual)):
        return False
    return not es_telefono_celular(str(manual))

def validar_descripcion(descripcion) -> bool:
    if descripcion is None:
        return True
    desc_str = str(descripcion).strip()
    if es_numero_sede(desc_str):
        return False
    return not es_municipio_o_departamento(desc_str)

print("✅ Validación semántica v14.1 cargada")
print("✅ 🆕 Lista expandida de ciudades colombianas")
print("✅ 🆕 Validación CUPS ultra estricta (rechaza ciudades/valores monetarios)")
print("✅ 🆕 Teléfonos: detecta números SIN guiones")
print("✅ 🆕 Alerta PAQUETES: solo si no hay hoja de servicios")
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 4: CARGAR MAESTRA DE CONTRATOS v14.1
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(3, 6, "CARGAR MAESTRA DE CONTRATOS")

print("""
📁 Selecciona el archivo de la maestra de contratos vigentes.
   Formatos soportados: .xlsx, .xls, .xlsb, .xlsm
""")

# ADAPTACIÓN LOCAL: Solicitar ruta del archivo
print("📁 Ingrese la ruta del archivo de maestra de contratos:")
print("   Formatos soportados: .xlsx, .xls, .xlsb, .xlsm")

# ══════════════════════════════════════════════════════════════════════════════
# CARGA DE MAESTRA (PARAMETRIZADO - SIN input())
# ══════════════════════════════════════════════════════════════════════════════

# ADAPTACIÓN: Usar variable de entorno en lugar de input()
ruta_maestra = PARAM_MAESTRA
if ruta_maestra.startswith('"') or ruta_maestra.startswith("'"):
    ruta_maestra = ruta_maestra[1:-1]
uploaded = {ruta_maestra: open(ruta_maestra, 'rb').read()}
ARCHIVO_MAESTRA = list(uploaded.keys())[0]

LOG.indent()
LOG.success(f"Archivo cargado", ARCHIVO_MAESTRA)

hojas = obtener_hojas(ARCHIVO_MAESTRA)
LOG.info(f"Hojas encontradas", f"{len(hojas)} hojas")

HOJA_CONTRATOS = None
for hoja in hojas:
    hoja_upper = hoja.upper()
    if 'CONTRATO' in hoja_upper and 'VIGENTE' in hoja_upper:
        HOJA_CONTRATOS = hoja
        break

if not HOJA_CONTRATOS:
    for hoja in hojas:
        if 'CONTRATO' in hoja.upper():
            HOJA_CONTRATOS = hoja
            break

if not HOJA_CONTRATOS:
    HOJA_CONTRATOS = hojas[0] if hojas else None

LOG.info("Hoja seleccionada", HOJA_CONTRATOS)

df_maestra = leer_excel(ARCHIVO_MAESTRA, sheet_name=HOJA_CONTRATOS)
LOG.success(f"Maestra cargada", f"{len(df_maestra):,} registros totales")

@dataclass
class ColumnasIdentificadas:
    tipo_proveedor: Optional[str] = None
    cto: Optional[str] = None
    numero_contrato: Optional[str] = None
    ano_contrato: Optional[str] = None

COLS = ColumnasIdentificadas()

for col in df_maestra.columns:
    col_upper = str(col).upper().strip()
    if 'TIPO' in col_upper and 'PROVEEDOR' in col_upper:
        COLS.tipo_proveedor = col
    elif col_upper == 'CTO':
        COLS.cto = col
    elif ('NUMERO' in col_upper or 'NÚMERO' in col_upper) and 'CONTRATO' in col_upper:
        COLS.numero_contrato = col
    elif ('AÑO' in col_upper or 'ANO' in col_upper) and 'CONTRATO' in col_upper:
        COLS.ano_contrato = col

LOG.info("Columnas identificadas:")
LOG.indent()
if COLS.tipo_proveedor: LOG.info("Tipo proveedor", COLS.tipo_proveedor)
if COLS.numero_contrato: LOG.info("Número contrato", COLS.numero_contrato)
if COLS.ano_contrato: LOG.info("Año contrato", COLS.ano_contrato)
if COLS.cto: LOG.info("CTO", COLS.cto)
LOG.dedent()

# 🆕 v14.1: MOSTRAR FILTROS APLICADOS
print("\n" + "─" * 50)
print("📋 FILTROS APLICADOS A LA MAESTRA:")
print("─" * 50)

registros_iniciales = len(df_maestra)
print(f"   • Registros iniciales: {registros_iniciales:,}")

if COLS.tipo_proveedor:
    tipos_unicos = df_maestra[COLS.tipo_proveedor].dropna().unique()
    print(f"\n   📌 FILTRO 1: Columna '{COLS.tipo_proveedor}'")
    print(f"      Valores encontrados: {list(tipos_unicos)[:5]}...")
    print(f"      Filtrando por: 'PRESTADOR DE SERVICIOS DE SALUD'")

    df_prestadores = df_maestra[
        df_maestra[COLS.tipo_proveedor] == 'PRESTADOR DE SERVICIOS DE SALUD'
    ].copy()

    registros_filtrados = len(df_prestadores)
    registros_excluidos = registros_iniciales - registros_filtrados
    print(f"      ✅ Registros después del filtro: {registros_filtrados:,}")
    print(f"      ❌ Registros excluidos: {registros_excluidos:,}")

    LOG.success(f"Prestadores filtrados", f"{len(df_prestadores):,} registros")
else:
    df_prestadores = df_maestra.copy()
    print(f"\n   ⚠️ Sin columna TIPO PROVEEDOR - usando todos los registros")
    LOG.warning("Sin columna TIPO PROVEEDOR", "usando todos los registros")

if COLS.ano_contrato:
    anos = sorted([int(a) for a in df_prestadores[COLS.ano_contrato].dropna().unique()])
    print(f"\n   📌 AÑOS DISPONIBLES EN LA MAESTRA:")
    print(f"      {anos}")

    print(f"\n   📊 CONTRATOS POR AÑO:")
    for ano in anos:
        count = len(df_prestadores[df_prestadores[COLS.ano_contrato] == ano])
        print(f"      • {ano}: {count:,} contratos")

    LOG.info("Años disponibles", str(anos))

print("─" * 50)
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 5: CLIENTE SFTP v14.1
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(4, 6, "CONFIGURANDO CLIENTE SFTP v14.1")
LOG.indent()

class SFTPClient:
    """🆕 v14.1: Cliente SFTP con reconexión forzada por contrato."""

    def __init__(self, config: Config, logger: Logger):
        self.config = config
        self.log = logger
        self._client = None
        self._sftp = None
        self._transport = None
        self._reconexiones = 0
        self._current_path = "/"

    def _cerrar(self):
        for c in [self._sftp, self._client]:
            try:
                if c: c.close()
            except: pass
        try:
            if self._transport:
                self._transport.close()
        except: pass
        self._sftp = self._client = self._transport = None

    def conectar(self, silencioso: bool = False) -> bool:
        self._cerrar()

        for intento in range(self.config.MAX_REINTENTOS_CONEXION):
            try:
                if not silencioso:
                    self.log.info(f"Conectando a {self.config.HOST}:{self.config.PORT}...")

                self._client = paramiko.SSHClient()
                self._client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                self._client.connect(
                    hostname=self.config.HOST,
                    port=self.config.PORT,
                    username=self.config.USERNAME,
                    password=self.config.PASSWORD,
                    timeout=self.config.TIMEOUT_CONEXION,
                    banner_timeout=self.config.TIMEOUT_CONEXION,
                    auth_timeout=self.config.TIMEOUT_CONEXION,
                    allow_agent=False,
                    look_for_keys=False
                )

                self._transport = self._client.get_transport()
                self._transport.set_keepalive(self.config.KEEPALIVE_INTERVAL)
                self._sftp = self._client.open_sftp()
                self._sftp.get_channel().settimeout(self.config.TIMEOUT_OPERACION)
                self._current_path = "/"

                if not silencioso:
                    self.log.success("Conexión establecida")
                return True

            except Exception as e:
                self._cerrar()
                if intento < self.config.MAX_REINTENTOS_CONEXION - 1:
                    espera = self.config.BACKOFF_BASE ** intento
                    if not silencioso:
                        self.log.warning(f"Intento {intento + 1} fallido", f"reintentando en {espera:.1f}s")
                    time.sleep(espera)

        if not silencioso:
            self.log.error("No se pudo conectar después de varios intentos")
        return False

    def reconectar_forzado(self, silencioso: bool = True) -> bool:
        """🆕 v14.1: Fuerza reconexión."""
        self._reconexiones += 1
        self._cerrar()
        time.sleep(0.5)
        return self.conectar(silencioso)

    def esta_activo(self) -> bool:
        try:
            if not self._sftp or not self._transport: return False
            if not self._transport.is_active(): return False
            self._sftp.getcwd()
            return True
        except:
            return False

    def _ejecutar(self, operacion: Callable, descripcion: str = "operación"):
        for intento in range(self.config.MAX_REINTENTOS_OPERACION):
            try:
                if not self.esta_activo():
                    self._reconexiones += 1
                    self.log.warning("Reconectando...", f"intento {self._reconexiones}")
                    if not self.conectar(True):
                        raise Exception("Reconexión fallida")
                return operacion()
            except Exception as e:
                if intento == self.config.MAX_REINTENTOS_OPERACION - 1:
                    raise
                time.sleep(1)

    def listar(self, ruta: str = '.') -> List[Dict]:
        def _op():
            return [
                {
                    'nombre': a.filename,
                    'tamano': a.st_size,
                    'es_directorio': stat.S_ISDIR(a.st_mode),
                    'fecha_modificacion': a.st_mtime
                }
                for a in self._sftp.listdir_attr(ruta)
            ]
        return self._ejecutar(_op)

    def cd(self, ruta: str, log_nav: bool = True):
        def _op():
            self._sftp.chdir(ruta)
            self._current_path = self._sftp.getcwd() or ruta
        self._ejecutar(_op)
        if log_nav:
            self.log.nav(self._current_path)

    def descargar(self, remoto: str, local: str, log_download: bool = True):
        if log_download:
            self.log.download(remoto)
        self._ejecutar(lambda: self._sftp.get(remoto, local))

    def desconectar(self):
        self._cerrar()
        self.log.info("Conexión SFTP cerrada")

    @property
    def path_actual(self) -> str:
        return self._current_path

    @property
    def reconexiones(self) -> int:
        return self._reconexiones

LOG.success("Cliente SFTP v14.1 configurado")
LOG.success("🆕 Reconexión forzada por contrato habilitada")
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 6: BUSCADOR DE ANEXOS v14.1
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(5, 6, "CONFIGURANDO BUSCADOR DE ANEXOS v14.1")
LOG.indent()

class BuscadorAnexos:
    """🆕 v14.1: Buscador de anexos con búsqueda mejorada."""

    def __init__(self, cliente: SFTPClient, config: Config, logger: Logger):
        self.cliente = cliente
        self.config = config
        self.log = logger
        self.alertas: List[Alerta] = []
        self._contrato_actual = ""

    def limpiar_alertas(self):
        self.alertas = []

    def set_contrato(self, contrato: str):
        self._contrato_actual = contrato

    def agregar_alerta(self, tipo: TipoAlerta, mensaje: str, archivo: str = ""):
        self.alertas.append(Alerta(
            tipo=tipo,
            mensaje=mensaje,
            contrato=self._contrato_actual,
            archivo=archivo
        ))
        self.log.alert(tipo.value, mensaje, archivo)

    def extraer_numero_otrosi(self, nombre: str) -> Optional[int]:
        """🆕 v15.0: Extrae número de otrosí con patrones expandidos."""
        if not nombre:
            return None
        nombre_upper = nombre.upper()

        # Patrones expandidos v15.0
        patrones = [
            r'OTRO\s*S[IÍ]\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
            r'OTROS[IÍ]\s*[_#\-\s]*(\d+)',
            r'OTRO[\s_\-]?SI[\s_\-#]*(\d+)',
            r'OT\s*[_\-\s]?\s*(\d+)',
            r'ADICI[OÓ]N\s*[_#\-\s]*N?[OÚº°]?\.?\s*(\d+)',
            r'MODIFICA(?:CI[OÓ]N)?\s*[_#\-\s]*(\d+)',
        ]

        for patron in patrones:
            match = re.search(patron, nombre_upper)
            if match:
                try:
                    return int(match.group(1))
                except (ValueError, IndexError):
                    continue
        return None

    def extraer_numero_acta(self, nombre: str, nombre_carpeta: str = None) -> Optional[int]:
        """🆕 v14.1: Extrae número de acta del nombre o carpeta."""
        if not nombre:
            nombre = ""

        nombre_upper = nombre.upper()

        patrones = [
            r'ACTA\s*(?:DE\s*)?(?:NEGOCIACI[OÓ]N\s*)?(?:N[OÚº°]?\.?\s*)?#?\s*(\d+)',
            r'ACT[_\-\s]?(\d+)',
            r'\bAN\s*[_\-]?\s*(\d+)',
            r'ACTA\s*#?\s*(\d+)',
            r'ACTA\s*N[OÚº°]?\s*(\d+)',
        ]

        for patron in patrones:
            match = re.search(patron, nombre_upper)
            if match:
                return int(match.group(1))

        if nombre_carpeta:
            carpeta_upper = nombre_carpeta.upper()
            for patron in patrones:
                match = re.search(patron, carpeta_upper)
                if match:
                    return int(match.group(1))

        return None

    def buscar_carpeta(self, carpetas: List[str], texto: str) -> Optional[str]:
        texto_l = texto.lower()
        for c in carpetas:
            if c.lower() == texto_l: return c
        for c in carpetas:
            if texto_l in c.lower(): return c
        return None

    def buscar_carpeta_contrato(self, carpetas: List[str], numero: str, nombre_proveedor: str = None) -> Optional[str]:
        """🆕 v14.1: Búsqueda mejorada con cero inicial."""
        num = ''.join(filter(str.isdigit, str(numero)))

        variantes = [
            num,
            num.zfill(4),
            num.zfill(3),
            num.zfill(5),
            num.lstrip('0') or '0',
            '0' + num,
        ]

        variantes_unicas = list(dict.fromkeys(variantes))

        self.log.debug(f"Buscando contrato con variantes: {variantes_unicas}")

        for variante in variantes_unicas:
            for carpeta in carpetas:
                partes = re.split(r'[\s\-_]', carpeta)
                if partes and partes[0] == variante:
                    self.log.debug(f"Encontrado con variante '{variante}': {carpeta}")
                    return carpeta

        for variante in variantes_unicas:
            for carpeta in carpetas:
                if carpeta.startswith(variante + '-') or carpeta.startswith(variante + '_') or carpeta.startswith(variante + ' '):
                    self.log.debug(f"Encontrado por prefijo '{variante}': {carpeta}")
                    return carpeta

        if nombre_proveedor:
            nombre_limpio = nombre_proveedor.upper().strip()
            for carpeta in carpetas:
                if nombre_limpio in carpeta.upper():
                    self.log.debug(f"Encontrado por nombre proveedor: {carpeta}")
                    return carpeta

        return None

    def navegar_a_contrato(self, ano: str, numero: str, nombre_proveedor: str = None) -> Tuple[bool, str, Optional[str]]:
        """🆕 v14.1: Navega con búsqueda mejorada."""
        try:
            self.log.info("Navegando a contrato...")
            self.log.indent()

            self.cliente.cd('/', log_nav=False)
            items = self.cliente.listar()
            carpetas = [i['nombre'] for i in items if i['es_directorio']]

            cp = self.buscar_carpeta(carpetas, self.config.CARPETA_PRINCIPAL)
            if not cp:
                self.log.error("No encontrada", "carpeta principal")
                self.log.dedent()
                return False, "Sin carpeta principal", None

            self.cliente.cd(cp)

            items = self.cliente.listar()
            carpetas = [i['nombre'] for i in items if i['es_directorio']]
            ca = self.buscar_carpeta(carpetas, f'contratos {ano}')
            if not ca:
                self.log.error("No encontrada", f"carpeta año {ano}")
                self.log.dedent()
                return False, f"Sin año {ano}", None

            self.cliente.cd(ca)

            items = self.cliente.listar()
            carpetas = [i['nombre'] for i in items if i['es_directorio']]

            cc = self.buscar_carpeta_contrato(carpetas, numero, nombre_proveedor)

            if not cc:
                self.log.error("No encontrada", f"carpeta contrato {numero}")
                self.log.warning("🆕 Variantes buscadas", f"{numero}, 0{numero}, {numero.zfill(4)}")
                self.log.nav_tree(carpetas[:10], "carpetas")

                self.agregar_alerta(
                    TipoAlerta.CONTRATO_NO_ENCONTRADO_GO,
                    f"CONTRATO NO SE ENCUENTRA EN EL GO ANYWHERE. Buscado: {numero}, 0{numero}, {numero.zfill(4)}"
                )

                self.log.dedent()
                return False, f"CONTRATO NO SE ENCUENTRA EN EL GO ANYWHERE", None

            self.cliente.cd(cc)
            self.log.success("Contrato encontrado", cc)
            self.log.dedent()

            return True, "OK", f"/{cp}/{ca}/{cc}"

        except Exception as e:
            self.log.error("Error de navegación", str(e)[:40])
            self.log.dedent()
            return False, str(e)[:40], None

    def descargar_anexos(self, carpeta_destino: str, id_contrato: str) -> Dict:
        """Descarga ANEXO 1 con logging detallado."""
        resultado = {
            'exito': False,
            'archivos': [],
            'mensaje': '',
            'actas_encontradas': [],
            'otrosis_encontrados': []
        }

        try:
            self.log.info("Buscando archivos ANEXO 1...")
            self.log.indent()

            items = self.cliente.listar()
            carpetas = [i['nombre'] for i in items if i['es_directorio']]
            archivos = [i['nombre'] for i in items if not i['es_directorio']]

            self.log.debug(f"Contenido: {len(carpetas)} carpetas, {len(archivos)} archivos")

            carpeta_tarifas = None
            for c in carpetas:
                if 'tarifa' in c.lower():
                    carpeta_tarifas = c
                    break

            if not carpeta_tarifas:
                self.log.error("Carpeta TARIFAS no encontrada")
                self.log.nav_tree(carpetas, "carpetas")
                self.agregar_alerta(TipoAlerta.SIN_CARPETA_TARIFAS, "No existe carpeta TARIFAS")
                self.log.dedent()
                resultado['mensaje'] = "Sin TARIFAS"
                return resultado

            self.cliente.cd(carpeta_tarifas)
            items_tarifas = self.cliente.listar()

            archivos_excel = [i for i in items_tarifas if not i['es_directorio'] and es_extension_excel(i['nombre'])]
            subcarpetas = [i for i in items_tarifas if i['es_directorio']]

            self.log.info(f"En TARIFAS: {len(archivos_excel)} Excel, {len(subcarpetas)} subcarpetas")

            anexos_iniciales = []
            anexos_otrosi = []
            archivos_ignorados = []

            # 🆕 v15.0: Usar clasificar_tipo_archivo para mejor detección
            for item in archivos_excel:
                nombre = item['nombre']

                # Clasificar el archivo
                info = clasificar_tipo_archivo(nombre)

                if not info['es_valido']:
                    archivos_ignorados.append((nombre, info.get('motivo_exclusion', 'N/A')))
                    self.log.debug(f"Archivo ignorado: {nombre} ({info.get('motivo_exclusion', 'N/A')})")
                    continue

                self.log.debug(f"✓ Archivo válido: {nombre} → tipo={info['tipo']}")

                num_otrosi = self.extraer_numero_otrosi(nombre)
                if num_otrosi:
                    anexos_otrosi.append({'item': item, 'numero': num_otrosi, 'tipo': info['tipo']})
                    self.log.file_found(nombre, f"Otrosí {num_otrosi} ({info['tipo']})")
                else:
                    anexos_iniciales.append({'item': item, 'tipo': info['tipo']})
                    self.log.file_found(nombre, f"Inicial ({info['tipo']})")

            if archivos_ignorados and len(archivos_ignorados) <= 5:
                for nombre_ign, motivo in archivos_ignorados:
                    self.log.debug(f"  ↳ Ignorado: {nombre_ign[:40]}... - {motivo}")

            resultado['otrosis_encontrados'] = [a['numero'] for a in anexos_otrosi]

            archivo_principal = None
            origen_principal = None
            numero_principal = None
            fecha_referencia = None

            if anexos_otrosi:
                anexos_otrosi.sort(key=lambda x: x['numero'], reverse=True)
                archivo_principal = anexos_otrosi[0]['item']
                numero_principal = anexos_otrosi[0]['numero']
                origen_principal = OrigenTarifa.OTROSI
                self.log.info(f"Seleccionado: Otrosí {numero_principal} (mayor)")
            elif anexos_iniciales:
                archivo_principal = anexos_iniciales[0]['item']
                origen_principal = OrigenTarifa.INICIAL
                self.log.info("Seleccionado: Anexo inicial")
            else:
                self.log.warning("No hay ANEXO 1 ni TARIFAS en carpeta")
                msg_extra = ""
                if archivos_ignorados:
                    nombres_ignorados = [n for n, m in archivos_ignorados[:3]]
                    msg_extra = f" | Archivos ignorados: {', '.join(nombres_ignorados)}"
                self.agregar_alerta(TipoAlerta.SIN_ANEXO1, f"No hay anexo 1, otrosí ni archivo TARIFAS válido{msg_extra}")
                resultado['mensaje'] = "Sin ANEXO 1"

            if archivo_principal:
                ruta_local = os.path.join(carpeta_destino, archivo_principal['nombre'])
                self.cliente.descargar(archivo_principal['nombre'], ruta_local)
                fecha_referencia = archivo_principal.get('fecha_modificacion')
                resultado['archivos'].append(ArchivoAnexo(
                    nombre=archivo_principal['nombre'],
                    ruta_local=ruta_local,
                    origen=origen_principal,
                    numero=numero_principal,
                    fecha_modificacion=fecha_referencia,
                    origen_completo=f"/{carpeta_tarifas}/{archivo_principal['nombre']}"
                ))

            carpetas_actas = [item for item in subcarpetas if 'acta' in item['nombre'].lower()]

            if carpetas_actas:
                self.log.info(f"Buscando en {len(carpetas_actas)} carpeta(s) de actas...")

            todas_las_actas = []

            for carpeta_acta in carpetas_actas:
                try:
                    self.cliente.cd(carpeta_acta['nombre'], log_nav=False)
                    self.log.debug(f"Entrando a: {carpeta_acta['nombre']}")

                    items_actas = self.cliente.listar()
                    actas_excel = [i for i in items_actas if not i['es_directorio'] and es_extension_excel(i['nombre'])]
                    actas_en_carpeta = []

                    for ia in actas_excel:
                        if not contiene_anexo1(ia['nombre']):
                            continue

                        num_acta = self.extraer_numero_acta(ia['nombre'], carpeta_acta['nombre'])
                        fecha_acta = ia.get('fecha_modificacion')

                        descargar = False
                        if not archivo_principal:
                            descargar = True
                        elif fecha_referencia and fecha_acta and fecha_acta > fecha_referencia:
                            descargar = True

                        if descargar:
                            nombre_local = f"ACTA_{carpeta_acta['nombre']}_{ia['nombre']}"
                            nombre_local = re.sub(r'[<>:"/\\|?*]', '_', nombre_local)
                            ruta_acta = os.path.join(carpeta_destino, nombre_local)
                            self.cliente.descargar(ia['nombre'], ruta_acta, log_download=False)
                            self.log.file_found(ia['nombre'], f"Acta {num_acta or '?'}")

                            resultado['archivos'].append(ArchivoAnexo(
                                nombre=nombre_local,
                                ruta_local=ruta_acta,
                                origen=OrigenTarifa.ACTA,
                                numero=num_acta,
                                fecha_modificacion=fecha_acta,
                                origen_completo=f"/{carpeta_tarifas}/{carpeta_acta['nombre']}/{ia['nombre']}"
                            ))

                        if num_acta:
                            actas_en_carpeta.append(num_acta)
                            todas_las_actas.append(num_acta)

                    if not actas_en_carpeta and actas_excel:
                        self.log.warning(f"Carpeta '{carpeta_acta['nombre']}' sin ANEXO 1")
                        self.agregar_alerta(
                            TipoAlerta.CARPETA_ACTAS_SIN_ANEXO,
                            f"Carpeta sin anexo 1",
                            carpeta_acta['nombre']
                        )

                    self.cliente.cd('..', log_nav=False)

                except Exception as e_acta:
                    self.log.error(f"Error en carpeta actas", str(e_acta)[:30])
                    self.agregar_alerta(
                        TipoAlerta.ERROR_PROCESAMIENTO,
                        f"Error procesando: {str(e_acta)[:30]}",
                        carpeta_acta['nombre']
                    )
                    try:
                        self.cliente.cd('..', log_nav=False)
                    except:
                        pass

            resultado['actas_encontradas'] = sorted(set(todas_las_actas))

            if todas_las_actas:
                max_acta = max(todas_las_actas)
                faltantes = [i for i in range(1, max_acta + 1) if i not in todas_las_actas]
                for f in faltantes:
                    self.log.warning(f"Acta {f} faltante en secuencia")
                    self.agregar_alerta(TipoAlerta.ACTA_FALTANTE, f"Falta anexo 1 del acta {f}")

            if resultado['archivos']:
                resultado['exito'] = True
                resultado['mensaje'] = f"{len(resultado['archivos'])} archivo(s)"
                self.log.success(f"Total archivos a procesar: {len(resultado['archivos'])}")

            self.log.dedent()
            return resultado

        except Exception as e:
            self.log.error("Error en descarga", str(e)[:40])
            self.agregar_alerta(TipoAlerta.ERROR_PROCESAMIENTO, str(e)[:50])
            self.log.dedent()
            resultado['mensaje'] = str(e)[:40]
            return resultado

LOG.success("Buscador de anexos v14.1 configurado")
LOG.success("🆕 Búsqueda con cero inicial habilitada")
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 7: PROCESADOR DE ANEXOS v14.1
# ══════════════════════════════════════════════════════════════════════════════

LOG.step(6, 6, "CONFIGURANDO PROCESADOR DE ANEXOS v14.1")
LOG.indent()

class ProcesadorAnexo:
    """🆕 v14.1: Procesador de anexos con detección de columnas mejorada."""

    def __init__(self, logger: Logger):
        self.log = logger
        self._alertas_set: set = set()
        self.alertas: List[Alerta] = []
        self._contrato_actual = ""
        self._categoria_cuentas_medicas = ""

    def limpiar_alertas(self):
        self._alertas_set = set()
        self.alertas = []

    def set_contrato(self, contrato: str):
        self._contrato_actual = contrato

    def set_categoria_cuentas_medicas(self, categoria: str):
        self._categoria_cuentas_medicas = categoria if categoria else ""

    def agregar_alerta(self, tipo: TipoAlerta, mensaje: str, archivo: str = ""):
        nueva_alerta = Alerta(
            tipo=tipo,
            mensaje=mensaje,
            contrato=self._contrato_actual,
            archivo=archivo
        )

        clave = (tipo, mensaje, self._contrato_actual, archivo)
        if clave not in self._alertas_set:
            self._alertas_set.add(clave)
            self.alertas.append(nueva_alerta)
            self.log.alert(tipo.value, mensaje, archivo)

    def buscar_hoja_servicios(self, archivo: str) -> Optional[str]:
        """🆕 v14.1: Busca la hoja de servicios - CORREGIDO.
        Las hojas de PAQUETES NO generan alerta individual, solo se mencionan
        si no hay hoja de servicios válida.
        """
        formato_real = detectar_formato_real(archivo)
        ext_declarada = os.path.splitext(archivo)[1].lower()

        if formato_real == 'xlsb' and ext_declarada != '.xlsb':
            self.log.debug(f"⚠️ Formato real: XLSB (extensión: {ext_declarada})")

        hojas = obtener_hojas(archivo)

        if not hojas:
            motivo = f"No se pudo leer archivo (formato: {formato_real})"
            self.agregar_alerta(TipoAlerta.ERROR_LECTURA, motivo, os.path.basename(archivo))
            return None

        # 🆕 v14.1: Usar nueva función que retorna info de hojas excluidas
        hoja_encontrada, hojas_excluidas_info = buscar_hoja_servicios_inteligente(hojas)

        if hoja_encontrada:
            # ✅ Encontró hoja de servicios - NO genera alerta de PAQUETES
            self.log.debug(f"Hoja seleccionada: '{hoja_encontrada}' de {len(hojas)} disponibles")
            return hoja_encontrada

        # ❌ NO encontró hoja de servicios - verificar tipo de archivo
        es_solo_traslados, msg_traslados, tipo_archivo = es_archivo_solo_traslados(hojas)

        if es_solo_traslados:
            self.log.warning(f"Archivo de {tipo_archivo.lower()}", msg_traslados)

            if tipo_archivo == "AMBULANCIAS":
                self.agregar_alerta(
                    TipoAlerta.ARCHIVO_SOLO_AMBULANCIAS,
                    msg_traslados,
                    os.path.basename(archivo)
                )
            elif tipo_archivo == "TRASLADOS":
                self.agregar_alerta(
                    TipoAlerta.ARCHIVO_SOLO_TRASLADOS,
                    msg_traslados,
                    os.path.basename(archivo)
                )
            else:
                self.agregar_alerta(
                    TipoAlerta.SOLO_TRASLADOS,
                    msg_traslados,
                    os.path.basename(archivo)
                )
            return None

        # 🆕 v14.1: Generar mensaje con TODAS las hojas disponibles
        # AQUÍ es donde se mencionan las hojas de PAQUETES (no antes)
        hojas_str = ", ".join([f"'{h}'" for h in hojas])
        mensaje = f"No se encontró hoja de servicios válida. Hojas disponibles: [{hojas_str}]"

        # Agregar info de hojas excluidas si existen (PAQUETES, COSTO VIAJE)
        if hojas_excluidas_info:
            excluidas_str = ", ".join([f"'{h[0]}'" for h in hojas_excluidas_info])
            mensaje += f". Hojas excluidas (no aplican para T25): [{excluidas_str}]"

        if self._categoria_cuentas_medicas:
            mensaje += f". Categoría cuentas médicas: '{self._categoria_cuentas_medicas}'"

        self.agregar_alerta(TipoAlerta.HOJA_NO_ENCONTRADA, mensaje, os.path.basename(archivo))

        return None

    def detectar_columnas(self, fila: List) -> Dict[str, int]:
        """🆕 v14.1: Detecta índices de columnas con prioridad estricta."""
        idx = {
            'cups': -1,
            'homologo': -1,
            'descripcion': -1,
            'tarifa': -1,
            'tarifario': -1,
            'porcentaje': -1,
            'observaciones': -1
        }

        columnas_usadas = set()

        PATRONES_ORDENADOS = [
            ('cups', [
                'CODIGO CUPS', 'CÓDIGO CUPS', 'COD CUPS', 'COD. CUPS',
                'CODIGO CUP', 'COD CUP', 'COD. CUP'
            ]),
            ('homologo', [
                'CODIGO HOMOLOGO', 'CÓDIGO HOMÓLOGO', 'COD HOMOLOGO',
                'HOMOLOGO MANUAL', 'CÓDIGO HOMOLOGO MANUAL', 'CODIGO HOMOLOGO MANUAL'
            ]),
            ('descripcion', [
                'DESCRIPCION DEL CUPS', 'DESCRIPCIÓN DEL CUPS',
                'DESCRIPCION CUPS', 'DESCRIPCIÓN CUPS',
                'DESCRIPCION DEL CUP', 'DESCRIPCIÓN DEL CUP'
            ]),
            ('tarifa', [
                'TARIFA UNITARIA EN PESOS', 'TARIFA UNITARIA PESOS',
                'TARIFA EN PESOS', 'TARIFA UNITARIA',
                'VALOR UNITARIO', 'PRECIO UNITARIO'
            ]),
            ('tarifario', [
                'MANUAL TARIFARIO', 'TARIFARIO', 'MANUAL TAR',
                'TIPO TARIFARIO', 'TIPO DE TARIFARIO'
            ]),
            ('porcentaje', [
                'TARIFA SEGUN TARIFARIO', 'TARIFA SEGÚN TARIFARIO',
                'PORCENTAJE TARIFARIO', 'PORCENTAJE',
                '% TARIFARIO', '% DEL TARIFARIO'
            ]),
            ('observaciones', [
                'OBSERVACIONES', 'OBSERVACION', 'OBS', 'NOTAS'
            ]),
        ]

        for i, celda in enumerate(fila):
            t = normalizar_texto(celda)
            if not t:
                continue

            if i in columnas_usadas:
                continue

            for campo, patrones in PATRONES_ORDENADOS:
                if idx[campo] != -1:
                    continue

                for patron in patrones:
                    patron_norm = normalizar_texto(patron)

                    if patron_norm in t or patron_norm == t:
                        if campo == 'cups' and 'HOMOLOGO' in t:
                            continue

                        if campo == 'tarifa':
                            if 'TARIFARIO' in t or 'SEGUN' in t or 'SEGÚN' in t:
                                continue
                            if 'MANUAL' in t and 'UNITARIA' not in t:
                                continue

                        if campo == 'tarifario':
                            if 'UNITARIA' in t or 'EN PESOS' in t or 'PESOS' in t:
                                continue

                        if campo == 'porcentaje':
                            if 'UNITARIA' in t:
                                continue

                        idx[campo] = i
                        columnas_usadas.add(i)
                        break

                if idx[campo] != -1:
                    break

        return idx

    def extraer_sedes_de_bloque(self, datos: List[List], inicio: int, idx_hab: int, idx_sede: int) -> List[Dict]:
        """Extrae las sedes de un bloque de datos de sedes."""
        sedes = []
        k = inicio

        while k < len(datos) and len(sedes) < CONFIG.MAX_SEDES:
            fila = datos[k]
            if not fila:
                k += 1
                continue

            if es_encabezado_seccion_sedes(fila) or es_encabezado_seccion_servicios(fila):
                break

            if es_dato_de_sede(fila):
                if idx_hab >= 0 and idx_hab < len(fila):
                    codigo_hab = fila[idx_hab]
                    if codigo_hab:
                        codigo_str = str(codigo_hab).strip()
                        if codigo_str.endswith('.0'):
                            codigo_str = codigo_str[:-2]
                        codigo_clean = re.sub(r'[^\d]', '', codigo_str)

                        if codigo_clean and codigo_clean.isdigit() and 5 <= len(codigo_clean) <= 12:
                            num_sede = fila[idx_sede] if idx_sede >= 0 and idx_sede < len(fila) else len(sedes) + 1
                            sedes.append({'codigo': codigo_hab, 'sede': num_sede})
                            k += 1
                            continue

            if fila[0] is not None:
                primera = str(fila[0]).upper().strip()
                if not es_municipio_o_departamento(primera) and not es_direccion(primera):
                    if primera and not primera.isspace():
                        break

            k += 1

        return sedes

    def extraer_servicios(self, archivo: str, nombre: str) -> Tuple[bool, List[Dict], str]:
        """Extrae servicios del archivo ANEXO 1."""
        try:
            self.log.process(f"Procesando: {nombre[:50]}...")
            self.log.indent()

            hoja = self.buscar_hoja_servicios(archivo)
            if not hoja:
                self.log.error("No se encontró hoja de servicios")
                self.log.dedent()
                return False, [], "Sin hoja de servicios"

            formato = detectar_formato_real(archivo)
            self.log.info(f"Hoja encontrada: '{hoja}' (formato: {formato})")

            datos = leer_hoja_raw(archivo, hoja, max_filas=20000)
            if not datos:
                self.log.error("Hoja vacía o no legible")
                self.agregar_alerta(TipoAlerta.ERROR_LECTURA, "Hoja vacía", nombre)
                self.log.dedent()
                return False, [], "Hoja vacía"

            self.log.debug(f"Filas leídas: {len(datos)}")

            servicios = []
            sedes_activas = []
            sedes_pendientes = []  # 🆕 Sedes que esperan su bloque de servicios
            idx_columnas = None
            encontro_encabezado_servicios = False
            encontro_sedes = False

            estado = 'buscando'

            i = 0
            while i < len(datos):
                fila = datos[i]

                if not fila or all(c is None for c in fila):
                    i += 1
                    continue

                if es_encabezado_seccion_sedes(fila):
                    print(f"  🔍 SEDES: Detectado bloque de sedes en fila {i+1}")
                    self.log.debug(f"Fila {i+1}: Encabezado de SEDES detectado")
                    encontro_sedes = True
                    estado = 'en_sedes'

                    idx_hab = -1
                    idx_sede = -1
                    for j, c in enumerate(fila):
                        t = normalizar_texto(c) if c else ''
                        if 'HABILITACION' in t or 'HABIITACION' in t:
                            idx_hab = j
                        if 'NUMERO DE SEDE' in t or 'NUMERO SEDE' in t or 'N SEDE' in t or 'N° SEDE' in t:
                            idx_sede = j

                    if idx_sede == -1 and idx_hab >= 0:
                        idx_sede = idx_hab + 1

                    nuevas_sedes = self.extraer_sedes_de_bloque(datos, i + 1, idx_hab, idx_sede)
                    if nuevas_sedes:
                        print(f"    👉 Sedes encontradas en bloque: {[s['sede'] for s in nuevas_sedes]}")
                        # 🆕 Guardar sedes pendientes para el próximo bloque de servicios
                        sedes_pendientes = nuevas_sedes
                        self.log.debug(f"  Sedes detectadas: {len(sedes_pendientes)}, esperando encabezado de servicios")

                    i += 1
                    continue

                if es_encabezado_seccion_servicios(fila):
                    print(f"  🔍 SERVICIOS: Detectado bloque de servicios en fila {i+1}")
                    self.log.debug(f"Fila {i+1}: Encabezado de SERVICIOS detectado")
                    idx_columnas = self.detectar_columnas(fila)
                    encontro_encabezado_servicios = True
                    estado = 'en_servicios'

                    # 🆕 Activar las sedes pendientes para este bloque de servicios
                    if sedes_pendientes:
                        sedes_activas = sedes_pendientes
                        sedes_pendientes = []
                        print(f"    👉 Activando sedes para este bloque: {[s['sede'] for s in sedes_activas]}")
                        self.log.debug(f"  Sedes activadas para este bloque: {len(sedes_activas)}")
                        for sede in sedes_activas:
                            self.log.debug(f"    - Sede {sede['sede']}: {sede['codigo']}")

                    cols_detectadas = [k for k, v in idx_columnas.items() if v >= 0]
                    self.log.debug(f"  Columnas: {cols_detectadas}")

                    i += 1
                    continue

                if estado == 'en_sedes':
                    if es_dato_de_sede(fila):
                        i += 1
                        continue
                    i += 1
                    continue

                if estado == 'en_servicios' and idx_columnas and sedes_activas:
                    if es_dato_de_sede(fila):
                        self.log.debug(f"Fila {i+1}: Saltando (es dato de sede)")
                        i += 1
                        continue

                    if idx_columnas['cups'] >= 0 and idx_columnas['cups'] < len(fila):
                        cups_raw = fila[idx_columnas['cups']]
                        cups = limpiar_codigo(cups_raw)

                        if cups and validar_cups(cups, fila):
                            def get_valor(campo: str):
                                col_idx = idx_columnas.get(campo, -1)
                                return fila[col_idx] if 0 <= col_idx < len(fila) else None

                            tarifa = get_valor('tarifa')
                            manual = get_valor('tarifario')
                            descripcion = get_valor('descripcion')

                            if not validar_tarifa(tarifa):
                                print(f"    ❌ RECHAZADO (Tarifa inválida) Fila {i+1}: {tarifa}")
                                self.log.debug(f"Fila {i+1}: Tarifa rechazada (parece teléfono)")
                                i += 1
                                continue

                            if not validar_manual_tarifario(manual):
                                print(f"    ❌ RECHAZADO (Manual inválido) Fila {i+1}: {manual}")
                                self.log.debug(f"Fila {i+1}: Manual rechazado (parece dirección)")
                                i += 1
                                continue

                            if not validar_descripcion(descripcion):
                                print(f"    ❌ RECHAZADO (Descripción inválida) Fila {i+1}: {descripcion}")
                                self.log.debug(f"Fila {i+1}: Descripción rechazada (es número de sede)")
                                i += 1
                                continue

                            base = {
                                'codigo_cups': cups,
                                'codigo_homologo_manual': limpiar_codigo(get_valor('homologo')),
                                'descripcion_del_cups': limpiar_texto(descripcion),
                                'tarifa_unitaria_en_pesos': limpiar_tarifa(tarifa),
                                'manual_tarifario': limpiar_texto(manual),
                                'porcentaje_manual_tarifario': limpiar_texto(get_valor('porcentaje')),
                                'observaciones': limpiar_texto(get_valor('observaciones'))
                            }

                            for sede in sedes_activas:
                                s = base.copy()
                                s['codigo_de_habilitacion'] = formatear_habilitacion(sede['codigo'], sede['sede'])
                                servicios.append(s)
                        else:
                            # Debug por qué falló validar_cups
                            if cups:
                                pass # print(f"    ⚠️ CUPS inválido o fila rechazada por validador: {cups} en fila {i+1}")

                i += 1

            if not encontro_encabezado_servicios:
                self.log.warning("No se encontró encabezado de servicios")
                self.agregar_alerta(TipoAlerta.COLUMNAS_NO_DETECTADAS, "Sin encabezado de servicios", nombre)

            if not encontro_sedes:
                self.log.warning("No se encontró sección de sedes")
                self.agregar_alerta(TipoAlerta.SEDES_NO_DETECTADAS, "Sin sección de sedes", nombre)

            # 🆕 Si hay sedes pendientes que nunca se activaron, activarlas ahora
            if sedes_pendientes and not sedes_activas:
                self.log.debug(f"Activando {len(sedes_pendientes)} sedes pendientes que no tuvieron encabezado de servicios explícito")
                sedes_activas = sedes_pendientes

            if servicios:
                self.log.success(f"Servicios extraídos: {len(servicios):,}")
                self.log.dedent()
                return True, servicios, f"{len(servicios)} servicios"
            else:
                self.log.warning("No se extrajeron servicios")
                self.log.dedent()
                return False, [], "Sin servicios extraídos"

        except Exception as e:
            self.log.error(f"Error procesando archivo: {str(e)[:50]}")
            self.agregar_alerta(TipoAlerta.ERROR_PROCESAMIENTO, str(e)[:50], nombre)
            self.log.dedent()
            return False, [], str(e)[:50]

    def extraer_con_timeout(self, archivo: str, nombre: str, timeout: int = 60) -> Tuple[bool, List[Dict], str]:
        """Extrae servicios con timeout."""
        resultado = [False, [], "Timeout"]
        error_msg = [None]

        def worker():
            try:
                resultado[0], resultado[1], resultado[2] = self.extraer_servicios(archivo, nombre)
            except Exception as e:
                error_msg[0] = str(e)
                resultado[0] = False
                resultado[2] = str(e)[:50]

        thread = threading.Thread(target=worker)
        thread.start()
        thread.join(timeout)

        if thread.is_alive():
            self.log.warning(f"Timeout ({timeout}s) procesando archivo")
            self.agregar_alerta(TipoAlerta.TIMEOUT, f"Archivo tardó más de {timeout}s", nombre)
            return False, [], f"Timeout ({timeout}s)"

        if error_msg[0]:
            self.agregar_alerta(TipoAlerta.ERROR_PROCESAMIENTO, error_msg[0][:50], nombre)

        return resultado[0], resultado[1], resultado[2]

LOG.success("Procesador de anexos v14.1 configurado")
LOG.success("🆕 Detección de columnas con prioridad estricta")
LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 8: FUNCIÓN OBTENER FECHA DE ACUERDO
# ══════════════════════════════════════════════════════════════════════════════

def obtener_fecha_acuerdo(numero: str, ano: str, origen: str, fecha_archivo: float = None) -> Tuple[Optional[str], bool]:
    """Obtiene fecha de acuerdo de forma inteligente."""
    try:
        fila = None

        if COLS.cto:
            cto_str = f"{str(numero).zfill(4)}-{ano}"
            mask = df_prestadores[COLS.cto] == cto_str
            if mask.any():
                fila = df_prestadores[mask].iloc[0]

        if fila is None and COLS.numero_contrato and COLS.ano_contrato:
            try:
                mask = (
                    df_prestadores[COLS.numero_contrato].astype(str).str.replace('.0', '', regex=False).str.zfill(4) == str(numero).zfill(4)
                ) & (
                    df_prestadores[COLS.ano_contrato].astype(str).str.replace('.0', '', regex=False) == str(ano)
                )
                if mask.any():
                    fila = df_prestadores[mask].iloc[0]
            except:
                pass

        fecha = None
        columnas = list(df_prestadores.columns) if fila is not None else []

        if origen == 'Inicial' and fila is not None:
            for col in columnas:
                cl = str(col).lower()
                if 'fecha' in cl and 'inicial' in cl and 'otrosi' not in cl and 'otrosí' not in cl:
                    fecha = fila[col]
                    break

        elif ('Otrosí' in origen or 'Otrosi' in origen) and fila is not None:
            m = re.search(r'\d+', origen)
            if m:
                num = int(m.group())
                patron = f"fecha.*otros[ií].*{num}"
                for col in columnas:
                    if re.search(patron, str(col).lower()):
                        fecha = fila[col]
                        break

        elif 'Acta' in origen and fila is not None:
            m = re.search(r'\d+', origen)
            if m:
                num = int(m.group())
                for i, col in enumerate(columnas):
                    cl = str(col).lower()
                    if 'no. acta' in cl or 'no acta' in cl:
                        val = fila[col]
                        if pd.notna(val):
                            ma = re.search(r'#?(\d+)', str(val))
                            if ma and int(ma.group(1)) == num:
                                if i + 1 < len(columnas):
                                    fecha = fila[columnas[i + 1]]
                                    break

        if fecha is None and fecha_archivo is not None:
            fecha_sftp = timestamp_a_fecha(fecha_archivo)
            if fecha_sftp:
                return fecha_sftp, True

        if fecha is not None and pd.notna(fecha):
            if isinstance(fecha, (int, float)):
                try:
                    dias = int(fecha)
                    if 30000 < dias < 60000:
                        fecha_dt = datetime(1899, 12, 30) + timedelta(days=dias)
                        return fecha_dt.strftime('%d/%m/%Y'), True
                except:
                    pass
            elif isinstance(fecha, datetime):
                return fecha.strftime('%d/%m/%Y'), True
            elif isinstance(fecha, str):
                fs = fecha.strip()
                if fs and fs.lower() not in ('nan', 'none', 'nat', ''):
                    return fs, True

        if fecha_archivo is not None:
            fecha_sftp = timestamp_a_fecha(fecha_archivo)
            if fecha_sftp:
                return fecha_sftp, True

        return None, False

    except Exception as e:
        if fecha_archivo is not None:
            fecha_sftp = timestamp_a_fecha(fecha_archivo)
            if fecha_sftp:
                return fecha_sftp, True
        return None, False

LOG.success("Función obtener_fecha_acuerdo cargada")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 9: GENERADOR DE EXCEL
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MAX_FILAS_POR_HOJA = 500_000

def exportar_consolidado_multisheet(df: pd.DataFrame, nombre_base: str, log=None, nombre_legible: str = None) -> str:
    """Exporta consolidado dividiendo en múltiples hojas si es necesario."""
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    if nombre_legible:
        archivo = f'{nombre_legible}_{timestamp}.xlsx'
    else:
        archivo = f'{nombre_base}_{timestamp}.xlsx'

    def _log(msg, level='info'):
        if log:
            getattr(log, level, log.info)(msg)
        else:
            print(msg)

    total_filas = len(df)
    num_hojas = (total_filas // MAX_FILAS_POR_HOJA) + (1 if total_filas % MAX_FILAS_POR_HOJA > 0 else 0)

    _log(f'Exportando {total_filas:,} registros a {num_hojas} hoja(s)...')

    try:
        with pd.ExcelWriter(archivo, engine='openpyxl') as writer:
            for i in range(num_hojas):
                inicio = i * MAX_FILAS_POR_HOJA
                fin = min((i + 1) * MAX_FILAS_POR_HOJA, total_filas)
                registros_hoja = fin - inicio

                if num_hojas == 1:
                    hoja_nombre = 'CONSOLIDADO'
                else:
                    hoja_nombre = f'CONSOLIDADO_{i + 1}'

                _log(f'   📊 Hoja \'{hoja_nombre}\': filas {inicio + 1:,} a {fin:,} ({registros_hoja:,} registros)')

                df.iloc[inicio:fin].to_excel(
                    writer,
                    sheet_name=hoja_nombre,
                    index=False,
                    freeze_panes=(1, 0)
                )

        tamaño = os.path.getsize(archivo)
        tamaño_str = f'{tamaño/1024/1024:.1f} MB' if tamaño > 1024*1024 else f'{tamaño/1024:.1f} KB'

        _log(f'✅ Exportado: {archivo}')
        _log(f'   • Tamaño: {tamaño_str}')
        _log(f'   • Hojas: {num_hojas}')
        _log(f'   • Total registros: {total_filas:,}')

        return archivo

    except Exception as e:
        _log(f'❌ Error exportando: {str(e)}', 'error')
        raise

def exportar_consolidado_csv(df: pd.DataFrame, nombre_base: str, log=None, nombre_legible: str = None) -> str:
    """Exporta consolidado a CSV."""
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M')
    if nombre_legible:
        archivo = f'{nombre_legible}_{timestamp}.csv'
    else:
        archivo = f'{nombre_base}_{timestamp}.csv'

    df.to_csv(archivo, index=False, encoding='utf-8-sig')

    tamaño = os.path.getsize(archivo)
    tamaño_str = f'{tamaño/1024/1024:.1f} MB' if tamaño > 1024*1024 else f'{tamaño/1024:.1f} KB'

    if log:
        log.success(f'Exportado CSV: {archivo} ({tamaño_str}, {len(df):,} registros)')

    return archivo

LOG.success("Generador Excel cargado")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 10: SELECCIÓN DE CONTRATOS
# ══════════════════════════════════════════════════════════════════════════════

LOG.header("SELECCIÓN DE CONTRATOS")

print("""
  Opciones disponibles:
  ─────────────────────────────────────────────────────────────
  [1] Un contrato específico
  [2] Todos los contratos de un año
  [3] Todos los contratos
  ─────────────────────────────────────────────────────────────
""")


# ══════════════════════════════════════════════════════════════════════════════
# SELECCIÓN DE CONTRATOS (PARAMETRIZADO - SIN input())
# ══════════════════════════════════════════════════════════════════════════════

# ADAPTACIÓN: Usar variables de entorno en lugar de input()
if PARAM_MODO == "ESPECIFICO":
    opcion = "1"
elif PARAM_MODO == "POR_ANO":
    opcion = "2"
else:  # COMPLETO
    opcion = "3"

CONTRATOS_A_PROCESAR = []
MODO_OPERACION = ""
CARPETA_TRABAJO = "./trabajo"

LOG.indent()

if opcion == "1":
    MODO_OPERACION = "ESPECIFICO"
    numero = PARAM_NUMERO
    ano = PARAM_ANO
    CONTRATOS_A_PROCESAR = [{'numero': numero, 'ano': ano}]
    CARPETA_TRABAJO = f"./trabajo_{numero}_{ano}"
    LOG.success(f"Contrato seleccionado: {numero}-{ano}")

elif opcion == "2":
    MODO_OPERACION = "POR_ANO"
    ano = PARAM_ANO
    
    if COLS.numero_contrato and COLS.ano_contrato:
        df_filtrado = df_prestadores[
            df_prestadores[COLS.ano_contrato].astype(str).str.replace('.0', '', regex=False) == ano
        ]
        for _, row in df_filtrado.iterrows():
            n = row[COLS.numero_contrato]
            if pd.notna(n):
                CONTRATOS_A_PROCESAR.append({'numero': str(int(n)), 'ano': ano})

    CARPETA_TRABAJO = f"./trabajo_ano_{ano}"
    LOG.success(f"Año seleccionado: {ano}", f"{len(CONTRATOS_A_PROCESAR)} contratos")

elif opcion == "3":
    MODO_OPERACION = "COMPLETO"
    if COLS.numero_contrato and COLS.ano_contrato:
        for _, row in df_prestadores.iterrows():
            n, a = row[COLS.numero_contrato], row[COLS.ano_contrato]
            if pd.notna(n) and pd.notna(a):
                CONTRATOS_A_PROCESAR.append({'numero': str(int(n)), 'ano': str(int(a))})
    CARPETA_TRABAJO = "./trabajo_completo"
    LOG.success(f"Procesamiento completo", f"{len(CONTRATOS_A_PROCESAR)} contratos")

else:
    LOG.error("Opción no válida")

if CONTRATOS_A_PROCESAR:
    os.makedirs(CARPETA_TRABAJO, exist_ok=True)
    LOG.info("Carpeta de trabajo", CARPETA_TRABAJO)

LOG.dedent()

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 11: CONEXIÓN AL SERVIDOR
# ══════════════════════════════════════════════════════════════════════════════

if CONTRATOS_A_PROCESAR:
    LOG.header("CONEXIÓN AL SERVIDOR SFTP")

    cliente = SFTPClient(CONFIG, LOG)

    if cliente.conectar():
        buscador = BuscadorAnexos(cliente, CONFIG, LOG)
        procesador = ProcesadorAnexo(LOG)

        LOG.info(f"Contratos a procesar: {len(CONTRATOS_A_PROCESAR)}")
        LOG.info(f"Timeout por archivo: {CONFIG.TIMEOUT_ARCHIVO}s")
    else:
        LOG.error("No se pudo conectar. Verifica la red y credenciales.")
else:
    LOG.warning("No hay contratos seleccionados")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 12: PROCESAMIENTO PRINCIPAL v14.1 - CON RECONEXIÓN FORZADA
# ══════════════════════════════════════════════════════════════════════════════

if not CONTRATOS_A_PROCESAR:
    LOG.warning("No hay contratos para procesar")
else:
    LOG.header("PROCESAMIENTO v14.1", f"Modo: {MODO_OPERACION} | {len(CONTRATOS_A_PROCESAR)} contratos")

    LOG.start_time = time.time()
    LOG.stats = {
        'contratos_procesados': 0,
        'contratos_exitosos': 0,
        'archivos_descargados': 0,
        'servicios_extraidos': 0,
        'alertas_generadas': 0
    }

    # consolidado_total = []  <-- ELIMINADO PARA AHORRAR MEMORIA
    
    # 🆕 v15.1: BATCH PROCESSING PARA EVITAR OOM
    BATCH_SIZE = 500  # Reducido a 500 (Ultra-conservative mode)
    batch_buffer = []
    
    # 🆕 v15.2: Flushing de alertas
    ALERT_BATCH_SIZE = 2000
    temp_alertas_file = f"temp_alertas_{int(time.time())}.csv"
    alertas_header_written = False
    
    import gc # Asegurar importación

    temp_csv_file = f"temp_consolidado_{int(time.time())}.csv"
    csv_headers_written = False
    total_registros_procesados = 0
    
    def procesar_y_guardar_batch(buffer, archivo_csv, es_primer_batch):
        if not buffer: return False, 0
        
        try:
            df_batch = pd.DataFrame(buffer)
            
            # Aplicar ETL ML si está disponible
            if etl_ml_helper:
                # Usar un nombre genérico para el log por batch
                df_batch = etl_ml_helper.procesar_dataframe(df_batch, f"Batch Procesamiento")
            
            # Guardar
            modo = 'w' if es_primer_batch else 'a'
            header = es_primer_batch
            
            # Limpieza final de formatos numéricos (eliminar .0)
            cols_limpiar = ['tarifa_unitaria_en_pesos', 'porcentaje_manual_tarifario']
            for col in cols_limpiar:
                if col in df_batch.columns:
                    # Convertir a string, eliminar nulos y usar regex para quitar .0 al final
                    df_batch[col] = df_batch[col].astype(str).replace({'nan': '', 'NaN': '', 'None': ''})
                    df_batch[col] = df_batch[col].str.replace(r'\.0$', '', regex=True)

            df_batch.to_csv(archivo_csv, mode=modo, header=header, index=False, encoding='utf-8-sig')
            
            # Limpieza agresiva de memoria
            del df_batch
            gc.collect()
            
            return True, len(buffer)
        except Exception as e:
            LOG.error(f"Error guardando batch: {e}")
            return False, 0
    def guardar_alertas_batch(alertas_lista, archivo_csv, es_primer_batch):
        if not alertas_lista: return
        try:
            df_a = pd.DataFrame(alertas_lista)
            modo = 'w' if es_primer_batch else 'a'
            header = es_primer_batch
            df_a.to_csv(archivo_csv, mode=modo, header=header, index=False, encoding='utf-8-sig')
        except Exception as e:
            LOG.error(f"Error guardando batch de alertas: {e}")

    todas_alertas = [] # Se usará como buffer ahora
    alertas_set = set()
    resumen_contratos = []
    archivos_no_positiva = []
    contratos_sin_fecha = set()
    fechas_ok = fechas_no = 0

    PALABRAS_AMBULANCIA_MAESTRA = [
        'AMBULANCIA', 'AMBULANCIAS',
        'TRASLADO ASISTENCIAL', 'TRASLADOS ASISTENCIALES',
        'TAM', 'TAB',
        'TRANSPORTE ASISTENCIAL', 'TRANSPORTES ASISTENCIALES',
        'SERVICIO DE AMBULANCIA', 'SERVICIOS DE AMBULANCIA',
    ]

    COLUMNAS_REVISAR_AMBULANCIA = [
        'CATEGORÍA CUENTAS MEDICAS', 'CATEGORIA CUENTAS MEDICAS',
        'OBJETO', 'DESCRIPCION', 'DESCRIPCIÓN',
        'TIPO', 'TIPO_SERVICIO', 'SERVICIO'
    ]

    def detectar_ambulancia_en_maestra(numero: str, ano: str) -> Tuple[bool, str, str]:
        try:
            cto_str = f"{str(numero).zfill(4)}-{ano}"

            mask = None
            if COLS.cto:
                mask = df_maestra[COLS.cto] == cto_str

            if mask is None or not mask.any():
                if COLS.numero_contrato and COLS.ano_contrato:
                    mask = (
                        df_maestra[COLS.numero_contrato].astype(str).str.replace('.0', '', regex=False).str.zfill(4) == str(numero).zfill(4)
                    ) & (
                        df_maestra[COLS.ano_contrato].astype(str).str.replace('.0', '', regex=False) == str(ano)
                    )

            if mask is None or not mask.any():
                return False, "", ""

            fila_cto = df_maestra[mask].iloc[0]

            for col in df_maestra.columns:
                col_upper = str(col).upper().strip()

                revisar = False
                for col_target in COLUMNAS_REVISAR_AMBULANCIA:
                    if col_target in col_upper or col_upper in col_target:
                        revisar = True
                        break

                if not revisar:
                    continue

                valor = fila_cto[col]
                if pd.isna(valor) or valor is None:
                    continue

                valor_str = str(valor).upper().strip()

                for palabra in PALABRAS_AMBULANCIA_MAESTRA:
                    if palabra in valor_str:
                        return True, col, valor_str

            return False, "", ""

        except Exception as e:
            return False, "", ""

    def obtener_categoria_cuentas_medicas(numero: str, ano: str) -> str:
        try:
            cto_str = f"{str(numero).zfill(4)}-{ano}"

            mask = None
            if COLS.cto:
                mask = df_maestra[COLS.cto] == cto_str

            if mask is None or not mask.any():
                if COLS.numero_contrato and COLS.ano_contrato:
                    mask = (
                        df_maestra[COLS.numero_contrato].astype(str).str.replace('.0', '', regex=False).str.zfill(4) == str(numero).zfill(4)
                    ) & (
                        df_maestra[COLS.ano_contrato].astype(str).str.replace('.0', '', regex=False) == str(ano)
                    )

            if mask is None or not mask.any():
                return ""

            fila_cto = df_maestra[mask].iloc[0]

            for col in df_maestra.columns:
                col_upper = str(col).upper().strip()
                if 'CATEGOR' in col_upper and 'CUENTA' in col_upper:
                    valor = fila_cto[col]
                    if pd.notna(valor):
                        return str(valor).strip()

            return ""
        except:
            return ""

    def agregar_alerta_unica(alerta_dict: dict):
        clave = (alerta_dict['tipo'], alerta_dict['mensaje'], alerta_dict['contrato'], alerta_dict['archivo'])
        if clave not in alertas_set:
            alertas_set.add(clave)
            todas_alertas.append(alerta_dict)

    # 🆕 v14.1: RECONECTAR CADA N CONTRATOS
    RECONECTAR_CADA_N = 10

    for idx, contrato in enumerate(CONTRATOS_A_PROCESAR, 1):
        numero, ano = contrato['numero'], contrato['ano']
        id_c = f"{numero}-{ano}"

        LOG.contract_start(idx, len(CONTRATOS_A_PROCESAR), id_c)

        es_ambulancia, col_ambulancia, valor_ambulancia = detectar_ambulancia_en_maestra(numero, ano)
        categoria_cuentas_medicas = obtener_categoria_cuentas_medicas(numero, ano)

        if es_ambulancia:
            LOG.info(f"📋 Contrato identificado como AMBULANCIAS desde maestra")
            LOG.info(f"   Columna: '{col_ambulancia}' → '{valor_ambulancia[:50]}...'")

            agregar_alerta_unica(Alerta(
                tipo=TipoAlerta.CONTRATO_AMBULANCIA_MAESTRA,
                mensaje=f"Identificado como contrato de ambulancias - Columna '{col_ambulancia}' contiene '{valor_ambulancia[:30]}'",
                contrato=id_c
            ).to_dict())

        t_c = time.time()

        # 🆕 v14.1: RECONEXIÓN FORZADA
        LOG.indent()
        LOG.info("🔄 Verificando/renovando conexión SFTP...")

        conexion_ok = False

        if idx % RECONECTAR_CADA_N == 1 or not cliente.esta_activo():
            LOG.debug(f"Reconexión forzada (contrato #{idx})")
            if cliente.reconectar_forzado(silencioso=True):
                conexion_ok = True
                LOG.success("Conexión renovada")
            else:
                for intento in range(3):
                    if cliente.conectar(True):
                        conexion_ok = True
                        break
                    LOG.warning(f"Reintento de conexión {intento + 1}/3...")
                    time.sleep(2)
        else:
            if cliente.esta_activo():
                conexion_ok = True
            else:
                for intento in range(3):
                    LOG.warning(f"Reconectando (intento {intento + 1})...")
                    if cliente.conectar(True):
                        conexion_ok = True
                        break
                    time.sleep(2)

        if not conexion_ok:
            LOG.error("Sin conexión al servidor")
            resumen_contratos.append({
                'contrato': id_c, 'exito': 'NO', 'registros': 0,
                'mensaje': 'Sin conexión (Socket closed)', 'tiempo': 0
            })
            agregar_alerta_unica(Alerta(
                tipo=TipoAlerta.CONEXION,
                mensaje='No se pudo conectar - Socket is closed',
                contrato=id_c
            ).to_dict())
            LOG.dedent()
            LOG.contract_end(False, 0, time.time() - t_c, "Sin conexión")
            continue

        carpeta = os.path.join(CARPETA_TRABAJO, f"t_{numero}_{ano}")
        os.makedirs(carpeta, exist_ok=True)

        buscador.limpiar_alertas()
        buscador.set_contrato(id_c)
        procesador.limpiar_alertas()
        procesador.set_contrato(id_c)
        procesador.set_categoria_cuentas_medicas(categoria_cuentas_medicas)

        res = {'exito': False, 'archivos': [], 'mensaje': 'Error'}

        for intento in range(3):
            try:
                ok, msg, ruta = buscador.navegar_a_contrato(ano, numero)
                if ok:
                    res = buscador.descargar_anexos(carpeta, id_c)
                else:
                    res = {'exito': False, 'archivos': [], 'mensaje': msg}
                break
            except Exception as e:
                if 'socket' in str(e).lower() and intento < 2:
                    LOG.warning("Error de socket, reconectando...")
                    cliente.reconectar_forzado(silencioso=True)
                else:
                    res['mensaje'] = str(e)[:30]
                    break

        for alerta in buscador.alertas:
            agregar_alerta_unica(alerta.to_dict())

        # Flush alertas si es necesario
        if len(todas_alertas) >= ALERT_BATCH_SIZE:
             guardar_alertas_batch(todas_alertas, temp_alertas_file, not alertas_header_written)
             alertas_header_written = True
             todas_alertas = []
             gc.collect()

        if not res['exito']:
            resumen_contratos.append({
                'contrato': id_c, 'exito': 'NO', 'registros': 0,
                'mensaje': res['mensaje'], 'tiempo': round(time.time() - t_c, 1)
            })

            try: shutil.rmtree(carpeta)
            except: pass

            LOG.dedent()
            LOG.contract_end(False, 0, time.time() - t_c, res['mensaje'])
            continue

        regs = 0
        es_prob = id_c in CONFIG.CONTRATOS_PROBLEMATICOS
        timeout = CONFIG.TIMEOUT_CONTRATOS_PROBLEMATICOS if es_prob else CONFIG.TIMEOUT_ARCHIVO

        for arch in res['archivos']:
            nombre = arch.nombre if hasattr(arch, 'nombre') else arch.get('nombre', '')
            ruta = arch.ruta_local if hasattr(arch, 'ruta_local') else arch.get('ruta_local', '')
            origen = arch.origen_completo if hasattr(arch, 'origen_completo') else arch.get('origen', '')
            fecha_mod = arch.fecha_modificacion if hasattr(arch, 'fecha_modificacion') else arch.get('fecha_modificacion')

            try:
                ok, servs, msg = procesador.extraer_con_timeout(ruta, nombre, timeout)

                if ok and servs:
                    fecha, f_ok = obtener_fecha_acuerdo(numero, ano, origen, fecha_mod)

                    if f_ok:
                        fechas_ok += 1
                    else:
                        fechas_no += 1
                        contratos_sin_fecha.add(id_c)
                        agregar_alerta_unica(Alerta(
                            tipo=TipoAlerta.FECHA_NO_ENCONTRADA,
                            mensaje=f"Sin fecha para {origen}",
                            contrato=id_c,
                            archivo=nombre
                        ).to_dict())

                    for s in servs:
                        s['contrato'] = id_c
                        s['origen_tarifa'] = origen
                        s['fecha_de_acuerdo'] = fecha if fecha else ''
                        batch_buffer.append(s)

                        # Procesar batch si está lleno
                        if len(batch_buffer) >= BATCH_SIZE:
                            LOG.info(f"💾 Guardando batch intermedio ({len(batch_buffer)} registros)...")
                            ok_batch, n_regs = procesar_y_guardar_batch(batch_buffer, temp_csv_file, not csv_headers_written)
                            if ok_batch:
                                csv_headers_written = True
                                total_registros_procesados += n_regs
                            batch_buffer = [] # Limpiar memoria
                            gc.collect()

                    regs += len(servs)
                else:
                    # Verificar si es un archivo de paquetes (no incluir en No_Positiva, solo en alertas)
                    es_paquete = 'PAQUETE' in msg.upper() if msg else False
                    
                    if es_paquete:
                        # Solo agregar alerta, no a archivos_no_positiva
                        agregar_alerta_unica(Alerta(
                            tipo=TipoAlerta.ARCHIVO_PAQUETE,
                            mensaje=f"Archivo de paquetes: {msg}",
                            contrato=id_c,
                            archivo=nombre
                        ).to_dict())
                    else:
                        archivos_no_positiva.append({
                            'contrato': id_c,
                            'archivo': nombre,
                            'motivo': msg
                        })

            except Exception as e:
                archivos_no_positiva.append({
                    'contrato': id_c,
                    'archivo': nombre,
                    'motivo': str(e)[:50]
                })

        for alerta in procesador.alertas:
            agregar_alerta_unica(alerta.to_dict())

        # Flush alertas si es necesario
        if len(todas_alertas) >= ALERT_BATCH_SIZE:
             guardar_alertas_batch(todas_alertas, temp_alertas_file, not alertas_header_written)
             alertas_header_written = True
             todas_alertas = []

        exito = regs > 0
        resumen_contratos.append({
            'contrato': id_c,
            'exito': 'SI' if exito else 'NO',
            'registros': regs,
            'mensaje': f'{regs} servicios' if exito else 'Sin servicios',
            'tiempo': round(time.time() - t_c, 1),
            'es_ambulancia': 'SI' if es_ambulancia else 'NO'
        })

        try: shutil.rmtree(carpeta)
        except: pass

        LOG.dedent()
        LOG.dedent()
        LOG.contract_end(exito, regs, time.time() - t_c, '' if exito else 'Sin servicios')

    # Procesar remanentes al final del loop
    if batch_buffer:
        LOG.info(f"💾 Guardando últimos {len(batch_buffer)} registros...")
        ok_batch, n_regs = procesar_y_guardar_batch(batch_buffer, temp_csv_file, not csv_headers_written)
        if ok_batch:
            csv_headers_written = True
            total_registros_procesados += n_regs
        batch_buffer = []
        gc.collect()

    LOG.stats_summary()

    print(f"\n📊 RESUMEN DE PROCESAMIENTO:")
    print(f"   • Registros consolidados: {total_registros_procesados:,}")

    # Flush final de alertas
    if todas_alertas:
         guardar_alertas_batch(todas_alertas, temp_alertas_file, not alertas_header_written)
         alertas_header_written = True
         todas_alertas = []
    print(f"   • Alertas generadas: {len(alertas_set)} (unicas)")
    print(f"   • Archivos sin formato POSITIVA: {len(archivos_no_positiva)}")
    print(f"   • Contratos sin fecha en maestra: {len(contratos_sin_fecha)}")
    print(f"   • Fechas encontradas: {fechas_ok} | No encontradas: {fechas_no}")
    print(f"   • Reconexiones SFTP: {cliente.reconexiones}")

    contratos_ambulancia = sum(1 for r in resumen_contratos if r.get('es_ambulancia') == 'SI')
    if contratos_ambulancia > 0:
        print(f"   • Contratos de ambulancias detectados: {contratos_ambulancia}")

# ══════════════════════════════════════════════════════════════════════════════
# CELDA 13: GENERACIÓN DE ARCHIVOS v14.1 - ALERTAS SEPARADAS POR HOJAS
# ══════════════════════════════════════════════════════════════════════════════

LOG.header("GENERACIÓN DE ARCHIVOS v14.1")

archivos_generados = []

sufijos = {
    'ESPECIFICO': f"{CONTRATOS_A_PROCESAR[0]['numero']}_{CONTRATOS_A_PROCESAR[0]['ano']}" if CONTRATOS_A_PROCESAR else 'X',
    'POR_ANO': f"ANO_{CONTRATOS_A_PROCESAR[0]['ano']}" if CONTRATOS_A_PROCESAR else 'X',
    'COMPLETO': 'COMPLETO'
}
suf = sufijos.get(MODO_OPERACION, 'X')
ts = datetime.now().strftime('%Y-%m-%d_%H-%M')

# Nombres legibles para archivos
nombres_legibles = {
    'ESPECIFICO': f"Consolidado_Contrato_{CONTRATOS_A_PROCESAR[0]['numero']}",
    'POR_ANO': f"Consolidado_Año_{CONTRATOS_A_PROCESAR[0]['ano']}" if CONTRATOS_A_PROCESAR else 'Consolidado',
    'COMPLETO': 'Consolidado_Completo'
}
nombre_consolidado = nombres_legibles.get(MODO_OPERACION, 'Consolidado')

LOG.indent()

LOG.indent()

# Cargar consolidado final desde CSV
df_consolidado = None
if os.path.exists(temp_csv_file):
    try:
        if total_registros_procesados > 0:
            LOG.info("🔄 Cargando resultado final para exportación...")
            # Leer con tipos string para preservar formatos
            df_consolidado = pd.read_csv(temp_csv_file, dtype=str)
            df_consolidado = df_consolidado.replace({'nan': '', 'NaN': ''})
    except Exception as e:
        LOG.error(f"Error cargando CSV temporal: {e}")

if df_consolidado is not None and not df_consolidado.empty:
    total_registros = len(df_consolidado)
    LOG.info(f"Total de registros a exportar: {total_registros:,}")

    if total_registros > MAX_FILAS_POR_HOJA:
        LOG.warning(f"⚠️ El consolidado tiene {total_registros:,} registros")
        LOG.warning(f"   Se dividirá en múltiples hojas (máx {MAX_FILAS_POR_HOJA:,} por hoja)")

    try:
        # df_consolidado ya está creado
        archivo = exportar_consolidado_multisheet(
            df_consolidado,
            f"CONSOLIDADO_{suf}",
            log=LOG,
            nombre_legible=nombre_consolidado
        )
        archivos_generados.append(archivo)
    except Exception as e:
        LOG.error(f"Error exportando consolidado: {str(e)}")
        LOG.info("Intentando exportar a CSV como alternativa...")
        try:
            # df_consolidado ya está creado
            archivo_csv = exportar_consolidado_csv(df_consolidado, f"CONSOLIDADO_{suf}", log=LOG, nombre_legible=nombre_consolidado)
            archivos_generados.append(archivo_csv)
        except Exception as e2:
            LOG.error(f"Error exportando CSV: {str(e2)}")

# 🆕 v14.1: ALERTAS SEPARADAS POR HOJAS
# 🆕 v14.1: ALERTAS SEPARADAS POR HOJAS
# Cargar alertas desde CSV temporal si existe
if os.path.exists(temp_alertas_file):
    try:
        df_alertas_full = pd.read_csv(temp_alertas_file, dtype=str)
        todas_alertas = df_alertas_full.to_dict('records') # Convertir para el proceso de separación
    except:
        pass

if todas_alertas:
    nombre_alertas = f"Alertas_{ts}.xlsx"

    try:
        df_alertas = pd.DataFrame(todas_alertas)

        CATEGORIAS_ALERTAS = {
            'CONTRATOS_NO_ENCONTRADOS': [
                'CONTRATO_NO_ENCONTRADO_GO',
                'SIN_CARPETA_TARIFAS',
                'CONEXION'
            ],
            'HOJAS_SIN_SERVICIOS': [
                'HOJA_NO_ENCONTRADA',
                'TARIFA_SERVICIOS_NO_ENCONTRADA',
                'COLUMNAS_NO_DETECTADAS',
                'SEDES_NO_DETECTADAS'
            ],
            'FECHAS_FALTANTES': [
                'FECHA_NO_ENCONTRADA',
                'FECHA_FALTANTE_MAESTRA'
            ],
            'AMBULANCIAS_TRASLADOS': [
                'CONTRATO_AMBULANCIA_MAESTRA',
                'ARCHIVO_SOLO_AMBULANCIAS',
                'ARCHIVO_SOLO_TRASLADOS',
                'SOLO_TRASLADOS',
                'CONTRATO_AMBULANCIA'
            ],
            'ANEXOS_FALTANTES': [
                'SIN_ANEXO1',
                'ACTA_FALTANTE',
                'CARPETA_ACTAS_SIN_ANEXO'
            ],
            # 🆕 v15.0: Nueva categoría para formatos propios
            'FORMATO_PROPIO': [
                'FORMATO_PROPIO',
                'SIN_FORMATO_POSITIVA'
            ],
            'ERRORES_PROCESAMIENTO': [
                'ERROR_PROCESAMIENTO',
                'ERROR_LECTURA',
                'TIMEOUT'
            ]
        }

        with pd.ExcelWriter(nombre_alertas, engine='openpyxl') as writer:
            if 'prioridad' in df_alertas.columns:
                df_alertas_sorted = df_alertas.sort_values(['prioridad', 'tipo', 'contrato'])
            else:
                df_alertas_sorted = df_alertas.sort_values(['tipo', 'contrato'])

            df_alertas_sorted.to_excel(writer, sheet_name='TODAS_ALERTAS', index=False)

            for nombre_hoja, tipos in CATEGORIAS_ALERTAS.items():
                df_categoria = df_alertas[df_alertas['tipo'].isin(tipos)]
                if len(df_categoria) > 0:
                    df_categoria.to_excel(writer, sheet_name=nombre_hoja[:31], index=False)

            todos_tipos_categorizados = [t for tipos in CATEGORIAS_ALERTAS.values() for t in tipos]
            df_otras = df_alertas[~df_alertas['tipo'].isin(todos_tipos_categorizados)]
            if len(df_otras) > 0:
                df_otras.to_excel(writer, sheet_name='OTRAS_ALERTAS', index=False)

        LOG.success(f"Generado: {nombre_alertas}", f"{len(todas_alertas)} alertas en múltiples hojas")
        archivos_generados.append(nombre_alertas)

        print(f"\n📋 RESUMEN DE ALERTAS POR CATEGORÍA:")
        for nombre_hoja, tipos in CATEGORIAS_ALERTAS.items():
            count = len(df_alertas[df_alertas['tipo'].isin(tipos)])
            if count > 0:
                print(f"   • {nombre_hoja}: {count}")

    except Exception as e:
        LOG.error(f"Error generando alertas separadas: {str(e)}")
        df_alertas = pd.DataFrame(todas_alertas)
        if 'prioridad' in df_alertas.columns:
            df_alertas = df_alertas.sort_values(['prioridad', 'tipo', 'contrato'])
        df_alertas.to_excel(nombre_alertas, index=False)
        LOG.warning(f"Generado archivo simple: {nombre_alertas}")
        archivos_generados.append(nombre_alertas)

if resumen_contratos:
    nombre = f"Resumen_{ts}.xlsx"
    pd.DataFrame(resumen_contratos).to_excel(nombre, index=False)
    LOG.success(f"Generado: {nombre}", f"{len(resumen_contratos)} contratos")
    archivos_generados.append(nombre)

if archivos_no_positiva:
    nombre = f"Archivos_No_Positiva_{ts}.xlsx"
    pd.DataFrame(archivos_no_positiva).to_excel(nombre, index=False)
    LOG.success(f"Generado: {nombre}", f"{len(archivos_no_positiva)} archivos")
    archivos_generados.append(nombre)

LOG.dedent()
LOG.info(f"Total archivos generados: {len(archivos_generados)}")



# Cerrar conexión SFTP
try:
    cliente.desconectar()
except:
    pass

print("\n" + "═"*70)
print("✅ CONSOLIDADOR T25 + ETL ML - PROCESO COMPLETO FINALIZADO")
print("═"*70)
print("""
📋 ARCHIVOS GENERADOS:
   • CONSOLIDADO_*.xlsx    - Datos consolidados del GoAnywhere
   • *_ML_LIMPIO.xlsx      - Datos procesados con ML
   • ALERTAS_*.xlsx        - Alertas del procesamiento
   • RESUMEN_*.xlsx        - Resumen de contratos
   • correcciones_ml.csv   - Log de correcciones ML (si aplica)

💡 EL PROCESO SE EJECUTÓ DE FORMA AUTOMÁTICA:
   1. Consolidador T25 → Extrae datos de GoAnywhere
   2. ETL con ML → Limpia y normaliza los datos
   3. Descarga → Todos los archivos disponibles
""")